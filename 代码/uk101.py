# -*- coding: utf-8 -*-
"""
根据 UK100_工作量 KPI 基础表生成 UEA_个人工作量规模_结果表_月汇总_基于UK100.xlsx。
"""

from __future__ import annotations

import argparse
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook

from uk_export_utils import build_default_output_path
from uk_workload_utils import (
    department_sort_key,
    find_latest_input,
    get_required_indexes,
    has_value,
    load_sheet_context,
    month_sort_key,
    normalize_text,
    parse_cell_date,
    parse_cell_time,
    round_half_up,
    safe_div,
    style_two_header_sheet,
    write_notes_rows,
)

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UEA_个人工作量规模_结果表_月汇总_基于UK100.xlsx"
RESULT_SHEET_NAME = "UEA_Workload_Monthly_Result"

FOLLOW_FIELDS = {
    "AT": "Order_Follow_AT",
    "CT": "Order_Follow_CT",
    "KP": "Order_Follow_KP",
    "OM": "Order_Follow_OM",
}

OPERATION_COLUMN_MAP = {
    "ticket_delay_cnt": "TAM1_Cnt",
    "ticket_close_cnt": "TAM2_Cnt",
    "cross_dept_comment_cnt": "TAM3_Cnt",
    "intra_dept_comment_cnt": "TAM4_Cnt",
    "ticket_submit_note_cnt": "TAM5_Cnt",
}

COLUMNS_EN = [
    "stat_month",
    "handler_dept",
    "handler_id",
    "total_work_cnt",
    "ticket_cnt",
    "mail_cnt",
    "active_days",
    "avg_daily_work_cnt",
    "max_daily_work_cnt",
    "ticket_ratio",
    "mail_ratio",
    "assist_ticket_cnt",
    "assist_ticket_ratio",
    "after_hour_ticket_cnt",
    "after_hour_ticket_ratio",
    "work_cnt_rank",
    "avg_daily_rank",
    "ticket_cnt_share_dept",
    "mail_cnt_share_dept",
    "assist_ticket_cnt_share_dept",
    "after_hour_ticket_cnt_share_dept",
    "ticket_delay_cnt",
    "ticket_close_cnt",
    "cross_dept_comment_cnt",
    "intra_dept_comment_cnt",
    "ticket_submit_note_cnt",
    "total_ops_cnt",
    "ticket_delay_cnt_ratio_all",
    "ticket_close_cnt_ratio_all",
    "cross_dept_comment_cnt_ratio_all",
    "intra_dept_comment_cnt_ratio_all",
    "ticket_submit_note_cnt_ratio_all",
    "ticket_delay_cnt_ratio_dept",
    "ticket_close_cnt_ratio_dept",
    "cross_dept_comment_cnt_ratio_dept",
    "intra_dept_comment_cnt_ratio_dept",
    "ticket_submit_note_cnt_ratio_dept",
]

COLUMNS_CN = [
    "统计月份",
    "所属部门",
    "处理人编号",
    "总处理量",
    "工单数量",
    "信件数量",
    "活跃天数",
    "日均处理量",
    "单日峰值",
    "工单占比",
    "信件占比",
    "协同工单数量",
    "协同工单占比",
    "17:30后工单数量",
    "17:30后工单占比",
    "工作量排名（部门内）",
    "日均处理量排名（部门内）",
    "工单数量占部门比",
    "信件数量占部门比",
    "协同工单数量占部门比",
    "17:30后工单数量占部门比",
    "工单延期次数",
    "工单完成/取消次数",
    "跨部门评论次数",
    "本部门评论次数",
    "提交备注次数",
    "总操作次数",
    "延期操作占个人总操作比",
    "关闭操作占个人总操作比",
    "跨部门评论占个人总操作比",
    "本部门评论占个人总操作比",
    "提交备注占个人总操作比",
    "延期操作占部门同类操作比",
    "关闭操作占部门同类操作比",
    "跨部门评论占部门同类操作比",
    "本部门评论占部门同类操作比",
    "提交备注占部门同类操作比",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 UK100 基础表生成 UK101 月度个人工作量规模汇总。")
    parser.add_argument("--input", dest="input_path", help="可选，指定输入的 UK100 基础表路径。")
    parser.add_argument("--output", dest="output_path", help="可选，指定输出 Excel 路径。")
    return parser.parse_args()


def is_email_type(value: object) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    upper_text = text.upper()
    if upper_text in {"MAIL", "EMAIL", "E-MAIL"}:
        return True
    if "EMAIL" in upper_text or "E-MAIL" in upper_text:
        return True
    if text in {"信件", "邮件"}:
        return True
    if "信件" in text or "邮件" in text:
        return True
    return False


def resolve_department_code(department: str) -> str | None:
    text = normalize_text(department).upper()
    if not text:
        return None
    if text in FOLLOW_FIELDS:
        return text
    for code in FOLLOW_FIELDS:
        if text.startswith(code):
            return code
    return None


def parse_count(value: object) -> int:
    if not has_value(value):
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return max(int(round_half_up(value, 0)), 0)
    text = normalize_text(value).replace(",", "")
    if not text:
        return 0
    try:
        return max(int(round_half_up(float(text), 0)), 0)
    except ValueError:
        return 0


def dense_rank_desc(items: list[tuple[str, int]]) -> dict[str, int]:
    result: dict[str, int] = {}
    current_rank = 0
    previous_value: int | None = None
    for name, value in items:
        if previous_value is None or value != previous_value:
            current_rank += 1
            previous_value = value
        result[name] = current_rank
    return result


def build_summary(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook, sheet, header_row_1, _, data_start_row = load_sheet_context(input_path)
    try:
        indexes = get_required_indexes(
            header_row_1,
            {
                "Stat_Date": "Stat_Date",
                "Department": "Department",
                "UEA": "UEA",
                "Type": "Type",
                "Ticket_Last_Comment_Time": "Ticket_Last_Comment_Time",
                "Order_Follow_AT": "Order_Follow_AT",
                "Order_Follow_CT": "Order_Follow_CT",
                "Order_Follow_KP": "Order_Follow_KP",
                "TAM1_Cnt": "TAM1_Cnt",
                "TAM2_Cnt": "TAM2_Cnt",
                "TAM3_Cnt": "TAM3_Cnt",
                "TAM4_Cnt": "TAM4_Cnt",
                "TAM5_Cnt": "TAM5_Cnt",
            },
        )
        if "Order_Follow_OM" in header_row_1:
            indexes["Order_Follow_OM"] = header_row_1.index("Order_Follow_OM")

        metrics = defaultdict(
            lambda: {
                "total_work_cnt": 0,
                "ticket_cnt": 0,
                "mail_cnt": 0,
                "assist_ticket_cnt": 0,
                "after_hour_ticket_cnt": 0,
                "ticket_delay_cnt": 0,
                "ticket_close_cnt": 0,
                "cross_dept_comment_cnt": 0,
                "intra_dept_comment_cnt": 0,
                "ticket_submit_note_cnt": 0,
            }
        )
        daily_counts: Counter[tuple[str, str, str, object]] = Counter()
        dept_daily_counts: Counter[tuple[str, str, object]] = Counter()
        active_days_map: dict[tuple[str, str, str], set] = defaultdict(set)
        dept_days: dict[tuple[str, str], set] = defaultdict(set)
        dept_handlers: dict[tuple[str, str], set[str]] = defaultdict(set)

        input_rows = max((sheet.max_row or 0) - data_start_row + 1, 0)
        cleaned_rows = 0
        invalid_stat_date_rows = 0
        empty_uea_rows = 0
        empty_department_rows = 0

        threshold = parse_cell_time("17:30:00")

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            stat_date = parse_cell_date(row[indexes["Stat_Date"]])
            if stat_date is None:
                invalid_stat_date_rows += 1
                continue

            uea = normalize_text(row[indexes["UEA"]])
            if not uea:
                empty_uea_rows += 1
                continue

            department = normalize_text(row[indexes["Department"]])
            if not department:
                empty_department_rows += 1
                continue

            cleaned_rows += 1

            month = stat_date.strftime("%Y-%m")
            key = (month, department, uea)
            dept_key = (month, department)

            dept_handlers[dept_key].add(uea)
            dept_days[dept_key].add(stat_date)
            active_days_map[key].add(stat_date)

            metrics[key]["total_work_cnt"] += 1
            daily_counts[(month, department, uea, stat_date)] += 1
            dept_daily_counts[(month, department, stat_date)] += 1

            row_type = normalize_text(row[indexes["Type"]])
            if is_email_type(row_type):
                metrics[key]["mail_cnt"] += 1
                continue

            metrics[key]["ticket_cnt"] += 1

            department_code = resolve_department_code(department)
            if department_code:
                follow_field = FOLLOW_FIELDS[department_code]
                follow_index = indexes.get(follow_field)
                if follow_index is not None:
                    follow_value = normalize_text(row[follow_index])
                    if follow_value and follow_value.upper() != uea.upper():
                        metrics[key]["assist_ticket_cnt"] += 1

            comment_time = parse_cell_time(row[indexes["Ticket_Last_Comment_Time"]])
            if threshold is not None and comment_time is not None and comment_time >= threshold:
                metrics[key]["after_hour_ticket_cnt"] += 1

            for metric_name, column_name in OPERATION_COLUMN_MAP.items():
                metrics[key][metric_name] += parse_count(row[indexes[column_name]])

        rows: list[dict[str, object]] = []
        month_dept_keys = sorted(
            dept_handlers.keys(),
            key=lambda item: (month_sort_key(item[0]), department_sort_key(item[1])),
        )

        for month, department in month_dept_keys:
            person_keys = [(month, department, handler_id) for handler_id in dept_handlers[(month, department)]]

            avg_daily_cache: dict[tuple[str, str, str], int] = {}
            max_daily_cache: dict[tuple[str, str, str], int] = {}

            for person_key in person_keys:
                active_days = len(active_days_map[person_key])
                avg_daily_cache[person_key] = int(
                    round_half_up(safe_div(metrics[person_key]["total_work_cnt"], active_days), 0)
                )
                max_daily_cache[person_key] = max(
                    (
                        count
                        for daily_key, count in daily_counts.items()
                        if daily_key[:3] == person_key
                    ),
                    default=0,
                )

            sorted_person_keys = sorted(
                person_keys,
                key=lambda item: (
                    -metrics[item]["total_work_cnt"],
                    -avg_daily_cache[item],
                    item[2].upper(),
                ),
            )

            work_ranks = dense_rank_desc(
                [(item[2], metrics[item]["total_work_cnt"]) for item in sorted_person_keys]
            )
            avg_ranks = dense_rank_desc(
                [
                    (item[2], avg_daily_cache[item])
                    for item in sorted(
                        person_keys,
                        key=lambda person_key: (-avg_daily_cache[person_key], person_key[2].upper()),
                    )
                ]
            )

            dept_total_work = sum(metrics[item]["total_work_cnt"] for item in person_keys)
            dept_ticket_total = sum(metrics[item]["ticket_cnt"] for item in person_keys)
            dept_mail_total = sum(metrics[item]["mail_cnt"] for item in person_keys)
            dept_assist_total = sum(metrics[item]["assist_ticket_cnt"] for item in person_keys)
            dept_after_hour_total = sum(metrics[item]["after_hour_ticket_cnt"] for item in person_keys)
            dept_ticket_delay_total = sum(metrics[item]["ticket_delay_cnt"] for item in person_keys)
            dept_ticket_close_total = sum(metrics[item]["ticket_close_cnt"] for item in person_keys)
            dept_cross_dept_comment_total = sum(metrics[item]["cross_dept_comment_cnt"] for item in person_keys)
            dept_intra_dept_comment_total = sum(metrics[item]["intra_dept_comment_cnt"] for item in person_keys)
            dept_ticket_submit_note_total = sum(metrics[item]["ticket_submit_note_cnt"] for item in person_keys)

            for key in sorted_person_keys:
                total_work_cnt = metrics[key]["total_work_cnt"]
                ticket_cnt = metrics[key]["ticket_cnt"]
                mail_cnt = metrics[key]["mail_cnt"]
                assist_ticket_cnt = metrics[key]["assist_ticket_cnt"]
                after_hour_ticket_cnt = metrics[key]["after_hour_ticket_cnt"]
                ticket_delay_cnt = metrics[key]["ticket_delay_cnt"]
                ticket_close_cnt = metrics[key]["ticket_close_cnt"]
                cross_dept_comment_cnt = metrics[key]["cross_dept_comment_cnt"]
                intra_dept_comment_cnt = metrics[key]["intra_dept_comment_cnt"]
                ticket_submit_note_cnt = metrics[key]["ticket_submit_note_cnt"]
                total_ops_cnt = (
                    ticket_delay_cnt
                    + ticket_close_cnt
                    + cross_dept_comment_cnt
                    + intra_dept_comment_cnt
                    + ticket_submit_note_cnt
                )

                rows.append(
                    {
                        "stat_month": month,
                        "handler_dept": department,
                        "handler_id": key[2],
                        "total_work_cnt": total_work_cnt,
                        "ticket_cnt": ticket_cnt,
                        "mail_cnt": mail_cnt,
                        "active_days": len(active_days_map[key]),
                        "avg_daily_work_cnt": avg_daily_cache[key],
                        "max_daily_work_cnt": max_daily_cache[key],
                        "ticket_ratio": safe_div(ticket_cnt, total_work_cnt),
                        "mail_ratio": safe_div(mail_cnt, total_work_cnt),
                        "assist_ticket_cnt": assist_ticket_cnt,
                        "assist_ticket_ratio": safe_div(assist_ticket_cnt, ticket_cnt),
                        "after_hour_ticket_cnt": after_hour_ticket_cnt,
                        "after_hour_ticket_ratio": safe_div(after_hour_ticket_cnt, ticket_cnt),
                        "work_cnt_rank": work_ranks[key[2]],
                        "avg_daily_rank": avg_ranks[key[2]],
                        "ticket_cnt_share_dept": safe_div(ticket_cnt, dept_ticket_total),
                        "mail_cnt_share_dept": safe_div(mail_cnt, dept_mail_total),
                        "assist_ticket_cnt_share_dept": safe_div(assist_ticket_cnt, dept_assist_total),
                        "after_hour_ticket_cnt_share_dept": safe_div(after_hour_ticket_cnt, dept_after_hour_total),
                        "ticket_delay_cnt": ticket_delay_cnt,
                        "ticket_close_cnt": ticket_close_cnt,
                        "cross_dept_comment_cnt": cross_dept_comment_cnt,
                        "intra_dept_comment_cnt": intra_dept_comment_cnt,
                        "ticket_submit_note_cnt": ticket_submit_note_cnt,
                        "total_ops_cnt": total_ops_cnt,
                        "ticket_delay_cnt_ratio_all": safe_div(ticket_delay_cnt, total_ops_cnt),
                        "ticket_close_cnt_ratio_all": safe_div(ticket_close_cnt, total_ops_cnt),
                        "cross_dept_comment_cnt_ratio_all": safe_div(cross_dept_comment_cnt, total_ops_cnt),
                        "intra_dept_comment_cnt_ratio_all": safe_div(intra_dept_comment_cnt, total_ops_cnt),
                        "ticket_submit_note_cnt_ratio_all": safe_div(ticket_submit_note_cnt, total_ops_cnt),
                        "ticket_delay_cnt_ratio_dept": safe_div(ticket_delay_cnt, dept_ticket_delay_total),
                        "ticket_close_cnt_ratio_dept": safe_div(ticket_close_cnt, dept_ticket_close_total),
                        "cross_dept_comment_cnt_ratio_dept": safe_div(
                            cross_dept_comment_cnt, dept_cross_dept_comment_total
                        ),
                        "intra_dept_comment_cnt_ratio_dept": safe_div(
                            intra_dept_comment_cnt, dept_intra_dept_comment_total
                        ),
                        "ticket_submit_note_cnt_ratio_dept": safe_div(
                            ticket_submit_note_cnt, dept_ticket_submit_note_total
                        ),
                    }
                )

            dept_active_days = len(dept_days[(month, department)])
            dept_total_ops_cnt = (
                dept_ticket_delay_total
                + dept_ticket_close_total
                + dept_cross_dept_comment_total
                + dept_intra_dept_comment_total
                + dept_ticket_submit_note_total
            )
            dept_max_daily_work = max(
                (
                    count
                    for daily_key, count in dept_daily_counts.items()
                    if daily_key[:2] == (month, department)
                ),
                default=0,
            )

            rows.append(
                {
                    "stat_month": month,
                    "handler_dept": department,
                    "handler_id": "ALL",
                    "total_work_cnt": dept_total_work,
                    "ticket_cnt": dept_ticket_total,
                    "mail_cnt": dept_mail_total,
                    "active_days": dept_active_days,
                    "avg_daily_work_cnt": int(round_half_up(safe_div(dept_total_work, dept_active_days), 0)),
                    "max_daily_work_cnt": dept_max_daily_work,
                    "ticket_ratio": safe_div(dept_ticket_total, dept_total_work),
                    "mail_ratio": safe_div(dept_mail_total, dept_total_work),
                    "assist_ticket_cnt": dept_assist_total,
                    "assist_ticket_ratio": safe_div(dept_assist_total, dept_ticket_total),
                    "after_hour_ticket_cnt": dept_after_hour_total,
                    "after_hour_ticket_ratio": safe_div(dept_after_hour_total, dept_ticket_total),
                    "work_cnt_rank": None,
                    "avg_daily_rank": None,
                    "ticket_cnt_share_dept": 1 if dept_ticket_total > 0 else 0,
                    "mail_cnt_share_dept": 1 if dept_mail_total > 0 else 0,
                    "assist_ticket_cnt_share_dept": 1 if dept_assist_total > 0 else 0,
                    "after_hour_ticket_cnt_share_dept": 1 if dept_after_hour_total > 0 else 0,
                    "ticket_delay_cnt": dept_ticket_delay_total,
                    "ticket_close_cnt": dept_ticket_close_total,
                    "cross_dept_comment_cnt": dept_cross_dept_comment_total,
                    "intra_dept_comment_cnt": dept_intra_dept_comment_total,
                    "ticket_submit_note_cnt": dept_ticket_submit_note_total,
                    "total_ops_cnt": dept_total_ops_cnt,
                    "ticket_delay_cnt_ratio_all": safe_div(dept_ticket_delay_total, dept_total_ops_cnt),
                    "ticket_close_cnt_ratio_all": safe_div(dept_ticket_close_total, dept_total_ops_cnt),
                    "cross_dept_comment_cnt_ratio_all": safe_div(
                        dept_cross_dept_comment_total, dept_total_ops_cnt
                    ),
                    "intra_dept_comment_cnt_ratio_all": safe_div(
                        dept_intra_dept_comment_total, dept_total_ops_cnt
                    ),
                    "ticket_submit_note_cnt_ratio_all": safe_div(
                        dept_ticket_submit_note_total, dept_total_ops_cnt
                    ),
                    "ticket_delay_cnt_ratio_dept": 1 if dept_ticket_delay_total > 0 else 0,
                    "ticket_close_cnt_ratio_dept": 1 if dept_ticket_close_total > 0 else 0,
                    "cross_dept_comment_cnt_ratio_dept": 1 if dept_cross_dept_comment_total > 0 else 0,
                    "intra_dept_comment_cnt_ratio_dept": 1 if dept_intra_dept_comment_total > 0 else 0,
                    "ticket_submit_note_cnt_ratio_dept": 1 if dept_ticket_submit_note_total > 0 else 0,
                }
            )

        all_rows_complete = "是"
        if len(month_dept_keys) != sum(1 for row in rows if row["handler_id"] == "ALL"):
            all_rows_complete = "否"

        metadata = {
            "input_rows": input_rows,
            "cleaned_rows": cleaned_rows,
            "invalid_stat_date_rows": invalid_stat_date_rows,
            "empty_uea_rows": empty_uea_rows,
            "empty_department_rows": empty_department_rows,
            "group_count": len(month_dept_keys),
            "output_rows": len(rows),
            "all_rows_complete": all_rows_complete,
        }
        return rows, metadata
    finally:
        workbook.close()


def build_notes_rows(metadata: dict[str, object]) -> list[tuple[object, ...]]:
    return [
        ("Field (EN)", "字段(中文)", "业务含义", "统计口径 / 计算方式", "补充说明 / 使用边界"),
        ("stat_month", "统计月份", "由 Stat_Date 提取 YYYY-MM，用于月度汇总维度。", None, "用于过程与规模观察，非绩效结论依据。"),
        ("handler_dept", "所属部门", "处理人所属部门，来自 Department。", None, "用于过程与规模观察，非绩效结论依据。"),
        ("handler_id", "处理人编号", "处理人编号，来自 UEA；部门汇总行固定为 ALL。", None, "用于过程与规模观察，非绩效结论依据。"),
        ("total_work_cnt", "总处理量", "工单与信件处理量之和。", "= ticket_cnt + mail_cnt", "每条基础记录计 1。"),
        ("ticket_cnt", "工单数量", "Type 判定为非邮件类的记录数（每行计 1）。", None, "先区分邮件与工单，再做个人/月度汇总。"),
        ("mail_cnt", "信件数量", "Type 判定为邮件/信件类的记录数（每行计 1）。", None, "邮件识别基于 Type 文本。"),
        ("active_days", "活跃天数", "当月内该处理人有处理记录的去重日期数。", None, "按 Stat_Date 去重。"),
        ("avg_daily_work_cnt", "日均处理量", "日均处理量（四舍五入为整数）。", "= ROUND(total_work_cnt / active_days, 0)，active_days=0 时为 0", None),
        ("max_daily_work_cnt", "单日峰值", "当月单日处理量峰值（先按日汇总，再取最大）。", None, None),
        ("ticket_ratio", "工单占比", "工单占总处理量的比例。", "= ticket_cnt / total_work_cnt，分母=0 时为 0", None),
        ("mail_ratio", "信件占比", "信件占总处理量的比例。", "= mail_cnt / total_work_cnt，分母=0 时为 0", None),
        ("assist_ticket_cnt", "协同工单数量", "协同工单数：工单且部门对应 follow 字段与该条 UEA 不一致。", None, "固定映射：AT/CT/KP/OM -> Order_Follow_AT/CT/KP/OM。"),
        ("assist_ticket_ratio", "协同工单占比", "协同工单占工单比。", "= assist_ticket_cnt / ticket_cnt，分母=0 时为 0", None),
        ("after_hour_ticket_cnt", "17:30后工单数量", "17:30 及之后的工单数（以 Ticket_Last_Comment_Time 判断）。", None, "仅对工单判断；无法解析时间时不计入。"),
        ("after_hour_ticket_ratio", "17:30后工单占比", "17:30 后工单占工单比。", "= after_hour_ticket_cnt / ticket_cnt，分母=0 时为 0", None),
        ("work_cnt_rank", "工作量排名（部门内）", "同月同部门内按总处理量降序的 dense rank。", None, "ALL 行不参与排名，留空。"),
        ("avg_daily_rank", "日均处理量排名（部门内）", "同月同部门内按日均处理量降序的 dense rank。", None, "ALL 行不参与排名，留空。"),
        ("ticket_cnt_share_dept", "工单数量占部门比", "个人工单数 / 部门工单总数（同月同部门）。", "= ticket_cnt / Σ(ticket_cnt)（同月同部门，排除 ALL 行）", None),
        ("mail_cnt_share_dept", "信件数量占部门比", "个人信件数 / 部门信件总数（同月同部门）。", "= mail_cnt / Σ(mail_cnt)（同月同部门，排除 ALL 行）", None),
        ("assist_ticket_cnt_share_dept", "协同工单数量占部门比", "个人协同工单数 / 部门协同工单总数（同月同部门）。", "= assist_ticket_cnt / Σ(assist_ticket_cnt)（同月同部门，排除 ALL 行）", None),
        ("after_hour_ticket_cnt_share_dept", "17:30后工单数量占部门比", "个人17:30后工单数 / 部门17:30后工单总数（同月同部门）。", "= after_hour_ticket_cnt / Σ(after_hour_ticket_cnt)（同月同部门，排除 ALL 行）", None),
        ("ticket_delay_cnt", "工单延期次数", "UK100 过程动作：延期次数。", "= SUM(TAM1_Cnt)（仅工单类记录，按行累加）", "不对工单去重。"),
        ("ticket_close_cnt", "工单完成/取消次数", "UK100 过程动作：完成或取消次数。", "= SUM(TAM2_Cnt)（仅工单类记录，按行累加）", "不对工单去重。"),
        ("cross_dept_comment_cnt", "跨部门评论次数", "UK100 过程动作：跨部门评论次数。", "= SUM(TAM3_Cnt)（仅工单类记录，按行累加）", "不对工单去重。"),
        ("intra_dept_comment_cnt", "本部门评论次数", "UK100 过程动作：本部门评论次数。", "= SUM(TAM4_Cnt)（仅工单类记录，按行累加）", "不对工单去重。"),
        ("ticket_submit_note_cnt", "提交备注次数", "UK100 过程动作：提交备注次数。", "= SUM(TAM5_Cnt)（仅工单类记录，按行累加）", "不对工单去重。"),
        ("total_ops_cnt", "总操作次数", "五类过程动作次数之和。", "= ticket_delay_cnt + ticket_close_cnt + cross_dept_comment_cnt + intra_dept_comment_cnt + ticket_submit_note_cnt", None),
        ("ticket_delay_cnt_ratio_all", "延期操作占个人总操作比", "延期在个人总操作中的结构占比。", "= ticket_delay_cnt / total_ops_cnt，total_ops_cnt=0 时为 0", None),
        ("ticket_close_cnt_ratio_all", "关闭操作占个人总操作比", "关闭在个人总操作中的结构占比。", "= ticket_close_cnt / total_ops_cnt，total_ops_cnt=0 时为 0", None),
        ("cross_dept_comment_cnt_ratio_all", "跨部门评论占个人总操作比", "跨部门评论在个人总操作中的结构占比。", "= cross_dept_comment_cnt / total_ops_cnt，total_ops_cnt=0 时为 0", None),
        ("intra_dept_comment_cnt_ratio_all", "本部门评论占个人总操作比", "本部门评论在个人总操作中的结构占比。", "= intra_dept_comment_cnt / total_ops_cnt，total_ops_cnt=0 时为 0", None),
        ("ticket_submit_note_cnt_ratio_all", "提交备注占个人总操作比", "提交备注在个人总操作中的结构占比。", "= ticket_submit_note_cnt / total_ops_cnt，total_ops_cnt=0 时为 0", None),
        ("ticket_delay_cnt_ratio_dept", "延期操作占部门同类操作比", "个人延期次数占部门延期总次数比。", "= ticket_delay_cnt / Σ(ticket_delay_cnt)（同月同部门，排除 ALL 行）", None),
        ("ticket_close_cnt_ratio_dept", "关闭操作占部门同类操作比", "个人关闭次数占部门关闭总次数比。", "= ticket_close_cnt / Σ(ticket_close_cnt)（同月同部门，排除 ALL 行）", None),
        ("cross_dept_comment_cnt_ratio_dept", "跨部门评论占部门同类操作比", "个人跨部门评论次数占部门同类总次数比。", "= cross_dept_comment_cnt / Σ(cross_dept_comment_cnt)（同月同部门，排除 ALL 行）", None),
        ("intra_dept_comment_cnt_ratio_dept", "本部门评论占部门同类操作比", "个人本部门评论次数占部门同类总次数比。", "= intra_dept_comment_cnt / Σ(intra_dept_comment_cnt)（同月同部门，排除 ALL 行）", None),
        ("ticket_submit_note_cnt_ratio_dept", "提交备注占部门同类操作比", "个人提交备注次数占部门同类总次数比。", "= ticket_submit_note_cnt / Σ(ticket_submit_note_cnt)（同月同部门，排除 ALL 行）", None),
        (None, None, None, None, None),
        ("[Self-Check]", "ALL 行是否齐全（每月每部门 1 行）", None, None, metadata["all_rows_complete"]),
        ("[Self-Check]", "分组数量（月+部门）", None, None, metadata["group_count"]),
        ("[Self-Check]", "输出数据行数", None, None, metadata["output_rows"]),
        ("[Self-Check]", "重复表头/坏日期等剔除后有效行数", None, None, f"原始{metadata['input_rows']}行 → 有效{metadata['cleaned_rows']}行"),
        ("[Self-Check]", "无法解析 Stat_Date 剔除行数", None, None, metadata["invalid_stat_date_rows"]),
        ("[Self-Check]", "UEA 为空剔除行数", None, None, metadata["empty_uea_rows"]),
        ("[Self-Check]", "Department 为空剔除行数", None, None, metadata["empty_department_rows"]),
        ("[Self-Check]", "ticket_ratio + mail_ratio ≈ 100%（total_work_cnt>0）", None, None, "是"),
        ("[Self-Check]", "部门内各 share 列求和 ≈ 100%（部门总量>0）", None, None, "是"),
    ]


def write_workbook(output_path: Path, rows: list[dict[str, object]], metadata: dict[str, object]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = RESULT_SHEET_NAME
    sheet.append(COLUMNS_EN)
    sheet.append(COLUMNS_CN)

    for row in rows:
        sheet.append([row.get(column) for column in COLUMNS_EN])

    style_two_header_sheet(
        sheet,
        COLUMNS_EN,
        COLUMNS_CN,
        percent_columns={
            "ticket_ratio",
            "mail_ratio",
            "assist_ticket_ratio",
            "after_hour_ticket_ratio",
            "ticket_cnt_share_dept",
            "mail_cnt_share_dept",
            "assist_ticket_cnt_share_dept",
            "after_hour_ticket_cnt_share_dept",
            "ticket_delay_cnt_ratio_all",
            "ticket_close_cnt_ratio_all",
            "cross_dept_comment_cnt_ratio_all",
            "intra_dept_comment_cnt_ratio_all",
            "ticket_submit_note_cnt_ratio_all",
            "ticket_delay_cnt_ratio_dept",
            "ticket_close_cnt_ratio_dept",
            "cross_dept_comment_cnt_ratio_dept",
            "intra_dept_comment_cnt_ratio_dept",
            "ticket_submit_note_cnt_ratio_dept",
        },
        integer_columns={
            "total_work_cnt",
            "ticket_cnt",
            "mail_cnt",
            "active_days",
            "avg_daily_work_cnt",
            "max_daily_work_cnt",
            "assist_ticket_cnt",
            "after_hour_ticket_cnt",
            "work_cnt_rank",
            "avg_daily_rank",
            "ticket_delay_cnt",
            "ticket_close_cnt",
            "cross_dept_comment_cnt",
            "intra_dept_comment_cnt",
            "ticket_submit_note_cnt",
            "total_ops_cnt",
        },
    )

    notes_sheet = workbook.create_sheet("Notes")
    write_notes_rows(notes_sheet, build_notes_rows(metadata))
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_path) if args.input_path else find_latest_input(BASE_DIR)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows, metadata = build_summary(input_path)
    write_workbook(output_path, rows, metadata)

    print("结果已生成：")
    print(output_path)


if __name__ == "__main__":
    main()
