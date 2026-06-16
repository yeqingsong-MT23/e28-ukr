# -*- coding: utf-8 -*-
"""
根据 UK200_转化漏斗KPI 基础表生成 UK201_转化漏斗概览.xlsx。

默认行为：
1. 自动读取脚本所在目录下最新的 UK200 开头基础表
2. 输出到 P:\\0_Report\\78_AI Report\\UKR_Report 下，文件名自动追加 -MMDDHHMM

示例：
python uk201.py
python uk201.py --input "C:\\\\path\\\\UK200_转化漏斗KPI-20260430.xlsx"
python uk201.py --output "C:\\\\path\\\\UK201_转化漏斗概览.xlsx"
"""

from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from uk_export_utils import build_default_output_path
from uk_funnel_utils import (
    detect_data_sheet,
    find_latest_input,
    load_stage_qpo_records,
    month_sort_key,
    normalize_text,
    parse_cell_date,
    safe_div,
    style_two_header_sheet,
)

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK201_转化漏斗概览.xlsx"

CANCEL_STATUS = {
    "SellerCancel",
    "CSCancel",
    "Long time-Cancel",
    "WootCancel",
}

IN_PROGRESS_STATUS = {
    "Hold",
    "New",
    "SubmitRBS",
}

DEAL_STATUS = {
    "End",
    "WaitingRPO",
    "WaitingPayment",
    "WaitingDeal",
    "OnDeal",
    "WaitingPO",
}

BASE_COLUMNS_EN = [
    "Month",
    "Total_QPO",
    "Cancel_QPO",
    "Deal_QPO",
    "InProgress_QPO",
    "Cancel_Rate",
    "Deal_Rate",
    "InProgress_Rate",
]

BASE_COLUMNS_CN = [
    "月份",
    "提报QPO总数",
    "取消QPO数",
    "成交QPO数",
    "跟进中QPO数",
    "取消占比",
    "成交占比",
    "跟进中占比",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据 UK200 基础表生成 UK201_转化漏斗概览。"
    )
    parser.add_argument(
        "--input",
        dest="input_path",
        help="UK200 基础表路径；不传时自动读取当前目录最新的 UK200 开头基础表",
    )
    parser.add_argument(
        "--output",
        dest="output_path",
        help=f"输出文件路径；不传时默认写入 {EXPORT_DIR}，文件名自动追加 -MMDDHHMM",
    )
    return parser.parse_args()


def has_chinese(text: object) -> bool:
    content = str(text or "")
    return any("\u4e00" <= char <= "\u9fff" for char in content)


def resolve_column_indexes(header_row: list[str]) -> dict[str, int]:
    matches: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(header_row):
        matches[normalize_text(header)].append(index)

    indexes: dict[str, int] = {}
    missing: list[str] = []

    required_single = {
        "Submission_Date": "Submission_Date",
        "QPO_ID": "QPO_ID",
        "QPO_Status": "QPO_Status",
    }

    for key, header_name in required_single.items():
        if matches.get(header_name):
            indexes[key] = matches[header_name][0]
        else:
            missing.append(header_name)

    node_headers = matches.get("WPO_Current_Node", [])
    if len(node_headers) >= 2:
        indexes["WPO_Current_Node"] = node_headers[0]
        indexes["WPO_Current_Node_CN"] = node_headers[1]
    else:
        missing.append("WPO_Current_Node(需存在代码列和中文列各1个)")

    if missing:
        raise ValueError(f"源表缺少必要字段：{missing}")

    return indexes


def map_status_category(status: str) -> str | None:
    if status in CANCEL_STATUS:
        return "Cancel"
    if status in IN_PROGRESS_STATUS:
        return "InProgress"
    if status in DEAL_STATUS:
        return "Deal"
    return None


def normalize_node(node: str) -> str:
    text = normalize_text(node).upper()
    return text if re.fullmatch(r"N\d{3}", text) else "Unknown"


def node_attribution_rank(node: str) -> tuple[int, int]:
    matched = re.fullmatch(r"N(\d{3})", normalize_text(node).upper())
    if not matched:
        return (0, -1)
    return (1, int(matched.group(1)))


def node_sort_key(node: str) -> tuple[int, int, str]:
    if node == "Unknown":
        return (0, -1, node)
    matched = re.fullmatch(r"N(\d{3})", node)
    if matched:
        return (1, int(matched.group(1)), node)
    return (2, 999999, node)


def get_row_value(row: tuple[object, ...], index: int) -> object:
    return row[index] if 0 <= index < len(row) else None


def build_summary(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    qpo_records, _ = load_stage_qpo_records(input_path)

    total_qpo: dict[str, set[str]] = defaultdict(set)
    category_qpo: dict[str, dict[str, set[str]]] = {
        "Cancel": defaultdict(set),
        "Deal": defaultdict(set),
        "InProgress": defaultdict(set),
    }
    qpo_month_map: dict[str, str] = {}
    qpo_final_category: dict[str, str | None] = {}
    unmapped_status = Counter()

    for record in qpo_records:
        qpo_id = normalize_text(record.get("QPO_ID"))
        if not qpo_id:
            continue

        submission_date = parse_cell_date(record.get("Submission_Date"))
        if submission_date is None:
            continue

        month = f"{submission_date.year:04d}-{submission_date.month:02d}"
        status = normalize_text(record.get("QPO_Status"))
        category = map_status_category(status)

        total_qpo[month].add(qpo_id)
        total_qpo["ALL"].add(qpo_id)
        qpo_month_map[qpo_id] = month
        qpo_final_category[qpo_id] = category

        if category is None:
            unmapped_status[status or "<blank>"] += 1
            continue

        category_qpo[category][month].add(qpo_id)
        category_qpo[category]["ALL"].add(qpo_id)

    cancel_best_node: dict[str, dict[str, tuple[tuple[int, int], int, str]]] = {"ALL": {}}
    node_cn_map: dict[str, str] = {"Unknown": "Unknown"}
    missing_node_cn_qpo_all: set[str] = set()
    unknown_node_qpo_all: set[str] = set()

    input_rows = 0
    valid_date_rows = 0
    row_sequence = 0

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = detect_data_sheet(workbook)

        try:
            header_row_1 = [
                normalize_text(value)
                for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            ]
        except StopIteration as exc:
            raise ValueError("源表缺少表头，无法生成报表。") from exc

        try:
            header_row_2 = [
                normalize_text(value)
                for value in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
            ]
        except StopIteration:
            header_row_2 = []

        data_start_row = 3 if header_row_2 and sum(has_chinese(value) for value in header_row_2) > 0 else 2
        column_indexes = resolve_column_indexes(header_row_1)

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            input_rows += 1

            submission_date = parse_cell_date(get_row_value(row, column_indexes["Submission_Date"]))
            if submission_date is None:
                continue

            valid_date_rows += 1

            qpo_id = normalize_text(get_row_value(row, column_indexes["QPO_ID"]))
            if not qpo_id:
                continue

            if qpo_final_category.get(qpo_id) != "Cancel":
                continue

            month = qpo_month_map.get(qpo_id)
            if not month:
                continue

            raw_node = normalize_text(get_row_value(row, column_indexes["WPO_Current_Node"]))
            raw_node_cn = normalize_text(get_row_value(row, column_indexes["WPO_Current_Node_CN"]))
            node = normalize_node(raw_node)

            if not raw_node_cn:
                missing_node_cn_qpo_all.add(qpo_id)

            if node == "Unknown":
                unknown_node_qpo_all.add(qpo_id)
            elif raw_node_cn:
                node_cn_map.setdefault(node, raw_node_cn)

            row_sequence += 1
            candidate = (node_attribution_rank(raw_node), row_sequence, node)

            for bucket in (month, "ALL"):
                bucket_map = cancel_best_node.setdefault(bucket, {})
                current = bucket_map.get(qpo_id)
                if current is None or candidate[:2] >= current[:2]:
                    bucket_map[qpo_id] = candidate
    finally:
        workbook.close()

    cancel_qpo_ids = {
        qpo_id
        for qpo_id, category in qpo_final_category.items()
        if category == "Cancel" and qpo_id in qpo_month_map
    }

    for qpo_id in cancel_qpo_ids:
        month = qpo_month_map[qpo_id]
        for bucket in (month, "ALL"):
            bucket_map = cancel_best_node.setdefault(bucket, {})
            if qpo_id not in bucket_map:
                row_sequence += 1
                bucket_map[qpo_id] = (node_attribution_rank(""), row_sequence, "Unknown")
                unknown_node_qpo_all.add(qpo_id)
                missing_node_cn_qpo_all.add(qpo_id)

    nodes_sorted = sorted(
        {value[2] for bucket_map in cancel_best_node.values() for value in bucket_map.values()},
        key=node_sort_key,
    )

    rows: list[dict[str, object]] = []
    month_keys = sorted((month for month in total_qpo if month != "ALL"), key=month_sort_key)

    for month in month_keys + ["ALL"]:
        total = len(total_qpo.get(month, set()))
        cancel = len(category_qpo["Cancel"].get(month, set()))
        deal = len(category_qpo["Deal"].get(month, set()))
        in_progress = len(category_qpo["InProgress"].get(month, set()))

        node_counter = Counter(
            candidate[2] for candidate in cancel_best_node.get(month, {}).values()
        )

        row_data: dict[str, object] = {
            "Month": month,
            "Total_QPO": total,
            "Cancel_QPO": cancel,
            "Deal_QPO": deal,
            "InProgress_QPO": in_progress,
            "Cancel_Rate": safe_div(cancel, total),
            "Deal_Rate": safe_div(deal, total),
            "InProgress_Rate": safe_div(in_progress, total),
        }

        for node in nodes_sorted:
            node_count = node_counter.get(node, 0)
            row_data[f"{node}_Cancel_QPO_Count"] = node_count
            row_data[f"{node}_Cancel_QPO_Rate"] = safe_div(node_count, total)

        rows.append(row_data)

    checks = []
    for month in month_keys + ["ALL"]:
        cancel = len(category_qpo["Cancel"].get(month, set()))
        deal = len(category_qpo["Deal"].get(month, set()))
        in_progress = len(category_qpo["InProgress"].get(month, set()))
        total = len(total_qpo.get(month, set()))
        node_sum = sum(
            Counter(candidate[2] for candidate in cancel_best_node.get(month, {}).values()).values()
        )
        checks.append(
            {
                "month": month,
                "category_le_total": cancel + deal + in_progress <= total,
                "node_sum_equals_cancel": node_sum == cancel,
            }
        )

    all_cancel = len(category_qpo["Cancel"].get("ALL", set()))
    distinct_qpo_count = len(total_qpo.get("ALL", set()))

    metadata = {
        "input_path": str(input_path),
        "nodes_sorted": nodes_sorted,
        "node_cn_map": node_cn_map,
        "input_rows": input_rows,
        "valid_date_rows": valid_date_rows,
        "distinct_qpo_count": distinct_qpo_count,
        "month_count": len(month_keys),
        "unmapped_statuses": dict(unmapped_status),
        "unmapped_status_qpo_count": sum(unmapped_status.values()),
        "missing_node_cn_rate": safe_div(len(missing_node_cn_qpo_all), distinct_qpo_count),
        "unknown_node_rate_in_cancel": safe_div(len(unknown_node_qpo_all), all_cancel),
        "checks": checks,
    }
    return rows, metadata


def build_column_headers(
    nodes_sorted: list[str], node_cn_map: dict[str, str]
) -> tuple[list[str], list[str]]:
    columns_en = list(BASE_COLUMNS_EN)
    columns_cn = list(BASE_COLUMNS_CN)

    for node in nodes_sorted:
        node_name_cn = node_cn_map.get(node, node)
        columns_en.extend(
            [
                f"{node}_Cancel_QPO_Count",
                f"{node}_Cancel_QPO_Rate",
            ]
        )
        columns_cn.extend(
            [
                f"{node_name_cn}取消QPO数量",
                f"{node_name_cn}取消占提报比例",
            ]
        )

    return columns_en, columns_cn


def write_main_sheet(
    workbook: Workbook,
    rows: list[dict[str, object]],
    metadata: dict[str, object],
) -> None:
    nodes_sorted = metadata["nodes_sorted"]
    node_cn_map = metadata["node_cn_map"]
    columns_en, columns_cn = build_column_headers(nodes_sorted, node_cn_map)

    sheet = workbook.active
    sheet.title = "UK201_转化漏斗概览"
    sheet.append(columns_en)
    sheet.append(columns_cn)

    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    style_two_header_sheet(sheet, columns_en, columns_cn)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    count_columns = {
        "Total_QPO",
        "Cancel_QPO",
        "Deal_QPO",
        "InProgress_QPO",
    }
    rate_columns = {
        "Cancel_Rate",
        "Deal_Rate",
        "InProgress_Rate",
    }

    for column in columns_en:
        if column.endswith("_Cancel_QPO_Count"):
            count_columns.add(column)
        elif column.endswith("_Cancel_QPO_Rate"):
            rate_columns.add(column)

    column_index = {column: idx + 1 for idx, column in enumerate(columns_en)}

    for row_num in range(3, sheet.max_row + 1):
        for column in columns_en:
            cell = sheet.cell(row=row_num, column=column_index[column])
            cell.alignment = center
            if column in count_columns:
                cell.number_format = "0"
            elif column in rate_columns:
                cell.number_format = "0.00%"

    preview_rows = min(sheet.max_row, 30)
    for idx, column in enumerate(columns_en, start=1):
        sample_lengths = [
            len(str(column)),
            len(str(columns_cn[idx - 1])),
        ]
        for row_num in range(3, preview_rows + 1):
            value = sheet.cell(row=row_num, column=idx).value
            if value is not None:
                sample_lengths.append(len(str(value)))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(sample_lengths) + 2, 35)

    sheet.freeze_panes = "A3"


def write_notes_sheet(workbook: Workbook, metadata: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Notes")
    nodes_sorted = metadata["nodes_sorted"]
    node_cn_map = metadata["node_cn_map"]
    unmapped_statuses = metadata["unmapped_statuses"]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = [
        ("Field", "Definition/Rule"),
        ("字段", "释义/口径"),
        ("Month", "YYYY-MM derived from Submission_Date (提报日期)"),
        ("Total_QPO", "Distinct QPO_ID count per Month after QPO-level deduplication"),
        ("Cancel_QPO", "Distinct QPO_ID where final mapped type = Cancel"),
        ("Deal_QPO", "Distinct QPO_ID where final mapped type = Deal"),
        ("InProgress_QPO", "Distinct QPO_ID where final mapped type = InProgress"),
        ("Cancel_Rate", "Cancel_QPO / Total_QPO"),
        ("Deal_Rate", "Deal_QPO / Total_QPO"),
        ("InProgress_Rate", "InProgress_QPO / Total_QPO"),
        (
            "<Node>_Cancel_QPO_Count",
            "For final Cancel QPOs, scan all source rows of the same QPO_ID and attribute to the deepest valid WPO_Current_Node in Nxxx format, then count by Month+Node",
        ),
        (
            "<Node>_Cancel_QPO_Rate",
            "Node_Cancel_QPO_Count / Total_QPO of the same Month (denominator is Total_QPO, not Cancel_QPO)",
        ),
        (
            "QPO_Type=Cancel",
            "QPO_Status ∈ {SellerCancel, CSCancel, Long time-Cancel, WootCancel}",
        ),
        (
            "QPO_Type=InProgress",
            "QPO_Status ∈ {Hold, New, SubmitRBS}",
        ),
        (
            "QPO_Type=Deal",
            "QPO_Status ∈ {End, WaitingRPO, WaitingPayment, WaitingDeal, OnDeal, WaitingPO}",
        ),
        ("QPO_Type=Other", "Any other QPO_Status; included in Total_QPO but excluded from the 3 mapped categories"),
        (
            "Max Node Attribution",
            "Choose the row with the largest numeric part in Nxxx as attribution node; unparseable treated as Unknown/smallest; ties keep last occurrence",
        ),
        (
            "Cancel Nodes Included",
            ", ".join(f"{node}={node_cn_map.get(node, node)}" for node in nodes_sorted) if nodes_sorted else "<none>",
        ),
        ("ALL", "Grand total across all months"),
        ("----", "----"),
        ("Input/Processing Stats", None),
        ("input_file", metadata["input_path"]),
        ("row_count_in", metadata["input_rows"]),
        ("row_count_valid_after_date_parse", metadata["valid_date_rows"]),
        ("distinct_qpo_count", metadata["distinct_qpo_count"]),
        ("month_count", metadata["month_count"]),
        ("unmapped_status_qpo_count", metadata["unmapped_status_qpo_count"]),
        ("missing_node_cn_rate", metadata["missing_node_cn_rate"]),
        ("unknown_node_rate_in_cancel", metadata["unknown_node_rate_in_cancel"]),
    ]

    if unmapped_statuses:
        rows.append(("unmapped_status_breakdown", "count by final QPO_Status"))
        for status, count in sorted(unmapped_statuses.items(), key=lambda item: (-item[1], item[0])):
            rows.append((status, count))

    rows.extend(
        [
            ("----", "----"),
            ("Self-checks", None),
        ]
    )

    for row in rows:
        sheet.append(row)

    for check in metadata["checks"]:
        status_1 = "PASS" if check["category_le_total"] else "FAIL"
        status_2 = "PASS" if check["node_sum_equals_cancel"] else "FAIL"
        sheet.append(
            (
                f"Check: {check['month']} Cancel+Deal+InProgress ≤ Total",
                status_1,
            )
        )
        sheet.append(
            (
                f"Check: {check['month']} Σ(Node Cancel Count) = Cancel_QPO",
                status_2,
            )
        )

    main_sheet = workbook["UK201_转化漏斗概览"]
    all_row = None
    for row in main_sheet.iter_rows(min_row=3, values_only=True):
        if normalize_text(row[0]) == "ALL":
            all_row = row
            break

    if all_row is None:
        all_row = ("ALL", 0, 0, 0, 0, 0, 0, 0)

    total_value = all_row[BASE_COLUMNS_EN.index("Total_QPO")] or 0
    cancel_value = all_row[BASE_COLUMNS_EN.index("Cancel_QPO")] or 0
    deal_value = all_row[BASE_COLUMNS_EN.index("Deal_QPO")] or 0
    inprogress_value = all_row[BASE_COLUMNS_EN.index("InProgress_QPO")] or 0
    cancel_rate_value = all_row[BASE_COLUMNS_EN.index("Cancel_Rate")] or 0
    deal_rate_value = all_row[BASE_COLUMNS_EN.index("Deal_Rate")] or 0
    inprogress_rate_value = all_row[BASE_COLUMNS_EN.index("InProgress_Rate")] or 0

    sheet.append(
        (
            "Check: ALL Rate recomputed (Cancel_Rate)",
            "PASS" if abs(cancel_rate_value - safe_div(cancel_value, total_value)) < 1e-12 else "FAIL",
        )
    )
    sheet.append(
        (
            "Check: ALL Rate recomputed (Deal_Rate)",
            "PASS" if abs(deal_rate_value - safe_div(deal_value, total_value)) < 1e-12 else "FAIL",
        )
    )
    sheet.append(
        (
            "Check: ALL Rate recomputed (InProgress_Rate)",
            "PASS" if abs(inprogress_rate_value - safe_div(inprogress_value, total_value)) < 1e-12 else "FAIL",
        )
    )

    for row_num in (1, 2):
        for col_num in range(1, 3):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center
            if isinstance(cell.value, float) and (
                "rate" in str(sheet.cell(row=cell.row, column=1).value).lower()
                or str(sheet.cell(row=cell.row, column=1).value).endswith("_rate")
            ):
                cell.number_format = "0.00%"

    sheet.column_dimensions["A"].width = 46
    sheet.column_dimensions["B"].width = 120
    sheet.freeze_panes = "A3"


def write_workbook(
    output_path: Path,
    rows: list[dict[str, object]],
    metadata: dict[str, object],
) -> None:
    workbook = Workbook()
    try:
        write_main_sheet(workbook, rows, metadata)
        write_notes_sheet(workbook, metadata)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def generate_report(input_path: Path, output_path: Path) -> None:
    rows, metadata = build_summary(input_path=input_path)
    write_workbook(output_path=output_path, rows=rows, metadata=metadata)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_path) if args.input_path else find_latest_input(BASE_DIR)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    generate_report(input_path=input_path, output_path=output_path)

    print("生成完成：")
    print(output_path)


if __name__ == "__main__":
    main()
