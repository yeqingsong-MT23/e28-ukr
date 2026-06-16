# -*- coding: utf-8 -*-
from __future__ import annotations

import datetime as dt
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EMAIL_TYPE = "邮件"
TICKET_ONLY_TYPE = "工单"
TICKET_LIKE_TYPES = {"工单", "批量工单", "U3工单", "U6工单"}
DEPARTMENT_ORDER = {"AT": 0, "CT": 1, "KP": 2, "OM": 3, "ALL": 99}


def parse_cell_date(value) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(".", "-").replace("/", "-")
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def parse_cell_time(value) -> dt.time | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.time()
    if isinstance(value, dt.time):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            return dt.datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def has_chinese(text: object) -> bool:
    content = str(text or "")
    return any("\u4e00" <= char <= "\u9fff" for char in content)


def has_value(value) -> bool:
    return value is not None and str(value).strip() != ""


def safe_div(numerator: float | int, denominator: float | int) -> float:
    return numerator / denominator if denominator else 0.0


def round_half_up(value: float | int, digits: int = 0) -> float:
    quant = Decimal("1") if digits == 0 else Decimal("1." + ("0" * digits))
    rounded = Decimal(str(value)).quantize(quant, rounding=ROUND_HALF_UP)
    return float(rounded)


def month_sort_key(month: str) -> tuple[int, int]:
    text = normalize_text(month)
    if text == "ALL":
        return (9999, 99)
    if "-" in text:
        year, mon = text.split("-", 1)
        return (int(year), int(mon))
    if len(text) >= 6 and text[:6].isdigit():
        return (int(text[:4]), int(text[4:6]))
    return (9998, 99)


def department_sort_key(department: str) -> tuple[int, str]:
    return (DEPARTMENT_ORDER.get(department, 50), department)


def find_latest_input(base_dir: Path) -> Path:
    files = sorted(
        (
            path
            for path in base_dir.glob("UK100_工作量 KPI-*.xlsx")
            if not path.name.startswith("~$")
        ),
        key=lambda item: (item.stat().st_mtime, item.name),
        reverse=True,
    )
    if not files:
        raise FileNotFoundError(
            f"未找到 UK100 基础表，请确认目录下存在 UK100_工作量 KPI-*.xlsx：{base_dir}"
        )
    return files[0]


def detect_data_sheet(workbook):
    if "Data" in workbook.sheetnames:
        return workbook["Data"]
    return workbook[workbook.sheetnames[0]]


def get_header_matches(header_row: list[str]) -> dict[str, list[int]]:
    matches: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(header_row):
        matches[normalize_text(header)].append(index)
    return matches


def get_required_indexes(
    header_row: list[str],
    required_single: dict[str, str],
    duplicate_requirements: dict[str, tuple[str, int]] | None = None,
) -> dict[str, int]:
    matches = get_header_matches(header_row)
    indexes: dict[str, int] = {}
    missing: list[str] = []

    for key, header_name in required_single.items():
        if matches.get(header_name):
            indexes[key] = matches[header_name][0]
        else:
            missing.append(header_name)

    for key, (header_name, position) in (duplicate_requirements or {}).items():
        header_indexes = matches.get(header_name, [])
        if len(header_indexes) > position:
            indexes[key] = header_indexes[position]
        else:
            missing.append(f"{header_name}[{position}]")

    if missing:
        raise ValueError(f"源表缺少必要字段：{missing}")
    return indexes


def load_sheet_context(input_path: Path):
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = detect_data_sheet(workbook)
    header_row_1 = [
        normalize_text(value)
        for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    header_row_2 = [
        normalize_text(value)
        for value in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    ]
    data_start_row = 3 if sum(has_chinese(value) for value in header_row_2) > 0 else 2
    return workbook, sheet, header_row_1, header_row_2, data_start_row


def dense_rank_desc(items: list[tuple[str, float | int]]) -> dict[str, int]:
    rank_map: dict[str, int] = {}
    current_rank = 0
    last_value = None
    for key, value in items:
        if last_value is None or value != last_value:
            current_rank += 1
            last_value = value
        rank_map[key] = current_rank
    return rank_map


def style_two_header_sheet(
    sheet,
    columns_en: list[str],
    columns_cn: list[str],
    *,
    date_columns: set[str] | None = None,
    percent_columns: set[str] | None = None,
    integer_columns: set[str] | None = None,
    decimal_formats: dict[str, str] | None = None,
) -> None:
    date_columns = date_columns or set()
    percent_columns = percent_columns or set()
    integer_columns = integer_columns or set()
    decimal_formats = decimal_formats or {}

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_num in (1, 2):
        for col_num in range(1, len(columns_en) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    column_index = {column: idx + 1 for idx, column in enumerate(columns_en)}
    for row_num in range(3, sheet.max_row + 1):
        for column in columns_en:
            cell = sheet.cell(row=row_num, column=column_index[column])
            cell.alignment = center
            if column in integer_columns:
                cell.number_format = "0"
            elif column in percent_columns:
                cell.number_format = "0.00%"
            elif column in decimal_formats:
                cell.number_format = decimal_formats[column]
            elif column in date_columns and cell.value is not None:
                cell.number_format = "yyyy-mm-dd"

    preview_rows = min(sheet.max_row, 60)
    for idx, column in enumerate(columns_en, start=1):
        sample_lengths = [len(str(column)), len(str(columns_cn[idx - 1]))]
        for row_num in range(3, preview_rows + 1):
            value = sheet.cell(row=row_num, column=idx).value
            if value is not None:
                sample_lengths.append(len(str(value)))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(sample_lengths) + 2, 36)

    sheet.freeze_panes = "A3"


def style_notes_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = left

    for idx, column_cells in enumerate(sheet.columns, start=1):
        width = max(len(str(cell.value or "")) for cell in column_cells[: min(sheet.max_row, 60)])
        sheet.column_dimensions[get_column_letter(idx)].width = min(width + 2, 48)


def write_notes_rows(sheet, rows: list[tuple[object, ...]]) -> None:
    for row in rows:
        sheet.append(list(row))
    style_notes_sheet(sheet)
