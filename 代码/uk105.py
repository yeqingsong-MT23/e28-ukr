# -*- coding: utf-8 -*-
"""
根据 UK100_工作量 KPI 基础表生成 UK105_UEA处理人邮件处理时效分析.xlsx。
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

from uk_export_utils import build_default_output_path
from uk_workload_utils import (
    find_latest_input,
    get_required_indexes,
    load_sheet_context,
    normalize_text,
    round_half_up,
    safe_div,
    style_two_header_sheet,
    write_notes_rows,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK105_UEA处理人邮件处理时效分析.xlsx"
RESULT_SHEET_NAME = "UK105_Result"
NOTES_SHEET_NAME = "Notes"
EMAIL_TYPE = "邮件"
ALL_VALUE = "All"

COLUMNS_EN = ["UEA", "Metric_Code", "Metric_Name", "Email_Count", "Total_Handle_Time_Min", "Avg_Handle_Time_Min"]
COLUMNS_CN = ["处理人", "邮件类型编码", "邮件节点名称", "处理数量", "总处理时长（分钟）", "平均处理时长（分钟）"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 UK100 基础表生成 UK105 邮件处理时效分析。")
    parser.add_argument("--input", dest="input_path", help="指定输入的 UK100 基础表路径")
    parser.add_argument("--output", dest="output_path", help="指定输出 Excel 路径")
    return parser.parse_args()


def _parse_handle_minutes(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)

    text = normalize_text(value)
    if not text:
        return None

    normalized = text.replace(",", "")
    try:
        return float(normalized)
    except (TypeError, ValueError):
        return None


def build_summary(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook, sheet, header_row_1, _, data_start_row = load_sheet_context(input_path)
    try:
        indexes = get_required_indexes(
            header_row_1,
            {
                "Type": "Type",
                "UEA": "UEA",
                "Metric_Code": "Metric_Code",
                "Metric_Name": "Metric_Name",
                "Handle_Time_Minutes": "Handle_Time_Minutes",
            },
        )

        grouped: defaultdict[tuple[str, str, str], dict[str, float | int]] = defaultdict(
            lambda: {"Email_Count": 0, "Total_Handle_Time_Min": 0.0}
        )

        loaded_rows = 0
        email_rows = 0
        skipped_missing_uea_rows = 0
        skipped_missing_metric_code_rows = 0
        skipped_missing_metric_name_rows = 0
        skipped_bad_handle_time_rows = 0

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            loaded_rows += 1

            row_type = normalize_text(row[indexes["Type"]])
            if row_type != EMAIL_TYPE:
                continue

            email_rows += 1

            uea = normalize_text(row[indexes["UEA"]])
            if not uea:
                skipped_missing_uea_rows += 1
                continue

            metric_code = normalize_text(row[indexes["Metric_Code"]])
            if not metric_code:
                skipped_missing_metric_code_rows += 1
                continue

            metric_name = normalize_text(row[indexes["Metric_Name"]])
            if not metric_name:
                skipped_missing_metric_name_rows += 1
                continue

            handle_minutes = _parse_handle_minutes(row[indexes["Handle_Time_Minutes"]])
            if handle_minutes is None:
                skipped_bad_handle_time_rows += 1
                continue

            for key_uea in (ALL_VALUE, uea):
                key = (metric_code, metric_name, key_uea)
                grouped[key]["Email_Count"] += 1
                grouped[key]["Total_Handle_Time_Min"] += handle_minutes

        rows: list[dict[str, object]] = []
        for metric_code, metric_name, uea in sorted(
            grouped.keys(),
            key=lambda item: (item[0], item[1], 0 if item[2] == ALL_VALUE else 1, item[2]),
        ):
            email_count = int(grouped[(metric_code, metric_name, uea)]["Email_Count"])
            total_handle = int(round_half_up(grouped[(metric_code, metric_name, uea)]["Total_Handle_Time_Min"], 0))
            avg_handle = round_half_up(safe_div(total_handle, email_count), 2) if email_count else 0.0
            rows.append(
                {
                    "UEA": uea,
                    "Metric_Code": metric_code,
                    "Metric_Name": metric_name,
                    "Email_Count": email_count,
                    "Total_Handle_Time_Min": total_handle,
                    "Avg_Handle_Time_Min": avg_handle,
                }
            )

        stats = {
            "Input_File": str(input_path),
            "Loaded_Rows": loaded_rows,
            "Email_Rows": email_rows,
            "Skipped_Missing_UEA_Rows": skipped_missing_uea_rows,
            "Skipped_Missing_Metric_Code_Rows": skipped_missing_metric_code_rows,
            "Skipped_Missing_Metric_Name_Rows": skipped_missing_metric_name_rows,
            "Skipped_Bad_Handle_Time_Rows": skipped_bad_handle_time_rows,
            "Output_Row_Count": len(rows),
            "All_Row_Count": sum(1 for row in rows if row["UEA"] == ALL_VALUE),
        }
        return rows, stats
    finally:
        workbook.close()


def build_notes_rows(input_path: Path, output_path: Path, stats: dict[str, object]) -> list[tuple[object, ...]]:
    return [
        ("Field", "中文名", "解释和含义"),
        ("UEA", "处理人", "邮件处理跟进人；同一 Metric_Code + Metric_Name 下会同时输出 UEA = All 的汇总行"),
        ("Metric_Code", "邮件类型编码", "邮件节点或类型的分类编码"),
        ("Metric_Name", "邮件节点名称", "与 Metric_Code 对应的节点名称"),
        ("Email_Count", "处理数量", "仅统计 Type=邮件 的记录数量；按行为记录计数，不按 Ticket_No 去重"),
        ("Total_Handle_Time_Min", "总处理时长（分钟）", "处理数量对应记录的 Handle_Time_Minutes 合计，单位为分钟"),
        ("Avg_Handle_Time_Min", "平均处理时长（分钟）", "Avg_Handle_Time_Min = Total_Handle_Time_Min / Email_Count，保留两位小数"),
        (None, None, None),
        ("口径说明", None, None),
        ("基础过滤条件", "—", "仅统计 Type = 邮件"),
        ("输出粒度", "—", "UEA + Metric_Code + Metric_Name"),
        ("All 行含义", "—", "同一 Metric_Code + Metric_Name 下，脚本会同时输出 UEA = All 的整体汇总行"),
        ("All 行排序", "—", "All 行在每个节点组内排序靠前"),
        ("计数口径", "—", "Email_Count 按行为记录计数，不对 Ticket_No 去重"),
        ("时长异常处理", "—", "Handle_Time_Minutes 为空、非数值或无法解析时，当前记录跳过"),
        ("数值格式说明", "—", "平均处理时长保留两位小数，其余数字为整数"),
        (None, None, None),
        ("运行元信息", None, None),
        ("Input_File", "输入文件", str(input_path)),
        ("Output_File", "输出文件", str(output_path)),
        ("Loaded_Rows", "加载行数", stats.get("Loaded_Rows", 0)),
        ("Email_Rows", "邮件记录行数", stats.get("Email_Rows", 0)),
        ("Skipped_Missing_UEA_Rows", "跳过-处理人为空行数", stats.get("Skipped_Missing_UEA_Rows", 0)),
        ("Skipped_Missing_Metric_Code_Rows", "跳过-邮件类型编码为空行数", stats.get("Skipped_Missing_Metric_Code_Rows", 0)),
        ("Skipped_Missing_Metric_Name_Rows", "跳过-邮件节点名称为空行数", stats.get("Skipped_Missing_Metric_Name_Rows", 0)),
        ("Skipped_Bad_Handle_Time_Rows", "跳过-处理时长异常行数", stats.get("Skipped_Bad_Handle_Time_Rows", 0)),
        ("Output_Row_Count", "输出结果行数", stats.get("Output_Row_Count", 0)),
        ("All_Row_Count", "All 汇总行数", stats.get("All_Row_Count", 0)),
    ]


def write_workbook(output_path: Path, rows: list[dict[str, object]], notes_rows: list[tuple[object, ...]]) -> None:
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = RESULT_SHEET_NAME
        sheet.append(COLUMNS_EN)
        sheet.append(COLUMNS_CN)

        for row in rows:
            sheet.append([row.get(column) for column in COLUMNS_EN])

        style_two_header_sheet(
            sheet,
            COLUMNS_EN,
            COLUMNS_CN,
            integer_columns={"Email_Count", "Total_Handle_Time_Min"},
            decimal_formats={"Avg_Handle_Time_Min": "0.00"},
        )

        notes_sheet = workbook.create_sheet(NOTES_SHEET_NAME)
        write_notes_rows(notes_sheet, notes_rows)

        workbook.save(output_path)
    finally:
        workbook.close()


def main() -> None:
    args = parse_args()

    input_path = Path(args.input_path).expanduser() if args.input_path else find_latest_input(BASE_DIR)
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    output_path = (
        Path(args.output_path).expanduser()
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    rows, stats = build_summary(input_path)
    notes_rows = build_notes_rows(input_path, output_path, stats)
    write_workbook(output_path, rows, notes_rows)

    print("结果已生成：")
    print(output_path)


if __name__ == "__main__":
    main()
