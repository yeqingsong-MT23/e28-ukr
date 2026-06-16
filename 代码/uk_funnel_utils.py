# -*- coding: utf-8 -*-
from __future__ import annotations

import datetime as dt
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUCCESS_STATUS = {
    "End",
    "WaitingRPO",
    "WaitingPayment",
    "WaitingDeal",
    "OnDeal",
    "WaitingPO",
}

IN_PROGRESS_STATUS = {
    "Hold",
    "New",
    "SubmitRBS",
}

WOOT_CANCEL_STATUS = {
    "CSCancel",
    "Long time-Cancel",
    "WootCancel",
}

SELLER_CANCEL_STATUS = {
    "SellerCancel",
}


def parse_cell_date(value) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(".", "-").replace("/", "-")
    for fmt in (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y%m%d",
    ):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def has_chinese(text: object) -> bool:
    content = str(text or "")
    return any("\u4e00" <= char <= "\u9fff" for char in content)


def has_value(value) -> bool:
    return value is not None and str(value).strip() != ""


def safe_div(numerator: float | int, denominator: float | int) -> float:
    return numerator / denominator if denominator else 0.0


def find_latest_input(base_dir: Path) -> Path:
    files = sorted(
        (
            path
            for path in base_dir.glob("UK200_转化漏斗KPI-*.xlsx")
            if not path.name.startswith("~$")
        ),
        key=lambda item: (item.stat().st_mtime, item.name),
        reverse=True,
    )
    if not files:
        raise FileNotFoundError(
            f"未找到 UK200 基础表，请确认目录下存在 UK200_转化漏斗KPI-*.xlsx：{base_dir}"
        )
    return files[0]


def detect_data_sheet(workbook):
    if "Data" in workbook.sheetnames:
        return workbook["Data"]
    return workbook[workbook.sheetnames[0]]


def get_header_matches(header_row: list[str]) -> dict[str, list[int]]:
    matches: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(header_row):
        matches[normalize_text(header)].append(index)
    return matches


def get_required_indexes(
    header_row: list[str],
    required_single: dict[str, str],
    duplicate_requirements: dict[str, tuple[str, int]] | None = None,
) -> dict[str, int]:
    matches = get_header_matches(header_row)
    indexes: dict[str, int] = {}
    missing: list[str] = []

    for key, header_name in required_single.items():
        if matches.get(header_name):
            indexes[key] = matches[header_name][0]
        else:
            missing.append(header_name)

    for key, (header_name, position) in (duplicate_requirements or {}).items():
        header_indexes = matches.get(header_name, [])
        if len(header_indexes) > position:
            indexes[key] = header_indexes[position]
        else:
            missing.append(f"{header_name}[{position}]")

    if missing:
        raise ValueError(f"源表缺少必要字段：{missing}")
    return indexes


def month_sort_key(month: str) -> tuple[int, int]:
    if month == "ALL":
        return (9999, 99)
    separator = "" if "-" not in month else "-"
    if separator:
        year, mon = month.split("-")
    else:
        year, mon = month[:4], month[4:]
    return (int(year), int(mon))


def style_two_header_sheet(
    sheet,
    columns_en: list[str],
    columns_cn: list[str],
    *,
    date_columns: set[str] | None = None,
) -> None:
    date_columns = date_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_num in (1, 2):
        for col_num in range(1, len(columns_en) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    count_columns = set()
    rate_columns = set()
    for column in columns_en:
        if any(
            token in column
            for token in ("_Count", "_QPO", "_QPOs", "Submitted_", "Total_", "Success_", "Cancel_", "InProgress_")
        ):
            if column.endswith("_Rate") or "Rate" in column or "Share" in column:
                rate_columns.add(column)
            elif column not in date_columns:
                count_columns.add(column)
        if "Rate" in column or "Share" in column:
            rate_columns.add(column)

    column_index = {column: idx + 1 for idx, column in enumerate(columns_en)}

    for row_num in range(3, sheet.max_row + 1):
        for column in columns_en:
            cell = sheet.cell(row=row_num, column=column_index[column])
            cell.alignment = center
            if column in count_columns:
                cell.number_format = "0"
            elif column in rate_columns:
                cell.number_format = "0.00%"
            elif column in date_columns and cell.value is not None:
                cell.number_format = "yyyy-mm-dd"

    preview_rows = min(sheet.max_row, 30)
    for idx, column in enumerate(columns_en, start=1):
        sample_lengths = [len(str(column)), len(str(columns_cn[idx - 1]))]
        for row_num in range(3, preview_rows + 1):
            value = sheet.cell(row=row_num, column=idx).value
            if value is not None:
                sample_lengths.append(len(str(value)))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(sample_lengths) + 2, 36)

    sheet.freeze_panes = "A3"


def map_simple_type(status: str) -> str | None:
    if status in WOOT_CANCEL_STATUS or status in SELLER_CANCEL_STATUS:
        return "Cancel"
    if status in SUCCESS_STATUS:
        return "Success"
    if status in IN_PROGRESS_STATUS:
        return "InProgress"
    return None


def map_split_type(status: str) -> str:
    if status in WOOT_CANCEL_STATUS:
        return "WootCancel"
    if status in SELLER_CANCEL_STATUS:
        return "SellerCancel"
    if status in SUCCESS_STATUS:
        return "Success"
    if status in IN_PROGRESS_STATUS:
        return "InProgress"
    return "Other"


def normalize_order_level(value: str) -> str:
    text = normalize_text(value)
    if text == "AA":
        return "A"
    return text


def load_stage_qpo_records(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    sheet = detect_data_sheet(workbook)

    header_row_1 = [
        normalize_text(value)
        for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    header_row_2 = [
        normalize_text(value)
        for value in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    ]
    data_start_row = 3 if sum(has_chinese(value) for value in header_row_2) > 0 else 2

    indexes = get_required_indexes(
        header_row_1,
        {
            "Submission_Date": "Submission_Date",
            "QPO_ID": "QPO_ID",
            "QPO_Status": "QPO_Status",
            "Order_Level": "Order_Level",
            "Pricing_Mode": "Pricing_Mode",
            "Initial_Quote": "Initial_Quote",
            "Second_Quote": "Second_Quote",
            "MVM_Approved_Price": "MVM_Approved_Price",
        },
    )

    qpo_state: dict[str, dict[str, object]] = {}
    input_rows = 0
    valid_rows = 0

    for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
        input_rows += 1
        submission_date = parse_cell_date(row[indexes["Submission_Date"]])
        if submission_date is None:
            continue
        valid_rows += 1

        qpo_id = normalize_text(row[indexes["QPO_ID"]])
        if not qpo_id:
            continue

        state = qpo_state.setdefault(
            qpo_id,
            {
                "submission_date": submission_date,
                "statuses": set(),
                "order_level": "",
                "pricing_mode": "",
                "has_initial": False,
                "has_second": False,
                "has_third": False,
            },
        )

        if submission_date < state["submission_date"]:
            state["submission_date"] = submission_date

        status = normalize_text(row[indexes["QPO_Status"]])
        state["statuses"].add(status)

        if not state["order_level"]:
            state["order_level"] = normalize_order_level(row[indexes["Order_Level"]])
        if not state["pricing_mode"]:
            state["pricing_mode"] = normalize_text(row[indexes["Pricing_Mode"]])

        state["has_initial"] = state["has_initial"] or has_value(row[indexes["Initial_Quote"]])
        state["has_second"] = state["has_second"] or has_value(row[indexes["Second_Quote"]])
        state["has_third"] = state["has_third"] or has_value(row[indexes["MVM_Approved_Price"]])

    excluded_qpo_ids: list[str] = []
    records: list[dict[str, object]] = []
    for qpo_id, state in qpo_state.items():
        split_types = {map_split_type(status) for status in state["statuses"]}
        if "WootCancel" in split_types:
            split_type = "WootCancel"
        elif "SellerCancel" in split_types:
            split_type = "SellerCancel"
        elif "Success" in split_types:
            split_type = "Success"
        elif "InProgress" in split_types:
            split_type = "InProgress"
        else:
            split_type = "Other"

        has_initial = bool(state["has_initial"])
        has_second = bool(state["has_second"])
        has_third = bool(state["has_third"])
        if split_type == "Success" and not (has_initial or has_second or has_third):
            excluded_qpo_ids.append(qpo_id)
            continue

        stage = "Third" if has_third else "Second" if has_second else "Initial"
        submission_date = state["submission_date"]
        records.append(
            {
                "qpo_id": qpo_id,
                "submission_date": submission_date,
                "month_dash": submission_date.strftime("%Y-%m"),
                "month_compact": submission_date.strftime("%Y%m"),
                "order_level": state["order_level"],
                "pricing_mode": state["pricing_mode"],
                "split_type": split_type,
                "simple_type": (
                    "Cancel" if split_type in {"WootCancel", "SellerCancel"} else split_type
                ),
                "stage": stage,
            }
        )

    metadata = {
        "input_rows": input_rows,
        "valid_rows": valid_rows,
        "excluded_qpo_ids": excluded_qpo_ids,
        "included_qpo_count": len(records),
    }
    return records, metadata


def stage_group_metrics(records: list[dict[str, object]]) -> dict[str, int | float]:
    submitted = len(records)
    success = sum(1 for record in records if record["simple_type"] == "Success")
    in_progress = sum(1 for record in records if record["simple_type"] == "InProgress")
    cancel = sum(1 for record in records if record["simple_type"] == "Cancel")
    return {
        "Submitted_QPO_Count": submitted,
        "Success_QPO_Count": success,
        "In_Progress_QPO_Count": in_progress,
        "Cancel_QPO_Count": cancel,
        "Success_Rate": safe_div(success, submitted),
        "In_Progress_Rate": safe_div(in_progress, submitted),
        "Cancel_Rate": safe_div(cancel, submitted),
    }


def build_stage_breakdown(
    records: list[dict[str, object]],
    *,
    include_stage_inprogress: bool,
    include_cancel_split: bool,
) -> dict[str, object]:
    result: dict[str, object] = {}
    stage_names = ("Initial", "Second", "Third")

    for stage_name in stage_names:
        stage_records = [record for record in records if record["stage"] == stage_name]
        stage_count = len(stage_records)
        success_count = sum(1 for record in stage_records if record["simple_type"] == "Success")
        cancel_count = sum(1 for record in stage_records if record["simple_type"] == "Cancel")
        in_progress_count = sum(
            1 for record in stage_records if record["simple_type"] == "InProgress"
        )
        result[f"{stage_name}_Stage_QPO_Count"] = stage_count
        result[f"{stage_name}_Stage_Success_QPO_Count"] = success_count
        result[f"{stage_name}_Stage_Success_Submit_Rate"] = safe_div(success_count, len(records))
        result[f"{stage_name}_Stage_Cancel_QPO_Count"] = cancel_count
        result[f"{stage_name}_Stage_Cancel_Submit_Rate"] = safe_div(cancel_count, len(records))

        if include_stage_inprogress:
            result[f"{stage_name}_Stage_In_Progress_QPO_Count"] = in_progress_count
            result[f"{stage_name}_Stage_In_Progress_Submit_Rate"] = safe_div(
                in_progress_count, len(records)
            )

        if include_cancel_split:
            woot_count = sum(1 for record in stage_records if record["split_type"] == "WootCancel")
            seller_count = sum(
                1 for record in stage_records if record["split_type"] == "SellerCancel"
            )
            result[f"{stage_name}_Stage_WootCancel_QPO_Count"] = woot_count
            result[f"{stage_name}_Stage_SellerCancel_QPO_Count"] = seller_count
            result[f"{stage_name}_Stage_WootCancel_Stage_Rate"] = safe_div(woot_count, stage_count)
            result[f"{stage_name}_Stage_WootCancel_Cancel_Rate"] = safe_div(
                woot_count, cancel_count
            )
            result[f"{stage_name}_Stage_SellerCancel_Stage_Rate"] = safe_div(
                seller_count, stage_count
            )
            result[f"{stage_name}_Stage_SellerCancel_Cancel_Rate"] = safe_div(
                seller_count, cancel_count
            )

    return result
