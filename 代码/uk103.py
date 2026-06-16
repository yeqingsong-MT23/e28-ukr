# -*- coding: utf-8 -*-
"""
根据 UK100_工作量 KPI 基础表生成 UK103_工单节点处理量分析表.xlsx。
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

from uk_export_utils import build_default_output_path
from uk_workload_utils import (
    department_sort_key,
    find_latest_input,
    get_required_indexes,
    load_sheet_context,
    month_sort_key,
    normalize_text,
    parse_cell_date,
    safe_div,
    style_two_header_sheet,
    write_notes_rows,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK103_工单节点处理量分析表.xlsx"
RESULT_SHEET_NAME = "UK103"
NOTES_SHEET_NAME = "Notes"
TICKET_ONLY_TYPE = "工单"
ALL_DEPARTMENT = "ALL"

COLUMNS_EN = [
    "Department",
    "Month",
    "Metric_Name",
    "Ticket_Unique_Cnt",
    "Ticket_Delay_Cnt",
    "Ticket_Close_Cnt",
    "Ticket_Comment_Cnt_CrossDept",
    "Ticket_Comment_Cnt_InDept",
    "Ticket_Note_Submit_Cnt",
    "Avg_Delay_per_Ticket",
    "Avg_CrossDept_Comment_per_Ticket",
    "Avg_InDept_Comment_per_Ticket",
    "Avg_Note_per_Ticket",
    "Avg_Close_per_Ticket",
]

COLUMNS_CN = [
    "部门",
    "月份",
    "节点名称",
    "工单数（去重）",
    "延期次数",
    "完成/取消次数",
    "跨部门评论次数",
    "本部门评论次数",
    "备注次数",
    "平均延期/工单",
    "平均跨部门评论/工单",
    "平均本部门评论/工单",
    "平均备注/工单",
    "平均完成/取消/工单",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 UK100 基础表生成 UK103 工单节点处理量分析表。")
    parser.add_argument("--input", dest="input_path", help="输入基础表路径；未传时自动识别最新 UK100 基础表")
    parser.add_argument("--output", dest="output_path", help="输出 Excel 路径；未传时输出到默认目录并自动追加时间戳")
    return parser.parse_args()


def to_int(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = normalize_text(value)
    if not text:
        return 0
    text = text.replace(",", "")
    try:
        return int(text)
    except ValueError:
        try:
            return int(float(text))
        except ValueError:
            return 0


def build_summary(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = None
    try:
        workbook, sheet, header_row_1, _, data_start_row = load_sheet_context(input_path)
        indexes = get_required_indexes(
            header_row_1,
            {
                "Stat_Date": "Stat_Date",
                "Department": "Department",
                "Type": "Type",
                "Metric_Name": "Metric_Name",
                "Ticket_No": "Ticket_No",
                "TAM1_Cnt": "TAM1_Cnt",
                "TAM2_Cnt": "TAM2_Cnt",
                "TAM3_Cnt": "TAM3_Cnt",
                "TAM4_Cnt": "TAM4_Cnt",
                "TAM5_Cnt": "TAM5_Cnt",
            },
        )

        grouped = defaultdict(
            lambda: {
                "tickets": set(),
                "Ticket_Delay_Cnt": 0,
                "Ticket_Close_Cnt": 0,
                "Ticket_Comment_Cnt_CrossDept": 0,
                "Ticket_Comment_Cnt_InDept": 0,
                "Ticket_Note_Submit_Cnt": 0,
            }
        )

        stats: dict[str, object] = {
            "Loaded_Rows": 0,
            "Ticket_Rows": 0,
            "Valid_Ticket_Rows": 0,
            "Skipped_Non_Ticket_Rows": 0,
            "Skipped_Missing_Stat_Date_Rows": 0,
            "Skipped_Missing_Department_Rows": 0,
            "Skipped_Missing_Metric_Name_Rows": 0,
            "Skipped_Missing_Ticket_No_Rows": 0,
            "Output_Row_Count": 0,
            "All_Row_Count": 0,
            "All_Generated": "No",
        }

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            stats["Loaded_Rows"] = int(stats["Loaded_Rows"]) + 1

            row_type = normalize_text(row[indexes["Type"]])
            if row_type != TICKET_ONLY_TYPE:
                stats["Skipped_Non_Ticket_Rows"] = int(stats["Skipped_Non_Ticket_Rows"]) + 1
                continue

            stats["Ticket_Rows"] = int(stats["Ticket_Rows"]) + 1

            stat_date = parse_cell_date(row[indexes["Stat_Date"]])
            if stat_date is None:
                stats["Skipped_Missing_Stat_Date_Rows"] = int(stats["Skipped_Missing_Stat_Date_Rows"]) + 1
                continue

            department = normalize_text(row[indexes["Department"]])
            if not department:
                stats["Skipped_Missing_Department_Rows"] = int(stats["Skipped_Missing_Department_Rows"]) + 1
                continue

            metric_name = normalize_text(row[indexes["Metric_Name"]])
            if not metric_name:
                stats["Skipped_Missing_Metric_Name_Rows"] = int(stats["Skipped_Missing_Metric_Name_Rows"]) + 1
                continue

            ticket_no = normalize_text(row[indexes["Ticket_No"]])
            if not ticket_no:
                stats["Skipped_Missing_Ticket_No_Rows"] = int(stats["Skipped_Missing_Ticket_No_Rows"]) + 1
                continue

            month = stat_date.strftime("%Y-%m")
            delay_cnt = to_int(row[indexes["TAM4_Cnt"]])
            close_cnt = to_int(row[indexes["TAM5_Cnt"]])
            cross_comment_cnt = to_int(row[indexes["TAM1_Cnt"]])
            indept_comment_cnt = to_int(row[indexes["TAM3_Cnt"]])
            note_cnt = to_int(row[indexes["TAM2_Cnt"]])

            for dept_key in (department, ALL_DEPARTMENT):
                key = (dept_key, month, metric_name)
                grouped[key]["tickets"].add(ticket_no)
                grouped[key]["Ticket_Delay_Cnt"] += delay_cnt
                grouped[key]["Ticket_Close_Cnt"] += close_cnt
                grouped[key]["Ticket_Comment_Cnt_CrossDept"] += cross_comment_cnt
                grouped[key]["Ticket_Comment_Cnt_InDept"] += indept_comment_cnt
                grouped[key]["Ticket_Note_Submit_Cnt"] += note_cnt

            stats["Valid_Ticket_Rows"] = int(stats["Valid_Ticket_Rows"]) + 1

        rows: list[dict[str, object]] = []
        for key in sorted(
            grouped,
            key=lambda item: (department_sort_key(item[0]), month_sort_key(item[1]), item[2]),
        ):
            values = grouped[key]
            ticket_unique_cnt = len(values["tickets"])
            rows.append(
                {
                    "Department": key[0],
                    "Month": key[1],
                    "Metric_Name": key[2],
                    "Ticket_Unique_Cnt": ticket_unique_cnt,
                    "Ticket_Delay_Cnt": values["Ticket_Delay_Cnt"],
                    "Ticket_Close_Cnt": values["Ticket_Close_Cnt"],
                    "Ticket_Comment_Cnt_CrossDept": values["Ticket_Comment_Cnt_CrossDept"],
                    "Ticket_Comment_Cnt_InDept": values["Ticket_Comment_Cnt_InDept"],
                    "Ticket_Note_Submit_Cnt": values["Ticket_Note_Submit_Cnt"],
                    "Avg_Delay_per_Ticket": safe_div(values["Ticket_Delay_Cnt"], ticket_unique_cnt),
                    "Avg_CrossDept_Comment_per_Ticket": safe_div(values["Ticket_Comment_Cnt_CrossDept"], ticket_unique_cnt),
                    "Avg_InDept_Comment_per_Ticket": safe_div(values["Ticket_Comment_Cnt_InDept"], ticket_unique_cnt),
                    "Avg_Note_per_Ticket": safe_div(values["Ticket_Note_Submit_Cnt"], ticket_unique_cnt),
                    "Avg_Close_per_Ticket": safe_div(values["Ticket_Close_Cnt"], ticket_unique_cnt),
                }
            )

        stats["Output_Row_Count"] = len(rows)
        stats["All_Row_Count"] = sum(1 for row in rows if row.get("Department") == ALL_DEPARTMENT)
        stats["All_Generated"] = "Yes" if int(stats["All_Row_Count"]) > 0 else "No"

        return rows, stats
    finally:
        if workbook is not None:
            workbook.close()


def build_notes_rows(input_path: Path, output_path: Path, stats: dict[str, object]) -> list[tuple[object, ...]]:
    return [
        ("Field EN", "字段中文", "计算口径", "业务解读"),
        ("Department", "部门", "所属部门；ALL=正式汇总行", "用于横向对比不同部门与全公司节点处理量；ALL 不是备注行"),
        ("Month", "月份", "从 Stat_Date 提取 yyyy-mm", "用于月度趋势分析"),
        ("Metric_Name", "节点名称", "工单节点名称（Metric_Name）", "区分不同节点/环节的工作量"),
        ("Ticket_Unique_Cnt", "工单数（去重）", "按 Ticket_No 去重计数", "衡量该节点覆盖的工单规模"),
        ("Ticket_Delay_Cnt", "延期次数", "sum(TAM4_Cnt)", "反映延期发生频次"),
        ("Ticket_Close_Cnt", "完成/取消次数", "sum(TAM5_Cnt)", "反映完成/取消动作量"),
        ("Ticket_Comment_Cnt_CrossDept", "跨部门评论次数", "sum(TAM1_Cnt)", "反映跨部门协作沟通强度"),
        ("Ticket_Comment_Cnt_InDept", "本部门评论次数", "sum(TAM3_Cnt)", "反映部门内部沟通强度"),
        ("Ticket_Note_Submit_Cnt", "备注次数", "sum(TAM2_Cnt)", "反映补充信息与记录强度"),
        ("Avg_Delay_per_Ticket", "平均延期/工单", "Ticket_Delay_Cnt ÷ Ticket_Unique_Cnt", "每单平均延期强度"),
        ("Avg_CrossDept_Comment_per_Ticket", "平均跨部门评论/工单", "Ticket_Comment_Cnt_CrossDept ÷ Ticket_Unique_Cnt", "每单跨部门沟通强度"),
        ("Avg_InDept_Comment_per_Ticket", "平均本部门评论/工单", "Ticket_Comment_Cnt_InDept ÷ Ticket_Unique_Cnt", "每单内部沟通强度"),
        ("Avg_Note_per_Ticket", "平均备注/工单", "Ticket_Note_Submit_Cnt ÷ Ticket_Unique_Cnt", "每单备注强度"),
        ("Avg_Close_per_Ticket", "平均完成/取消/工单", "Ticket_Close_Cnt ÷ Ticket_Unique_Cnt", "每单关闭动作强度"),
        (None, None, None, None),
        ("筛选规则", "口径规则", "仅保留 Type=工单；Metric_Name 和 Ticket_No 缺失记录不参与统计", "确保口径聚焦工单节点工作量"),
        ("动作映射", "TAM 映射", "TAM1=跨部门评论；TAM2=备注；TAM3=本部门评论；TAM4=延期；TAM5=完成/取消", "主表所有动作指标均按该映射聚合"),
        ("平均口径", "分母规则", "所有平均类指标均使用 动作总数 ÷ 去重工单数", "避免平均每单类指标分母漂移"),
        (None, None, None, None),
        ("Run Meta", "运行元信息", "说明", "取值"),
        ("Input_File", "输入文件", "本次读取的基础表文件", str(input_path)),
        ("Output_File", "输出文件", "本次生成的 Excel 文件路径", str(output_path)),
        ("Loaded_Rows", "加载行数", "基础表数据区总行数", stats.get("Loaded_Rows", 0)),
        ("Ticket_Rows", "工单行数", "Type=工单 的原始行数", stats.get("Ticket_Rows", 0)),
        ("Valid_Ticket_Rows", "有效工单行数", "通过关键字段校验并参与汇总的行数", stats.get("Valid_Ticket_Rows", 0)),
        ("Skipped_Non_Ticket_Rows", "剔除非工单行数", "Type 不等于 工单 的行数", stats.get("Skipped_Non_Ticket_Rows", 0)),
        ("Skipped_Missing_Stat_Date_Rows", "缺失日期剔除行数", "Type=工单 但 Stat_Date 无法识别的行数", stats.get("Skipped_Missing_Stat_Date_Rows", 0)),
        ("Skipped_Missing_Department_Rows", "缺失部门剔除行数", "Type=工单 但 Department 缺失的行数", stats.get("Skipped_Missing_Department_Rows", 0)),
        ("Skipped_Missing_Metric_Name_Rows", "缺失节点名剔除行数", "Type=工单 但 Metric_Name 缺失的行数", stats.get("Skipped_Missing_Metric_Name_Rows", 0)),
        ("Skipped_Missing_Ticket_No_Rows", "缺失工单号剔除行数", "Type=工单 但 Ticket_No 缺失的行数", stats.get("Skipped_Missing_Ticket_No_Rows", 0)),
        ("Output_Row_Count", "输出结果行数", "主表汇总结果总行数", stats.get("Output_Row_Count", 0)),
        ("All_Row_Count", "ALL 汇总行数", "Department=ALL 的正式汇总行数量", stats.get("All_Row_Count", 0)),
        ("All_Generated", "是否生成 ALL 汇总", "Yes=已生成正式 ALL 汇总；No=无有效数据或无汇总行", stats.get("All_Generated", "No")),
        (None, None, None, None),
        ("合规声明", None, None, "Instruction 中涉及的 UK + 数字均为企业内部编号。本 GPT 不用于规避监管、生成违法或受限内容、不处理政治、金融、医疗或法律合规敏感事务，仅用于企业内部数据整理与知识管理。"),
    ]


def write_workbook(output_path: Path, rows: list[dict[str, object]], notes_rows: list[tuple[object, ...]]) -> None:
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = RESULT_SHEET_NAME
        sheet.append(COLUMNS_EN)
        sheet.append(COLUMNS_CN)

        for row in rows:
            sheet.append([row.get(column) for column in COLUMNS_EN])

        style_two_header_sheet(
            sheet,
            COLUMNS_EN,
            COLUMNS_CN,
            integer_columns={
                "Ticket_Unique_Cnt",
                "Ticket_Delay_Cnt",
                "Ticket_Close_Cnt",
                "Ticket_Comment_Cnt_CrossDept",
                "Ticket_Comment_Cnt_InDept",
                "Ticket_Note_Submit_Cnt",
            },
            decimal_formats={
                "Avg_Delay_per_Ticket": "0.00",
                "Avg_CrossDept_Comment_per_Ticket": "0.00",
                "Avg_InDept_Comment_per_Ticket": "0.00",
                "Avg_Note_per_Ticket": "0.00",
                "Avg_Close_per_Ticket": "0.00",
            },
        )

        notes_sheet = workbook.create_sheet(NOTES_SHEET_NAME)
        write_notes_rows(notes_sheet, notes_rows)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def main() -> None:
    args = parse_args()

    if args.input_path:
        input_path = Path(args.input_path)
        if not input_path.exists():
            raise FileNotFoundError(f"未找到输入文件：{input_path}")
    else:
        input_path = find_latest_input(BASE_DIR)

    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )

    rows, stats = build_summary(input_path)
    notes_rows = build_notes_rows(input_path, output_path, stats)
    write_workbook(output_path, rows, notes_rows)

    print("结果已生成：")
    print(output_path)


if __name__ == "__main__":
    main()
