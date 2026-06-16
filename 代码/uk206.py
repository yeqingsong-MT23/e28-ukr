# -*- coding: utf-8 -*-
"""
根据 UK200_转化漏斗KPI 基础表生成 UK206｜各订单等级 × 报价阶段 转化漏斗.xlsx。
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from uk_export_utils import build_default_output_path
from uk_funnel_utils import (
    build_stage_breakdown,
    find_latest_input,
    load_stage_qpo_records,
    month_sort_key,
    safe_div,
    style_two_header_sheet,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK206｜各订单等级 × 报价阶段 转化漏斗.xlsx"
SHEET_NAME = "UK206_OrderLevel_QuoteStageFunnel"
NOTES_SHEET_NAME = "Notes"
REPORT_CODE = "UK206"
ORDER_LEVELS = ["A", "B", "C", "D"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据 UK200 基础表生成 UK206｜各订单等级 × 报价阶段 转化漏斗。"
    )
    parser.add_argument("--input", dest="input_path")
    parser.add_argument("--output", dest="output_path")
    return parser.parse_args()


def aggregate(records: list[dict[str, object]]) -> list[dict[str, object]]:
    grouped: dict[tuple[str, str], list[dict[str, object]]] = defaultdict(list)
    for record in records:
        if record["order_level"] in ORDER_LEVELS:
            grouped[(record["month_dash"], record["order_level"])].append(record)
            grouped[("ALL", record["order_level"])].append(record)
            grouped[("ALL", "ALL")].append(record)

    rows: list[dict[str, object]] = []
    months = sorted({record["month_dash"] for record in records}, key=month_sort_key)
    for month in months + ["ALL"]:
        levels = ORDER_LEVELS if month != "ALL" else ORDER_LEVELS + ["ALL"]
        for level in levels:
            bucket = grouped.get((month, level), [])
            if not bucket:
                continue
            submitted = len(bucket)
            success = sum(1 for record in bucket if record["simple_type"] == "Success")
            in_progress = sum(1 for record in bucket if record["simple_type"] == "InProgress")
            cancel = sum(1 for record in bucket if record["simple_type"] == "Cancel")
            row = {
                "Month": month,
                "Order_Level": level,
                "Submitted_QPO_Count": submitted,
                "Success_QPO_Count": success,
                "In_Progress_QPO_Count": in_progress,
                "Cancel_QPO_Count": cancel,
                "Success_Rate": safe_div(success, submitted),
                "In_Progress_Rate": safe_div(in_progress, submitted),
                "Cancel_Rate": safe_div(cancel, submitted),
            }
            row.update(
                build_stage_breakdown(
                    bucket,
                    include_stage_inprogress=False,
                    include_cancel_split=False,
                )
            )
            rows.append(row)
    return rows


def sanitize_note_value(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return ", ".join(f"{key}={sanitize_note_value(item)}" for key, item in value.items())
    if isinstance(value, (list, tuple, set)):
        return ", ".join(str(item) for item in value)
    return str(value)


def build_notes_rows(
    *,
    input_path: Path,
    output_path: Path,
    raw_records: list[dict[str, object]],
    summary_rows: list[dict[str, object]],
    metadata: dict[str, object],
) -> list[tuple[object, object, object]]:
    months = sorted({str(record["month_dash"]) for record in raw_records}, key=month_sort_key)
    if not months:
        month_range = "N/A"
    elif len(months) == 1:
        month_range = months[0]
    else:
        month_range = f"{months[0]} ~ {months[-1]}"

    notes_rows: list[tuple[object, object, object]] = [
        ("Report_Code", REPORT_CODE, "报表编码。"),
        ("Input_File", str(input_path), "本次运行使用的输入基础表。"),
        ("Output_File", str(output_path), "本次运行生成的输出文件。"),
        ("Main_Sheet", SHEET_NAME, "主结果 Sheet。"),
        ("Notes_Sheet", NOTES_SHEET_NAME, "口径说明与运行元信息 Sheet。"),
        ("Loaded_QPO_Count", len(raw_records), "load_stage_qpo_records 返回的 QPO 级记录数。"),
        (
            "Included_QPO_Count",
            metadata.get("included_qpo_count", len(raw_records)),
            "纳入统计的 QPO 数；仅保留订单等级 A/B/C/D，AA 并入 A。",
        ),
        (
            "Excluded_QPO_Count",
            len(raw_records) - int(metadata.get("included_qpo_count", len(raw_records))),
            "因订单等级为空或不在 A/B/C/D 范围内而被排除的 QPO 数。",
        ),
        ("Order_Level_List", ", ".join(ORDER_LEVELS), "固定输出顺序：A, B, C, D。"),
        ("Summary_Row_Count", len(summary_rows), "主表实际输出的数据行数，不含双表头。"),
        ("Month_Range", month_range, "Month 使用 month_dash 口径；主表另含 ALL 汇总行。"),
        (
            "Stage_Rule",
            "Third > Second > Initial",
            "阶段归属按最高进入阶段判定；由 Initial_Quote / Second_Quote / MVM_Approved_Price 是否存在推导。",
        ),
        (
            "Rate_Rule",
            "All rates denominator = Submitted_QPO_Count",
            "Success_Rate / In_Progress_Rate / Cancel_Rate 及各阶段 *_Submit_Rate 分母均为 Submitted_QPO_Count。",
        ),
        ("Month", "month_dash", "提报月份；主表包含各月份行及 ALL 汇总行。"),
        ("Order_Level", "A/B/C/D only; AA -> A", "订单等级仅保留 A/B/C/D；AA 会并入 A。"),
        ("Submitted_QPO_Count", "Count of included QPO", "当前 Month + Order_Level 分组的提报 QPO 数。"),
        ("Success_QPO_Count", "simple_type = Success", "当前分组内成交 QPO 数量。"),
        ("In_Progress_QPO_Count", "simple_type = InProgress", "当前分组内跟进中 QPO 数量。"),
        ("Cancel_QPO_Count", "simple_type = Cancel", "当前分组内取消 QPO 数量。"),
        ("Success_Rate", "Success_QPO_Count / Submitted_QPO_Count", "成交率；零分母时返回 0。"),
        ("In_Progress_Rate", "In_Progress_QPO_Count / Submitted_QPO_Count", "跟进中占比；零分母时返回 0。"),
        ("Cancel_Rate", "Cancel_QPO_Count / Submitted_QPO_Count", "取消率；零分母时返回 0。"),
    ]

    for key in sorted(metadata):
        notes_rows.append(
            (
                f"Meta_{key}",
                sanitize_note_value(metadata.get(key)),
                "load_stage_qpo_records 返回的附加元信息。",
            )
        )
    return notes_rows


def style_notes_sheet(sheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    alignment = Alignment(vertical="center", wrap_text=True)

    for row_num in (1, 2):
        for cell in sheet[row_num]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A3"
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 40
    sheet.column_dimensions["C"].width = 80


def write_workbook(
    rows: list[dict[str, object]],
    output_path: Path,
    *,
    notes_rows: list[tuple[object, object, object]] | None = None,
) -> None:
    columns_en = [
        "Month",
        "Order_Level",
        "Submitted_QPO_Count",
        "Success_QPO_Count",
        "In_Progress_QPO_Count",
        "Cancel_QPO_Count",
        "Success_Rate",
        "In_Progress_Rate",
        "Cancel_Rate",
        "Initial_Stage_QPO_Count",
        "Initial_Stage_Success_QPO_Count",
        "Initial_Stage_Success_Submit_Rate",
        "Initial_Stage_Cancel_QPO_Count",
        "Initial_Stage_Cancel_Submit_Rate",
        "Second_Stage_QPO_Count",
        "Second_Stage_Success_QPO_Count",
        "Second_Stage_Success_Submit_Rate",
        "Second_Stage_Cancel_QPO_Count",
        "Second_Stage_Cancel_Submit_Rate",
        "Third_Stage_QPO_Count",
        "Third_Stage_Success_QPO_Count",
        "Third_Stage_Success_Submit_Rate",
        "Third_Stage_Cancel_QPO_Count",
        "Third_Stage_Cancel_Submit_Rate",
    ]
    columns_cn = [
        "提报月份",
        "订单等级（AA+A合并为A）",
        "提报QPO数量",
        "成交QPO数量",
        "跟进中QPO数量",
        "取消QPO数量",
        "成交率（分母=提报QPO）",
        "跟进中占比（分母=提报QPO）",
        "取消率（分母=提报QPO）",
        "初次报价阶段QPO数量",
        "初次报价成交QPO数量",
        "初次报价成交占比（分母=提报QPO）",
        "初次报价取消QPO数量",
        "初次报价取消占比（分母=提报QPO）",
        "二次报价阶段QPO数量",
        "二次报价成交QPO数量",
        "二次报价成交占比（分母=提报QPO）",
        "二次报价取消QPO数量",
        "二次报价取消占比（分母=提报QPO）",
        "三次报价阶段QPO数量",
        "三次报价成交QPO数量",
        "三次报价成交占比（分母=提报QPO）",
        "三次报价取消QPO数量",
        "三次报价取消占比（分母=提报QPO）",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_NAME
    sheet.append(columns_en)
    sheet.append(columns_cn)
    for row in rows:
        sheet.append([row.get(column) for column in columns_en])
    style_two_header_sheet(sheet, columns_en, columns_cn)

    if notes_rows:
        notes_columns_en = ["Item", "Value", "Description"]
        notes_columns_cn = ["项目", "值", "说明"]
        notes_sheet = workbook.create_sheet(NOTES_SHEET_NAME)
        notes_sheet.append(notes_columns_en)
        notes_sheet.append(notes_columns_cn)
        for note_row in notes_rows:
            notes_sheet.append(list(note_row))
        style_two_header_sheet(notes_sheet, notes_columns_en, notes_columns_cn)
        style_notes_sheet(notes_sheet)

    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_path) if args.input_path else find_latest_input(BASE_DIR)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    records, metadata = load_stage_qpo_records(input_path)
    rows = aggregate(records)
    notes_rows = build_notes_rows(
        input_path=input_path,
        output_path=output_path,
        raw_records=records,
        summary_rows=rows,
        metadata=metadata,
    )
    write_workbook(rows, output_path, notes_rows=notes_rows)
    print("生成完成：")
    print(output_path)


if __name__ == "__main__":
    main()
