import ctypes
import datetime as dt
import os
import shutil
import subprocess
import sys
import tempfile
import time
from pathlib import Path
import re

import pyautogui
import pygetwindow as gw
import pyperclip
from openpyxl import Workbook, load_workbook

try:
    import pythoncom
    import win32com.client
except ImportError:
    pythoncom = None
    win32com = None


DB_PATH = Path(r"P:\0_Report\00_AMT中间表\UK100_工作量 KPI.accdb")
EXPORT_DIR = Path(r"P:\0_Report\00_AMT中间表\16_导出表格")
EXPORT_PATTERNS = ("UK100_工作量*.xlsx", "UK100_工作量*.xls")
EXPORT_WAIT_SECONDS = 600
EXPORT_RECENT_SECONDS = 120
POLL_INTERVAL_SECONDS = 10
TIME_RANGE_FORMAT = "%Y%m%d"
RANGE_SUFFIX_RE = re.compile(r"-\d{8}-\d{8}$")


def safe_print(*args, **kwargs):
    sep = kwargs.get("sep", " ")
    end = kwargs.get("end", "\n")
    text = sep.join(str(a) for a in args)
    try:
        sys.stdout.write(text + end)
    except UnicodeEncodeError:
        fallback = (text + end).encode("utf-8", errors="replace").decode("utf-8", errors="replace")
        sys.stdout.write(fallback)


def double_click_in_active_window(rel_x, rel_y, duration=0.3):
    win = gw.getActiveWindow()
    if not win:
        raise RuntimeError("当前没有激活窗口，无法双击。")

    abs_x = win.left + rel_x
    abs_y = win.top + rel_y
    pyautogui.moveTo(abs_x, abs_y, duration=duration)
    pyautogui.doubleClick()
    safe_print(f"已双击：窗口内坐标({rel_x}, {rel_y}) -> 屏幕坐标({abs_x}, {abs_y})")


def set_window_topmost(hwnd):
    hwnd_topmost = -1
    swp_nomove = 0x0002
    swp_nosize = 0x0001
    ctypes.windll.user32.SetWindowPos(hwnd, hwnd_topmost, 0, 0, 0, 0, swp_nomove | swp_nosize)


def activate_access_window(hwnd: int) -> None:
    sw_restore = 9
    user32 = ctypes.windll.user32
    user32.ShowWindow(hwnd, sw_restore)
    user32.SetForegroundWindow(hwnd)
    set_window_topmost(hwnd)
    time.sleep(1)


def ensure_pywin32() -> None:
    if pythoncom is None or win32com is None:
        raise RuntimeError("缺少 pywin32（pythoncom/win32com.client），无法通过 COM 正常控制 Access 退出。")


def get_access_hwnd(app) -> int:
    hwnd_value = app.hWndAccessApp
    if callable(hwnd_value):
        hwnd_value = hwnd_value()
    hwnd = int(hwnd_value)
    if hwnd <= 0:
        raise RuntimeError(f"无法获取有效的 Access 窗口句柄：{hwnd_value!r}")
    return hwnd


def get_window_process_id(hwnd: int) -> int:
    pid = ctypes.c_ulong()
    ctypes.windll.user32.GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
    if pid.value <= 0:
        raise RuntimeError(f"无法获取 Access 进程 PID：hwnd={hwnd}")
    return int(pid.value)


def is_process_running(pid: int) -> bool:
    synchronize = 0x00100000
    wait_timeout = 0x00000102
    process_handle = ctypes.windll.kernel32.OpenProcess(synchronize, False, pid)
    if not process_handle:
        return False
    try:
        return ctypes.windll.kernel32.WaitForSingleObject(process_handle, 0) == wait_timeout
    finally:
        ctypes.windll.kernel32.CloseHandle(process_handle)


def force_stop_process(pid: int) -> None:
    result = subprocess.run(
        ["powershell", "-NoProfile", "-Command", f"Stop-Process -Id {pid} -Force"],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0 and is_process_running(pid):
        raise RuntimeError(f"强制关闭 Access 进程失败：PID={pid}，{result.stderr.strip() or result.stdout.strip()}")


def open_access_application() -> tuple[object, int, int]:
    ensure_pywin32()
    pythoncom.CoInitialize()
    app = None
    try:
        app = win32com.client.DispatchEx("Access.Application")
        app.Visible = True
        app.OpenCurrentDatabase(str(DB_PATH))
        hwnd = get_access_hwnd(app)
        pid = get_window_process_id(hwnd)
        activate_access_window(hwnd)
        safe_print(f"已通过 COM 打开：{DB_PATH}")
        return app, hwnd, pid
    except Exception:
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
        raise


def close_access_application(app, hwnd: int | None = None, pid: int | None = None) -> None:
    if app is None:
        return
    try:
        try:
            app.CloseCurrentDatabase()
        except Exception:
            pass
        app.Quit()
        app = None
        deadline = time.time() + 10
        while time.time() <= deadline:
            window_exists = bool(hwnd) and bool(ctypes.windll.user32.IsWindow(hwnd))
            process_alive = bool(pid) and is_process_running(pid)
            if not window_exists and not process_alive:
                safe_print("已正常关闭 Access。")
                return
            time.sleep(0.5)

        if pid and is_process_running(pid):
            force_stop_process(pid)
            safe_print(f"Access 未正常退出，已强制关闭进程 PID={pid}。")
        else:
            safe_print("已正常关闭 Access。")
    finally:
        pythoncom.CoUninitialize()


def paste_date_into_active_field(hwnd: int, value: dt.date) -> None:
    pyperclip.copy(format_date(value))
    activate_access_window(hwnd)
    double_click_in_active_window(56, 83)
    pyautogui.hotkey("ctrl", "a")
    pyautogui.press("backspace")
    pyautogui.hotkey("ctrl", "v")
    pyautogui.press("enter")


def parse_date(text: str) -> dt.date:
    raw = text.strip().lstrip("\ufeff")
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return dt.datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"无法识别日期格式：{text}")


def format_date(value: dt.date) -> str:
    return value.strftime(TIME_RANGE_FORMAT)


def load_time_range() -> tuple[dt.date, dt.date]:
    time_range_file = Path(__file__).resolve().parent.parent / "时间范围.txt"
    if not time_range_file.exists():
        raise FileNotFoundError(f"未找到时间范围文件：{time_range_file}")

    lines = [line.strip().lstrip("\ufeff") for line in time_range_file.read_text(encoding="utf-8").splitlines()]
    values = [line for line in lines if line]
    if len(values) < 2:
        raise ValueError(f"时间范围文件内容不足两行：{time_range_file}")

    start_date = parse_date(values[0])
    end_date = parse_date(values[1])
    if end_date < start_date:
        raise ValueError(f"结束日期早于开始日期：{values[0]} > {values[1]}")
    return start_date, end_date


def iter_month_batches(start_date: dt.date, end_date: dt.date) -> list[tuple[dt.date, dt.date]]:
    batches: list[tuple[dt.date, dt.date]] = []
    current_start = start_date
    while current_start <= end_date:
        next_month = (current_start.replace(day=28) + dt.timedelta(days=4)).replace(day=1)
        current_end = min(end_date, next_month - dt.timedelta(days=1))
        batches.append((current_start, current_end))
        current_start = current_end + dt.timedelta(days=1)
    return batches


def build_range_suffix(start_date: dt.date, end_date: dt.date) -> str:
    return f"{format_date(start_date)}-{format_date(end_date)}"


def list_export_candidates(directory: Path) -> list[Path]:
    candidates: list[Path] = []
    for pattern in EXPORT_PATTERNS:
        candidates.extend(
            path for path in directory.glob(pattern) if path.is_file() and not path.name.startswith("~$")
        )
    return sorted(set(candidates))


def snapshot_export_candidates(directory: Path) -> dict[Path, int]:
    return {path: path.stat().st_mtime_ns for path in list_export_candidates(directory)}


def wait_for_recent_export(directory: Path, snapshot: dict[Path, int]) -> Path:
    deadline = time.time() + EXPORT_WAIT_SECONDS
    while time.time() <= deadline:
        now = time.time()
        recent_cutoff = now - EXPORT_RECENT_SECONDS
        changed_recent: list[Path] = []
        for path in list_export_candidates(directory):
            stat = path.stat()
            if snapshot.get(path) == stat.st_mtime_ns:
                continue
            if stat.st_mtime >= recent_cutoff:
                changed_recent.append(path)

        if changed_recent:
            return max(changed_recent, key=lambda item: (item.stat().st_mtime_ns, item.name))

        time.sleep(POLL_INTERVAL_SECONDS)

    raise FileNotFoundError(
        f"等待超过 {EXPORT_WAIT_SECONDS} 秒，仍未检测到最近 {EXPORT_RECENT_SECONDS} 秒内生成的 UK100 导出文件：{directory}"
    )


def remove_parent_uk100_kpi_excels(parent_dir: Path) -> None:
    removed_files: list[str] = []
    for pattern in EXPORT_PATTERNS:
        for file_path in parent_dir.glob(pattern):
            if file_path.is_file() and not file_path.name.startswith("~$"):
                file_path.unlink()
                removed_files.append(file_path.name)

    if removed_files:
        safe_print("已删除父级目录下旧 UK100 基础 Excel：")
        for file_name in sorted(removed_files):
            safe_print(f"- {file_name}")
    else:
        safe_print("父级目录下未发现需要删除的 UK100 基础 Excel。")


def remove_existing_uk100_excels(base_dir: Path) -> None:
    removed_files: list[str] = []
    for pattern in ("UK100*.xlsx", "UK100*.xls"):
        for file_path in base_dir.glob(pattern):
            if file_path.is_file() and not file_path.name.startswith("~$"):
                file_path.unlink()
                removed_files.append(file_path.name)

    if removed_files:
        safe_print("已删除旧UK100结果文件：")
        for file_name in sorted(removed_files):
            safe_print(f"- {file_name}")
    else:
        safe_print("当前目录下未发现需要删除的 UK100 开头 Excel。")


def run_access_export(start_date: dt.date, end_date: dt.date) -> Path:
    if not DB_PATH.exists():
        raise FileNotFoundError(f"文件不存在：{DB_PATH}")

    snapshot = snapshot_export_candidates(EXPORT_DIR)
    access_app = None
    hwnd = None
    pid = None
    try:
        access_app, hwnd, pid = open_access_application()
        activate_access_window(hwnd)
        double_click_in_active_window(218, 220)
        time.sleep(6)

        paste_date_into_active_field(hwnd, start_date)
        time.sleep(1)
        paste_date_into_active_field(hwnd, end_date)
        exported_file = wait_for_recent_export(EXPORT_DIR, snapshot)
    finally:
        close_access_application(access_app, hwnd, pid)

    safe_print(f"检测到导出文件：{exported_file}")
    return exported_file


def copy_batch_file(source_path: Path, temp_dir: Path, start_date: dt.date, end_date: dt.date) -> Path:
    target_name = f"{source_path.stem}-{build_range_suffix(start_date, end_date)}{source_path.suffix}"
    target_path = temp_dir / target_name
    shutil.copy2(source_path, target_path)
    return target_path


def remove_batch_cache(batch_files: list[Path], temp_dir: Path) -> None:
    removed_files: list[str] = []
    for batch_file in batch_files:
        if batch_file.exists():
            batch_file.unlink()
            removed_files.append(batch_file.name)

    if temp_dir.exists():
        shutil.rmtree(temp_dir)

    if removed_files:
        safe_print("已删除批次缓存文件：")
        for file_name in removed_files:
            safe_print(f"- {file_name}")
    safe_print(f"已删除批次缓存目录：{temp_dir.name}")


def row_contains_chinese(values: tuple[object, ...]) -> bool:
    for value in values:
        text = str(value or "")
        if any("\u4e00" <= char <= "\u9fff" for char in text):
            return True
    return False


def detect_data_start_row(sheet) -> int:
    if sheet.max_row <= 1:
        return 2
    second_row = tuple(cell for cell in next(sheet.iter_rows(min_row=2, max_row=2, values_only=True)))
    if row_contains_chinese(second_row):
        return 3
    return 2


def append_sheet_rows(target_sheet, source_sheet, *, include_headers: bool) -> None:
    min_row = 1 if include_headers else detect_data_start_row(source_sheet)
    for row in source_sheet.iter_rows(min_row=min_row, values_only=True):
        target_sheet.append(list(row))


def merge_batch_excels(batch_files: list[Path], output_path: Path) -> None:
    if not batch_files:
        raise ValueError("没有可汇总的批次文件。")

    if len(batch_files) == 1:
        shutil.copy2(batch_files[0], output_path)
        return

    merged_workbook = Workbook()
    merged_workbook.remove(merged_workbook.active)

    for batch_file in batch_files:
        source_workbook = load_workbook(batch_file, read_only=True, data_only=False)
        try:
            for sheet_name in source_workbook.sheetnames:
                source_sheet = source_workbook[sheet_name]
                if sheet_name not in merged_workbook.sheetnames:
                    target_sheet = merged_workbook.create_sheet(sheet_name)
                    append_sheet_rows(target_sheet, source_sheet, include_headers=True)
                    freeze_panes = getattr(source_sheet, "freeze_panes", None)
                    if freeze_panes:
                        target_sheet.freeze_panes = freeze_panes
                else:
                    append_sheet_rows(merged_workbook[sheet_name], source_sheet, include_headers=False)
        finally:
            source_workbook.close()

    merged_workbook.save(output_path)


def build_final_output_path(base_dir: Path, sample_file: Path, start_date: dt.date, end_date: dt.date) -> Path:
    stem = RANGE_SUFFIX_RE.sub("", sample_file.stem)
    return base_dir / f"{stem}-{build_range_suffix(start_date, end_date)}{sample_file.suffix}"


if os.name == "nt":
    os.system("chcp 65001 > nul")
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")


def main():
    base_dir = Path(__file__).resolve().parent.parent
    remove_parent_uk100_kpi_excels(base_dir)
    remove_existing_uk100_excels(base_dir)

    start_date, end_date = load_time_range()
    batches = iter_month_batches(start_date, end_date)
    safe_print(f"共需处理 {len(batches)} 个批次：{build_range_suffix(start_date, end_date)}")

    temp_dir = Path(tempfile.mkdtemp(prefix="uk100_batches_", dir=str(base_dir)))
    batch_files: list[Path] = []
    completed = False
    try:
        for index, (batch_start, batch_end) in enumerate(batches, start=1):
            safe_print(f"开始处理批次 {index}/{len(batches)}：{build_range_suffix(batch_start, batch_end)}")
            exported_file = run_access_export(batch_start, batch_end)
            copied_file = copy_batch_file(exported_file, temp_dir, batch_start, batch_end)
            batch_files.append(copied_file)
            safe_print(f"已缓存批次文件：{copied_file.name}")

        final_output = build_final_output_path(base_dir, batch_files[-1], start_date, end_date)
        merge_batch_excels(batch_files, final_output)
        safe_print(f"已生成最终 UK100 基础表：{final_output}")
        completed = True
    finally:
        if completed:
            remove_batch_cache(batch_files, temp_dir)


if __name__ == "__main__":
    main()
