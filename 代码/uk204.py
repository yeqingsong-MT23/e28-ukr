# -*- coding: utf-8 -*-
"""
根据 UK200_转化漏斗KPI 基础表生成 UK204_OM UEA转化漏斗概览.xlsx

默认行为：
1. 自动读取脚本所在目录下最新的 UK200 基础表
2. 输出到 P:\\0_Report\\78_AI Report\\UKR_Report 下，文件名自动追加 -MMDDHHMM

示例：
python uk204.py
python uk204.py --input "C:\\\\path\\\\UK200_转化漏斗KPI-20260430.xlsx"
python uk204.py --input "C:\\\\path\\\\UK200_转化漏斗KPI-20260430.xlsx" --output "C:\\\\path\\\\UK204.xlsx"
python uk204.py --start-date 2025-01-01
"""

from __future__ import annotations

import argparse
import datetime as dt
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from uk_export_utils import build_default_output_path
from uk_funnel_utils import (
    detect_data_sheet as funnel_detect_data_sheet,
    find_latest_input as funnel_find_latest_input,
    map_simple_type,
    month_sort_key as funnel_month_sort_key,
    normalize_text as funnel_normalize_text,
    parse_cell_date as funnel_parse_cell_date,
    safe_div,
    style_two_header_sheet,
)

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK204_OM UEA转化漏斗概览.xlsx"
DEFAULT_START_DATE = dt.date(2025, 1, 1)

BASE_COLUMNS_EN = [
    "OM_Flup",
    "Stat_Month",
    "Total_QPO_Count",
    "Success_QPO_Count",
    "Success_QPO_Rate",
    "Cancel_QPO_Count",
    "Cancel_QPO_Rate",
    "InProgress_QPO_Count",
    "InProgress_QPO_Rate",
]

BASE_COLUMNS_CN = [
    "OM跟进人",
    "统计月份",
    "提报QPO总数",
    "成交QPO数量",
    "成交QPO占比",
    "取消QPO数量",
    "取消QPO占比",
    "跟进中QPO数量",
    "跟进中QPO占比",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据 UK200 基础表生成 UK204_OM UEA转化漏斗概览。"
    )
    parser.add_argument(
        "--input",
        dest="input_path",
        help="UK200 基础表路径；不传时自动读取当前目录最新的 UK200 开头基础表",
    )
    parser.add_argument(
        "--output",
        dest="output_path",
        help=f"输出文件路径；不传时默认写入 {EXPORT_DIR}，文件名自动追加 -MMDDHHMM",
    )
    parser.add_argument(
        "--start-date",
        default=DEFAULT_START_DATE.isoformat(),
        help=f"纳入统计的最早提报日期，格式 YYYY-MM-DD，默认 {DEFAULT_START_DATE.isoformat()}",
    )
    return parser.parse_args()


def parse_cli_date(text: str) -> dt.date:
    try:
        return dt.date.fromisoformat(str(text).strip())
    except ValueError as exc:
        raise ValueError(f"无法识别开始日期：{text}，请使用 YYYY-MM-DD 格式。") from exc


def parse_cell_date(value) -> dt.date | None:
    return funnel_parse_cell_date(value)


def normalize_text(value: object) -> str:
    return funnel_normalize_text(value)


def find_latest_input(base_dir: Path) -> Path:
    return funnel_find_latest_input(base_dir)


def detect_data_sheet(workbook):
    return funnel_detect_data_sheet(workbook)


def month_sort_key(month: str) -> tuple[int, int]:
    return funnel_month_sort_key(month)


def has_chinese(text: object) -> bool:
    content = str(text or "")
    return any("\u4e00" <= char <= "\u9fff" for char in content)


def map_status_category(status: str) -> str | None:
    category = map_simple_type(status)
    if category in {"Success", "Cancel", "InProgress"}:
        return category
    return None


def node_sort_key(node: str) -> tuple[int, str]:
    digits = "".join(char for char in str(node) if char.isdigit())
    return (int(digits) if digits else 999999, str(node))


def resolve_column_indexes(header_row: list[str]) -> dict[str, int]:
    matches: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(header_row):
        matches[normalize_text(header)].append(index)

    required_single = {
        "Submission_Date": "Submission_Date",
        "QPO_ID": "QPO_ID",
        "QPO_Status": "QPO_Status",
        "OM_Flup": "OM_Flup",
    }

    indexes: dict[str, int] = {}
    missing: list[str] = []

    for key, header_name in required_single.items():
        if matches.get(header_name):
            indexes[key] = matches[header_name][0]
        else:
            missing.append(header_name)

    node_headers = matches.get("WPO_Current_Node", [])
    if len(node_headers) >= 2:
        indexes["WPO_Current_Node"] = node_headers[0]
        indexes["WPO_Current_Node_CN"] = node_headers[1]
    else:
        missing.append("WPO_Current_Node(需存在代码列和中文列各1个)")

    if missing:
        raise ValueError(f"源表缺少必要字段：{missing}")

    return indexes


def build_summary(
    input_path: Path,
    start_date: dt.date,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = detect_data_sheet(workbook)

        header_iter = sheet.iter_rows(min_row=1, max_row=2, values_only=True)
        header_row_1_raw = next(header_iter, None)
        if header_row_1_raw is None:
            raise ValueError("源表为空，无法读取表头。")
        header_row_1 = [normalize_text(value) for value in header_row_1_raw]

        header_row_2_raw = next(header_iter, None)
        header_row_2 = [normalize_text(value) for value in header_row_2_raw] if header_row_2_raw else []

        data_start_row = 3 if header_row_2 and sum(has_chinese(value) for value in header_row_2) > 0 else 2
        column_indexes = resolve_column_indexes(header_row_1)

        group_counts = defaultdict(
            lambda: {
                "Total": 0,
                "Success": 0,
                "Cancel": 0,
                "InProgress": 0,
                "nodes": Counter(),
            }
        )
        qpo_state: dict[tuple[str, str, str], dict[str, str | None]] = {}
        node_cn_map: dict[str, str] = {}
        missing_node_cn = set()
        unmapped_status = Counter()

        input_rows = 0
        valid_date_rows = 0
        after_start_date_rows = 0
        included_rows = 0
        min_date: dt.date | None = None
        max_date: dt.date | None = None

        def process_one_group(
            om: str,
            month: str,
            qpo_id: str,
            category: str | None,
            node: str,
            node_cn: str,
        ) -> None:
            key = (om, month, qpo_id)
            record = qpo_state.setdefault(
                key,
                {
                    "category": None,
                    "node": None,
                },
            )

            if node:
                if node_cn:
                    node_cn_map.setdefault(node, node_cn)
                else:
                    missing_node_cn.add(node)

            if category is None:
                return

            if record["category"] is None:
                record["category"] = category
                group_counts[(om, month)]["Total"] += 1
                group_counts[(om, month)][category] += 1

                if category == "Cancel" and node:
                    record["node"] = node
                    group_counts[(om, month)]["nodes"][node] += 1

            elif record["category"] == "Cancel" and record["node"] is None and node:
                record["node"] = node
                group_counts[(om, month)]["nodes"][node] += 1

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            input_rows += 1

            submission_date = parse_cell_date(row[column_indexes["Submission_Date"]])
            if submission_date is None:
                continue
            valid_date_rows += 1

            if submission_date < start_date:
                continue
            after_start_date_rows += 1

            qpo_id = normalize_text(row[column_indexes["QPO_ID"]])
            om_flup = normalize_text(row[column_indexes["OM_Flup"]])
            if not qpo_id or not om_flup:
                continue
            included_rows += 1

            if min_date is None or submission_date < min_date:
                min_date = submission_date
            if max_date is None or submission_date > max_date:
                max_date = submission_date

            status = normalize_text(row[column_indexes["QPO_Status"]])
            node = normalize_text(row[column_indexes["WPO_Current_Node"]])
            node_cn = normalize_text(row[column_indexes["WPO_Current_Node_CN"]])
            category = map_status_category(status)

            if category is None:
                unmapped_status[status or "<blank>"] += 1

            month = f"{submission_date.year:04d}-{submission_date.month:02d}"

            process_one_group(om_flup, month, qpo_id, category, node, node_cn)
            process_one_group(om_flup, "ALL", qpo_id, category, node, node_cn)

        all_nodes = set()
        for value in group_counts.values():
            all_nodes.update(value["nodes"].keys())
        nodes_sorted = sorted(all_nodes, key=node_sort_key)

        rows: list[dict[str, object]] = []
        validation_total_equals_sum = True

        for (om, month), value in sorted(
            group_counts.items(),
            key=lambda item: (item[0][0], month_sort_key(item[0][1])),
        ):
            total = value["Total"]
            success = value["Success"]
            cancel = value["Cancel"]
            in_progress = value["InProgress"]

            if total != success + cancel + in_progress:
                validation_total_equals_sum = False

            row_data: dict[str, object] = {
                "OM_Flup": om,
                "Stat_Month": month,
                "Total_QPO_Count": total,
                "Success_QPO_Count": success,
                "Success_QPO_Rate": safe_div(success, total),
                "Cancel_QPO_Count": cancel,
                "Cancel_QPO_Rate": safe_div(cancel, total),
                "InProgress_QPO_Count": in_progress,
                "InProgress_QPO_Rate": safe_div(in_progress, total),
            }

            for node in nodes_sorted:
                node_count = value["nodes"].get(node, 0)
                row_data[f"{node}_Cancel_QPO_Count"] = node_count
                row_data[f"{node}_Cancel_QPO_Rate"] = safe_div(node_count, total)

            rows.append(row_data)

        metadata = {
            "input_path": str(input_path),
            "start_date": start_date,
            "nodes_sorted": nodes_sorted,
            "node_cn_map": node_cn_map,
            "input_rows": input_rows,
            "valid_date_rows": valid_date_rows,
            "after_start_date_rows": after_start_date_rows,
            "included_rows": included_rows,
            "min_date": min_date,
            "max_date": max_date,
            "unmapped_status": unmapped_status,
            "missing_node_cn": sorted(missing_node_cn - set(node_cn_map.keys()), key=node_sort_key),
            "validation_total_equals_sum": validation_total_equals_sum,
        }
        return rows, metadata
    finally:
        workbook.close()


def build_column_headers(nodes_sorted: list[str], node_cn_map: dict[str, str]) -> tuple[list[str], list[str]]:
    columns_en = list(BASE_COLUMNS_EN)
    columns_cn = list(BASE_COLUMNS_CN)

    for node in nodes_sorted:
        node_name_cn = node_cn_map.get(node, node)
        columns_en.extend(
            [
                f"{node}_Cancel_QPO_Count",
                f"{node}_Cancel_QPO_Rate",
            ]
        )
        columns_cn.extend(
            [
                f"{node_name_cn}取消QPO数量",
                f"{node_name_cn}取消QPO占比",
            ]
        )

    return columns_en, columns_cn


def set_sheet_widths(sheet, preview_rows: int = 50, max_width: int = 35) -> None:
    for col_idx in range(1, sheet.max_column + 1):
        lengths = []
        for row_idx in range(1, min(sheet.max_row, preview_rows) + 1):
            value = sheet.cell(row=row_idx, column=col_idx).value
            if value is not None:
                lengths.append(len(str(value)))
        width = min(max(lengths or [10]) + 2, max_width)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width


def write_main_sheet(workbook: Workbook, rows: list[dict[str, object]], metadata: dict[str, object]) -> None:
    nodes_sorted = metadata["nodes_sorted"]
    node_cn_map = metadata["node_cn_map"]
    columns_en, columns_cn = build_column_headers(nodes_sorted, node_cn_map)

    sheet = workbook.active
    sheet.title = "UK204_OM UEA"
    sheet.append(columns_en)
    sheet.append(columns_cn)

    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    style_two_header_sheet(sheet, columns_en, columns_cn)
    sheet.freeze_panes = "A3"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    count_columns = {
        "Total_QPO_Count",
        "Success_QPO_Count",
        "Cancel_QPO_Count",
        "InProgress_QPO_Count",
    }
    rate_columns = {
        "Success_QPO_Rate",
        "Cancel_QPO_Rate",
        "InProgress_QPO_Rate",
    }

    for column in columns_en:
        if column.endswith("_Cancel_QPO_Count"):
            count_columns.add(column)
        elif column.endswith("_Cancel_QPO_Rate"):
            rate_columns.add(column)

    column_index = {column: idx + 1 for idx, column in enumerate(columns_en)}

    for row_num in range(3, sheet.max_row + 1):
        for column in columns_en:
            cell = sheet.cell(row=row_num, column=column_index[column])
            cell.alignment = center
            if column in count_columns:
                cell.number_format = "0"
            elif column in rate_columns:
                cell.number_format = "0.00%"

    set_sheet_widths(sheet, preview_rows=60, max_width=35)


def build_notes_rows(metadata: dict[str, object]) -> list[list[object]]:
    nodes_sorted: list[str] = metadata["nodes_sorted"]
    node_cn_map: dict[str, str] = metadata["node_cn_map"]
    start_date: dt.date = metadata["start_date"]
    min_date: dt.date | None = metadata["min_date"]
    max_date: dt.date | None = metadata["max_date"]
    unmapped_status: Counter = metadata["unmapped_status"]
    missing_node_cn: list[str] = metadata["missing_node_cn"]

    nodes_mapping_text = (
        "None"
        if not nodes_sorted
        else "; ".join(f"{node}:{node_cn_map.get(node, '<missing>')}" for node in nodes_sorted)
    )

    base_rows: list[list[object]] = [
        [
            "OM_Flup",
            "OM跟进人",
            "Dimension",
            "OM 的 UEA 跟进人（来自 UK200 字段 OM_Flup）",
            "OM_Flup 为空不参与汇总",
            "UK204_OM UEA",
        ],
        [
            "Stat_Month",
            "统计月份",
            "Dimension",
            f"统计月份由 Submission_Date 提取为 YYYY-MM；并额外生成 ALL 汇总行（>= {start_date.isoformat()}）",
            "示例：2025-01 / ALL",
            "UK204_OM UEA",
        ],
        [
            "Total_QPO_Count",
            "提报QPO总数",
            "Scale",
            "在该维度下按 QPO_ID 去重计数；仅统计能映射到 Success / Cancel / InProgress 的 QPO",
            "Total = Success + Cancel + InProgress",
            "UK204_OM UEA",
        ],
        [
            "Success_QPO_Count",
            "成交QPO数量",
            "Outcome",
            "QPO_Status 映射为 Success 的 QPO_ID 去重计数",
            "Success / Cancel / InProgress 为统一状态口径",
            "UK204_OM UEA",
        ],
        [
            "Success_QPO_Rate",
            "成交QPO占比",
            "Outcome",
            "Success_QPO_Count / Total_QPO_Count",
            "汇总行按汇总分子/分母重新计算，不能简单平均",
            "UK204_OM UEA",
        ],
        [
            "Cancel_QPO_Count",
            "取消QPO数量",
            "Outcome",
            "QPO_Status 映射为 Cancel 的 QPO_ID 去重计数",
            "取消节点分布详见动态节点列",
            "UK204_OM UEA",
        ],
        [
            "Cancel_QPO_Rate",
            "取消QPO占比",
            "Outcome",
            "Cancel_QPO_Count / Total_QPO_Count",
            "汇总行按汇总分子/分母重新计算，不能简单平均",
            "UK204_OM UEA",
        ],
        [
            "InProgress_QPO_Count",
            "跟进中QPO数量",
            "Outcome",
            "QPO_Status 映射为 InProgress 的 QPO_ID 去重计数",
            "未映射状态不纳入 Total",
            "UK204_OM UEA",
        ],
        [
            "InProgress_QPO_Rate",
            "跟进中QPO占比",
            "Outcome",
            "InProgress_QPO_Count / Total_QPO_Count",
            "汇总行按汇总分子/分母重新计算，不能简单平均",
            "UK204_OM UEA",
        ],
        [
            "Start_Date_Filter",
            "起始日期过滤",
            "Filter",
            "默认仅统计 Submission_Date >= 2025-01-01；可通过 --start-date 调整",
            start_date.isoformat(),
            "Notes",
        ],
        [
            "Submission_Date_Rule",
            "提报日期口径",
            "Rule",
            "Month 从 Submission_Date 提取；无有效 Submission_Date 的记录不参与统计",
            "YYYY-MM",
            "Notes",
        ],
        [
            "QPO_Status_Mapping",
            "状态映射",
            "Rule",
            "QPO_Status 统一映射为 Success / Cancel / InProgress",
            (
                "Success=End, WaitingRPO, WaitingPayment, WaitingDeal, OnDeal, WaitingPO; "
                "Cancel=SellerCancel, CSCancel, Long time-Cancel, WootCancel; "
                "InProgress=Hold, New, SubmitRBS"
            ),
            "Notes",
        ],
        [
            "Cancel_Node_Attribution",
            "取消节点归因",
            "Rule",
            "取消节点列来自运行期识别的 nodes_sorted；节点中文来自 WPO_Current_Node 中文列。按当前脚本口径，同一 QPO 若出现多个取消节点，按源表首次遇到的取消节点归因",
            nodes_mapping_text,
            "UK204_OM UEA",
        ],
        [
            "Dynamic_Node_Columns",
            "动态节点列展开",
            "Rule",
            "主表除 BASE_COLUMNS_EN 外，还会按运行期识别到的 nodes_sorted 动态展开 *_Cancel_QPO_Count / *_Cancel_QPO_Rate 列",
            "节点列表见 SELF_CHECK_REPORT",
            "UK204_OM UEA",
        ],
        [
            "Rate_Recalculation_Rule",
            "汇总比率重算规则",
            "Rule",
            "所有 Rate 均由汇总分子/汇总分母重新计算，不对分组比例做平均",
            "分母均为 Total_QPO_Count",
            "UK204_OM UEA",
        ],
    ]

    dynamic_rows: list[list[object]] = []
    for node in nodes_sorted:
        node_name_cn = node_cn_map.get(node, node)
        dynamic_rows.append(
            [
                f"{node}_Cancel_QPO_Count",
                f"{node_name_cn}取消QPO数量",
                "Cancel Node Distribution",
                f"仅对 Cancel 类 QPO，按 WPO_Current_Node={node} 归因的 QPO_ID 去重计数",
                f"节点中文={node_name_cn}",
                "UK204_OM UEA",
            ]
        )
        dynamic_rows.append(
            [
                f"{node}_Cancel_QPO_Rate",
                f"{node_name_cn}取消QPO占比",
                "Cancel Node Distribution",
                f"{node}_Cancel_QPO_Count / Total_QPO_Count",
                "分母不是 Cancel_QPO_Count",
                "UK204_OM UEA",
            ]
        )

    self_check_rows: list[list[object]] = [
        [None, None, None, None, None, None],
        ["SELF_CHECK_REPORT", "自检报告", None, None, None, "Notes"],
        ["Input_Rows", "输入行数", None, metadata["input_rows"], None, "Notes"],
        ["Valid_Date_Rows", "Submission_Date可解析行数", None, metadata["valid_date_rows"], None, "Notes"],
        ["Start_Date_Filter", "起始日期过滤", None, start_date.isoformat(), None, "Notes"],
        ["After_StartDate_Rows", "起始日期过滤后行数", None, metadata["after_start_date_rows"], None, "Notes"],
        ["Included_Rows", "纳入汇总行数", None, metadata["included_rows"], "需同时具备有效 QPO_ID 与 OM_Flup", "Notes"],
        [
            "Time_Range_Covered",
            "统计时间范围实际覆盖",
            None,
            f"{min_date.isoformat()} ~ {max_date.isoformat()}" if min_date and max_date else "None",
            f"仅统计 >= {start_date.isoformat()}",
            "Notes",
        ],
        [
            "Unmapped_QPO_Status",
            "未映射状态列表与数量",
            None,
            (
                "None"
                if not unmapped_status
                else "; ".join(f"{key}:{value}" for key, value in sorted(unmapped_status.items()))
            ),
            "未映射状态不纳入 Total",
            "Notes",
        ],
        [
            "Included_Nodes",
            "纳入的取消节点列表",
            None,
            "None" if not nodes_sorted else ", ".join(nodes_sorted),
            None,
            "Notes",
        ],
        [
            "Node_CN_Map",
            "节点中文映射",
            None,
            nodes_mapping_text,
            None,
            "Notes",
        ],
        [
            "Missing_Node_CN",
            "缺失中文映射的节点",
            None,
            "None" if not missing_node_cn else ", ".join(missing_node_cn),
            None,
            "Notes",
        ],
        [
            "Node_CN_Mapping_Complete",
            "节点中文映射完整性",
            None,
            "PASS" if not missing_node_cn else "FAIL",
            "如 FAIL 需补齐 WPO_Current_Node 中文列",
            "Notes",
        ],
        [
            "validation_total_equals_sum",
            "校验 Total=Success+Cancel+InProgress",
            None,
            "PASS" if metadata["validation_total_equals_sum"] else "FAIL",
            "如 FAIL 需排查未映射状态或重复归类",
            "Notes",
        ],
    ]

    return base_rows + dynamic_rows + self_check_rows


def write_notes_sheet(workbook: Workbook, metadata: dict[str, object]) -> None:
    sheet = workbook.create_sheet("Notes")
    headers_en = ["Field_EN", "Field_CN", "Field_Group", "Definition", "Enum_or_Notes", "Sheet"]
    headers_cn = ["字段英文", "字段中文", "字段分组", "口径说明", "枚举值/说明", "所属Sheet"]

    sheet.append(headers_en)
    sheet.append(headers_cn)

    for row in build_notes_rows(metadata):
        sheet.append(row)

    style_two_header_sheet(sheet, headers_en, headers_cn)
    sheet.freeze_panes = "A3"

    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = align

    widths = {
        1: 28,
        2: 24,
        3: 20,
        4: 62,
        5: 48,
        6: 16,
    }
    for idx, width in widths.items():
        sheet.column_dimensions[get_column_letter(idx)].width = width


def write_workbook(
    rows: list[dict[str, object]],
    metadata: dict[str, object],
    output_path: Path,
) -> None:
    workbook = Workbook()
    write_main_sheet(workbook, rows, metadata)
    write_notes_sheet(workbook, metadata)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def generate_report(input_path: Path, output_path: Path, start_date: dt.date) -> None:
    rows, metadata = build_summary(input_path=input_path, start_date=start_date)
    write_workbook(rows=rows, metadata=metadata, output_path=output_path)


def main() -> None:
    args = parse_args()
    start_date = parse_cli_date(args.start_date)

    input_path = Path(args.input_path) if args.input_path else find_latest_input(BASE_DIR)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )

    generate_report(input_path=input_path, output_path=output_path, start_date=start_date)

    print("生成完成：")
    print(output_path)


if __name__ == "__main__":
    main()
