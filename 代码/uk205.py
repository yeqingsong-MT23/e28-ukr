# -*- coding: utf-8 -*-
"""
根据 UK200 基础表生成 UK205_KP UEA转化漏斗概览.xlsx。
"""

from __future__ import annotations

import argparse
import datetime as dt
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from uk_export_utils import build_default_output_path
from uk_funnel_utils import (
    detect_data_sheet,
    find_latest_input,
    get_required_indexes,
    map_simple_type,
    month_sort_key,
    normalize_text,
    parse_cell_date,
    safe_div,
    style_two_header_sheet,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK205_KP UEA转化漏斗概览.xlsx"
DEFAULT_START_DATE = dt.date(2025, 1, 1)

BASE_COLUMNS_EN = [
    "KP_Flup",
    "Stat_Month",
    "Total_QPO_Count",
    "Success_QPO_Count",
    "Success_QPO_Rate",
    "Cancel_QPO_Count",
    "Cancel_QPO_Rate",
    "InProgress_QPO_Count",
    "InProgress_QPO_Rate",
]

BASE_COLUMNS_CN = [
    "KP_Flup",
    "统计月份",
    "提报QPO总数",
    "成交QPO数量",
    "成交QPO占比",
    "取消QPO数量",
    "取消QPO占比",
    "跟进中QPO数量",
    "跟进中QPO占比",
]


def _contains_chinese(value: object) -> bool:
    text = normalize_text(value)
    return any("\u4e00" <= ch <= "\u9fff" for ch in text)


def _second_row_is_cn_header(row_values: Iterable[object]) -> bool:
    values = list(row_values)
    if not values:
        return False
    chinese_count = sum(1 for value in values if _contains_chinese(value))
    non_blank_count = sum(1 for value in values if normalize_text(value))
    return chinese_count >= 2 and non_blank_count >= 2


def _node_sort_key(node: str) -> tuple[int, str]:
    digits = "".join(ch for ch in node if ch.isdigit())
    return (int(digits) if digits else 999999, node)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据 UK200 基础表生成 UK205_KP UEA转化漏斗概览。"
    )
    parser.add_argument("--input", dest="input_path", help="指定输入基础表路径")
    parser.add_argument("--output", dest="output_path", help="指定输出 Excel 路径")
    parser.add_argument(
        "--start-date",
        default=DEFAULT_START_DATE.isoformat(),
        help=f"起始提报日期，默认 {DEFAULT_START_DATE.isoformat()}",
    )
    return parser.parse_args()


def build_summary(
    input_path: Path, start_date: dt.date
) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = detect_data_sheet(workbook)

        row_iter = sheet.iter_rows(values_only=True)
        try:
            header_row_1 = [normalize_text(value) for value in next(row_iter)]
        except StopIteration as exc:
            raise ValueError("输入工作簿为空或缺少表头。") from exc

        try:
            raw_row_2 = next(row_iter)
        except StopIteration:
            raw_row_2 = tuple()

        header_row_2 = [normalize_text(value) for value in raw_row_2]
        data_start_row = 3 if _second_row_is_cn_header(raw_row_2) else 2

        try:
            indexes = get_required_indexes(
                header_row_1,
                {
                    "Submission_Date": "Submission_Date",
                    "QPO_ID": "QPO_ID",
                    "QPO_Status": "QPO_Status",
                    "KP_Flup": "KP_Flup",
                },
                {
                    "WPO_Current_Node": ("WPO_Current_Node", 0),
                    "WPO_Current_Node_CN": ("WPO_Current_Node", 1),
                },
            )
        except Exception as exc:
            raise ValueError(
                "输入表缺少必需字段。必需字段: Submission_Date, QPO_ID, QPO_Status, "
                "WPO_Current_Node, WPO_Current_Node.1, KP_Flup"
            ) from exc

        grouped = defaultdict(
            lambda: {
                "Total": 0,
                "Success": 0,
                "Cancel": 0,
                "InProgress": 0,
                "nodes": Counter(),
            }
        )
        qpo_state: dict[tuple[str, str, str], dict[str, str | None]] = {}
        node_cn_map: dict[str, str] = {}
        missing_node_cn: set[str] = set()
        unmapped_status = Counter()
        conflict_counter = 0

        input_rows = 0
        dated_rows = 0
        in_scope_rows = 0
        rows_with_required_keys = 0
        mapped_rows = 0

        min_date: dt.date | None = None
        max_date: dt.date | None = None

        def process_group(
            kp: str,
            month: str,
            qpo_id: str,
            category: str | None,
            node: str,
            node_cn: str,
        ) -> None:
            nonlocal conflict_counter
            key = (kp, month, qpo_id)
            state = qpo_state.setdefault(key, {"category": None, "node": None})

            if node:
                if node_cn:
                    node_cn_map.setdefault(node, node_cn)
                else:
                    missing_node_cn.add(node)

            if category is None:
                return

            if state["category"] is None:
                state["category"] = category
                grouped[(kp, month)]["Total"] += 1
                grouped[(kp, month)][category] += 1
                if category == "Cancel" and node:
                    state["node"] = node
                    grouped[(kp, month)]["nodes"][node] += 1
                return

            if state["category"] != category:
                conflict_counter += 1
                return

            if category == "Cancel" and state["node"] is None and node:
                state["node"] = node
                grouped[(kp, month)]["nodes"][node] += 1

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            input_rows += 1

            submission_date = parse_cell_date(row[indexes["Submission_Date"]])
            if submission_date is None:
                continue
            dated_rows += 1

            if submission_date < start_date:
                continue
            in_scope_rows += 1

            qpo_id = normalize_text(row[indexes["QPO_ID"]])
            kp_flup = normalize_text(row[indexes["KP_Flup"]])
            if not qpo_id or not kp_flup:
                continue
            rows_with_required_keys += 1

            if min_date is None or submission_date < min_date:
                min_date = submission_date
            if max_date is None or submission_date > max_date:
                max_date = submission_date

            status = normalize_text(row[indexes["QPO_Status"]])
            category = map_simple_type(status)
            if category is None:
                unmapped_status[status or "<blank>"] += 1
            else:
                mapped_rows += 1

            node = normalize_text(row[indexes["WPO_Current_Node"]])
            node_cn = normalize_text(row[indexes["WPO_Current_Node_CN"]])
            month = submission_date.strftime("%Y-%m")

            process_group(kp_flup, month, qpo_id, category, node, node_cn)
            process_group(kp_flup, "ALL", qpo_id, category, node, node_cn)

        all_nodes = sorted(
            {node for value in grouped.values() for node in value["nodes"] if node},
            key=_node_sort_key,
        )

        rows: list[dict[str, object]] = []
        for (kp, month), value in sorted(
            grouped.items(),
            key=lambda item: (item[0][0], month_sort_key(item[0][1])),
        ):
            total = value["Total"]
            row_data: dict[str, object] = {
                "KP_Flup": kp,
                "Stat_Month": month,
                "Total_QPO_Count": total,
                "Success_QPO_Count": value["Success"],
                "Success_QPO_Rate": safe_div(value["Success"], total),
                "Cancel_QPO_Count": value["Cancel"],
                "Cancel_QPO_Rate": safe_div(value["Cancel"], total),
                "InProgress_QPO_Count": value["InProgress"],
                "InProgress_QPO_Rate": safe_div(value["InProgress"], total),
            }
            for node in all_nodes:
                count = value["nodes"].get(node, 0)
                row_data[f"{node}_Cancel_QPO_Count"] = count
                row_data[f"{node}_Cancel_QPO_Rate"] = safe_div(count, total)
            rows.append(row_data)

        all_row_count = sum(1 for row in rows if row["Stat_Month"] == "ALL")
        detected_node_list = [
            f"{node} / {node_cn_map.get(node, node)}" if node_cn_map.get(node) else node
            for node in all_nodes
        ]

        metadata: dict[str, object] = {
            "nodes_sorted": all_nodes,
            "node_cn_map": node_cn_map,
            "input_rows": input_rows,
            "dated_rows": dated_rows,
            "in_scope_rows": in_scope_rows,
            "rows_with_required_keys": rows_with_required_keys,
            "mapped_rows": mapped_rows,
            "output_row_count": len(rows),
            "all_row_count": all_row_count,
            "min_date": min_date,
            "max_date": max_date,
            "unmapped_status": unmapped_status,
            "missing_node_cn": sorted(missing_node_cn - set(node_cn_map)),
            "conflict_counter": conflict_counter,
            "start_date": start_date,
            "detected_node_list": detected_node_list,
        }
        return rows, metadata
    finally:
        workbook.close()


def build_notes_rows(metadata: dict[str, object], input_path: Path) -> list[str | None]:
    start_date = metadata["start_date"]
    min_date = metadata["min_date"]
    max_date = metadata["max_date"]
    unmapped_status = metadata["unmapped_status"]
    missing_node_cn = metadata["missing_node_cn"]
    detected_node_list = metadata["detected_node_list"]

    return [
        "【输入文件】",
        f"- Input_File: {input_path}",
        f"- Input_File_Name: {input_path.name}",
        f"- Input_Row_Count: {metadata['input_rows']}",
        f"- Date_Parsed_Row_Count: {metadata['dated_rows']}",
        f"- In_Scope_Row_Count(Submission_Date>={start_date.isoformat()}): {metadata['in_scope_rows']}",
        f"- Required_Key_Row_Count(QPO_ID & KP_Flup non-blank): {metadata['rows_with_required_keys']}",
        f"- Mapped_Status_Row_Count: {metadata['mapped_rows']}",
        None,
        "【运行元信息】",
        f"- Date_Range: {min_date.isoformat() if min_date else 'None'} ~ {max_date.isoformat() if max_date else 'None'}",
        f"- Output_Row_Count: {metadata['output_row_count']}",
        f"- All_Row_Count: {metadata['all_row_count']}",
        f"- Unmapped_Status_Breakdown: {'无' if not unmapped_status else '; '.join(f'{k}:{v}' for k, v in sorted(unmapped_status.items()))}",
        f"- Detected_Node_List: {'无' if not detected_node_list else '; '.join(detected_node_list)}",
        None,
        "【口径说明】",
        f"- 默认只统计 Submission_Date >= {start_date.isoformat()} 的数据。",
        "- 输出粒度: KP_Flup × Stat_Month(YYYY-MM)。",
        "- KP_Flup 为空不参与统计。",
        "- 主表包含 ALL 汇总，用于展示每个 KP 在统计区间内累计表现。",
        "- Total_QPO_Count、Success_QPO_Count、Cancel_QPO_Count、InProgress_QPO_Count 均按 QPO_ID 去重统计。",
        "- 强制口径: Total = Success + Cancel + InProgress。",
        "- 若同一 KP_Flup×月份×QPO_ID 出现多种可映射状态，保留首次映射结果并记录冲突数量。",
        None,
        "【QPO_Status 状态映射规则】",
        "- 状态统一映射为 Success / Cancel / InProgress。",
        "- Cancel: ['CSCancel', 'Long time-Cancel', 'SellerCancel', 'WootCancel']",
        "- InProgress: ['Hold', 'New', 'SubmitRBS']",
        "- Success: ['End', 'OnDeal', 'WaitingDeal', 'WaitingPO', 'WaitingPayment', 'WaitingRPO']",
        f"- 未映射状态会跳过统计，并在 Notes 列出: {'无' if not unmapped_status else '; '.join(f'{k}:{v}' for k, v in sorted(unmapped_status.items()))}",
        None,
        "【取消节点分布规则】",
        "- 节点取消分布仅针对 Cancel QPO。",
        "- 节点维度来自 WPO_Current_Node（代码），中文名来自 WPO_Current_Node.1。",
        "- 每个节点输出两列: {Node}_Cancel_QPO_Count, {Node}_Cancel_QPO_Rate。",
        "- {Node}_Cancel_QPO_Rate 分母为 Total_QPO_Count，而不是 Cancel_QPO_Count。",
        f"- 节点中文缺失列表: {'无' if not missing_node_cn else ', '.join(missing_node_cn)}",
        None,
        "【字段要求】",
        "- 必需字段: Submission_Date, QPO_ID, QPO_Status, WPO_Current_Node, WPO_Current_Node.1, KP_Flup",
        None,
        "【自检】",
        f"- 同一 KP_Flup×月份内 QPO 多状态冲突数量(QPO_ID): {metadata['conflict_counter']}",
        "- 即使过滤后无数据，也会输出带双表头的主表和 Notes sheet。",
    ]


def write_main_sheet(
    workbook: Workbook, rows: list[dict[str, object]], metadata: dict[str, object]
) -> None:
    columns_en = list(BASE_COLUMNS_EN)
    columns_cn = list(BASE_COLUMNS_CN)

    for node in metadata["nodes_sorted"]:
        node_cn = metadata["node_cn_map"].get(node, node)
        columns_en.extend([f"{node}_Cancel_QPO_Count", f"{node}_Cancel_QPO_Rate"])
        columns_cn.extend([f"{node_cn}取消QPO数量", f"{node_cn}取消占提报比例"])

    sheet = workbook.active
    sheet.title = "UK205_KP UEA"
    sheet.append(columns_en)
    sheet.append(columns_cn)

    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    style_two_header_sheet(sheet, columns_en, columns_cn)


def write_notes_sheet(
    workbook: Workbook, metadata: dict[str, object], input_path: Path
) -> None:
    sheet = workbook.create_sheet("Notes")
    for line in build_notes_rows(metadata, input_path):
        sheet.append([line])
    sheet.column_dimensions["A"].width = 120
    sheet.freeze_panes = "A1"


def write_workbook(
    output_path: Path,
    rows: list[dict[str, object]],
    metadata: dict[str, object],
    input_path: Path,
) -> None:
    workbook = Workbook()
    write_main_sheet(workbook, rows, metadata)
    write_notes_sheet(workbook, metadata, input_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def generate_report(input_path: Path, output_path: Path, start_date: dt.date) -> None:
    rows, metadata = build_summary(input_path, start_date)
    write_workbook(output_path, rows, metadata, input_path)


def main() -> None:
    args = parse_args()

    input_path = Path(args.input_path) if args.input_path else find_latest_input(BASE_DIR)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )

    try:
        start_date = dt.date.fromisoformat(args.start_date)
    except ValueError as exc:
        raise ValueError(f"无效的 --start-date: {args.start_date}") from exc

    generate_report(input_path, output_path, start_date)
    print("生成完成：")
    print(output_path)


if __name__ == "__main__":
    main()
