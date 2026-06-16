# -*- coding: utf-8 -*-
"""
根据 UK100_工作量 KPI 基础表生成 UK104_QPO_工单处理量分析表。
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

from uk_export_utils import build_default_output_path
from uk_workload_utils import (
    find_latest_input,
    get_required_indexes,
    load_sheet_context,
    normalize_text,
    parse_cell_date,
    round_half_up,
    safe_div,
    style_two_header_sheet,
    write_notes_rows,
)

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK104_QPO_工单处理量分析表_20260430.xlsx"
TICKET_ONLY_TYPE = "工单"

DETAIL_COLUMNS_EN = [
    "Order_ID",
    "Latest_Stat_Date",
    "QPO_Order_Status",
    "QPO_Deal_Cnt",
    "Order_Level",
    "Seller_Code",
    "DCM",
    "KCP",
    "Ticket_Total",
    "TAM1_Total_QPO",
    "TAM2_Total_QPO",
    "TAM3_Total_QPO",
    "TAM4_Total_QPO",
    "TAM5_Total_QPO",
    "TAM1_AvgPerDeal_QPO",
    "TAM2_AvgPerDeal_QPO",
    "TAM3_AvgPerDeal_QPO",
    "TAM4_AvgPerDeal_QPO",
    "TAM5_AvgPerDeal_QPO",
]

DETAIL_COLUMNS_CN = [
    "QPO编号",
    "最新统计日期",
    "QPO订单状态",
    "运行Deal数量",
    "订单等级",
    "客户编号",
    "客户DCM等级",
    "KCP标签",
    "工单数量(按QPO内排重)",
    "TAM1总次数(QPO内求和)",
    "TAM2总次数(QPO内求和)",
    "TAM3总次数(QPO内求和)",
    "TAM4总次数(QPO内求和)",
    "TAM5总次数(QPO内求和)",
    "TAM1平均/Deal(QPO口径)",
    "TAM2平均/Deal(QPO口径)",
    "TAM3平均/Deal(QPO口径)",
    "TAM4平均/Deal(QPO口径)",
    "TAM5平均/Deal(QPO口径)",
]

SUMMARY_ORDER_COLUMNS_EN = [
    "Order_Level",
    "QPO_Cnt",
    "QPO_Share",
    "Deal_Cnt",
    "Deal_Share",
    "Ticket_Cnt",
    "Ticket_Share",
    "TAM1_Total",
    "TAM1_AvgPerTicket",
    "TAM1_AvgPerQPO",
    "TAM1_AvgPerDeal",
    "TAM2_Total",
    "TAM2_AvgPerTicket",
    "TAM2_AvgPerQPO",
    "TAM2_AvgPerDeal",
    "TAM3_Total",
    "TAM3_AvgPerTicket",
    "TAM3_AvgPerQPO",
    "TAM3_AvgPerDeal",
    "TAM4_Total",
    "TAM4_AvgPerTicket",
    "TAM4_AvgPerQPO",
    "TAM4_AvgPerDeal",
    "TAM5_Total",
    "TAM5_AvgPerTicket",
    "TAM5_AvgPerQPO",
    "TAM5_AvgPerDeal",
]

SUMMARY_ORDER_COLUMNS_CN = [
    "订单等级",
    "QPO数量",
    "QPO占比",
    "Deal总量(QPO口径)",
    "Deal占比",
    "工单总量(QPO口径)",
    "工单占比",
    "TAM1总次数",
    "TAM1平均/工单",
    "TAM1平均/QPO",
    "TAM1平均/Deal",
    "TAM2总次数",
    "TAM2平均/工单",
    "TAM2平均/QPO",
    "TAM2平均/Deal",
    "TAM3总次数",
    "TAM3平均/工单",
    "TAM3平均/QPO",
    "TAM3平均/Deal",
    "TAM4总次数",
    "TAM4平均/工单",
    "TAM4平均/QPO",
    "TAM4平均/Deal",
    "TAM5总次数",
    "TAM5平均/工单",
    "TAM5平均/QPO",
    "TAM5平均/Deal",
]

SUMMARY_DIM_COLUMNS_EN = [
    "DIM",
    "Seller_Cnt",
    "Seller_Share",
    "QPO_Cnt",
    "QPO_Share",
    "Deal_Cnt",
    "Deal_Share",
    "Ticket_Cnt",
    "Ticket_Share",
    "TAM1_Total",
    "TAM1_AvgPerTicket",
    "TAM1_AvgPerQPO",
    "TAM1_AvgPerDeal",
    "TAM1_AvgPerSeller",
    "TAM2_Total",
    "TAM2_AvgPerTicket",
    "TAM2_AvgPerQPO",
    "TAM2_AvgPerDeal",
    "TAM2_AvgPerSeller",
    "TAM3_Total",
    "TAM3_AvgPerTicket",
    "TAM3_AvgPerQPO",
    "TAM3_AvgPerDeal",
    "TAM3_AvgPerSeller",
    "TAM4_Total",
    "TAM4_AvgPerTicket",
    "TAM4_AvgPerQPO",
    "TAM4_AvgPerDeal",
    "TAM4_AvgPerSeller",
    "TAM5_Total",
    "TAM5_AvgPerTicket",
    "TAM5_AvgPerQPO",
    "TAM5_AvgPerDeal",
    "TAM5_AvgPerSeller",
]

SUMMARY_DCM_COLUMNS_CN = [
    "客户DCM等级",
    "客户数",
    "客户占比",
    "QPO数量",
    "QPO占比",
    "Deal总量(QPO口径)",
    "Deal占比",
    "工单总量(QPO口径)",
    "工单占比",
    "TAM1总次数",
    "TAM1平均/工单",
    "TAM1平均/QPO",
    "TAM1平均/Deal",
    "TAM1平均/客户",
    "TAM2总次数",
    "TAM2平均/工单",
    "TAM2平均/QPO",
    "TAM2平均/Deal",
    "TAM2平均/客户",
    "TAM3总次数",
    "TAM3平均/工单",
    "TAM3平均/QPO",
    "TAM3平均/Deal",
    "TAM3平均/客户",
    "TAM4总次数",
    "TAM4平均/工单",
    "TAM4平均/QPO",
    "TAM4平均/Deal",
    "TAM4平均/客户",
    "TAM5总次数",
    "TAM5平均/工单",
    "TAM5平均/QPO",
    "TAM5平均/Deal",
    "TAM5平均/客户",
]

SUMMARY_KCP_COLUMNS_CN = [
    "KCP标签",
    "客户数",
    "客户占比",
    "QPO数量",
    "QPO占比",
    "Deal总量(QPO口径)",
    "Deal占比",
    "工单总量(QPO口径)",
    "工单占比",
    "TAM1总次数",
    "TAM1平均/工单",
    "TAM1平均/QPO",
    "TAM1平均/Deal",
    "TAM1平均/客户",
    "TAM2总次数",
    "TAM2平均/工单",
    "TAM2平均/QPO",
    "TAM2平均/Deal",
    "TAM2平均/客户",
    "TAM3总次数",
    "TAM3平均/工单",
    "TAM3平均/QPO",
    "TAM3平均/Deal",
    "TAM3平均/客户",
    "TAM4总次数",
    "TAM4平均/工单",
    "TAM4平均/QPO",
    "TAM4平均/Deal",
    "TAM4平均/客户",
    "TAM5总次数",
    "TAM5平均/工单",
    "TAM5平均/QPO",
    "TAM5平均/Deal",
    "TAM5平均/客户",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 UK100 基础表生成 UK104_QPO 工单处理量分析表。")
    parser.add_argument("--input", dest="input_path", help="指定输入基础表路径")
    parser.add_argument("--output", dest="output_path", help="指定输出 Excel 路径")
    return parser.parse_args()


def numeric_or_zero(value) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = normalize_text(value).replace(",", "")
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def avg_per_deal(total: int, deal_cnt: int) -> float:
    value = safe_div(total, deal_cnt) if deal_cnt else total
    return round_half_up(value, 1)


def build_workbook_data(input_path: Path) -> dict[str, object]:
    workbook, sheet, header_row_1, _, data_start_row = load_sheet_context(input_path)
    try:
        indexes = get_required_indexes(
            header_row_1,
            {
                "Stat_Date": "Stat_Date",
                "Type": "Type",
                "Order_ID": "Order_ID",
                "Ticket_No": "Ticket_No",
                "QPO_Order_Status": "QPO_Order_Status",
                "QPO_BD_Run_Cnt": "QPO_BD_Run_Cnt",
                "Order_Level": "Order_Level",
                "Seller_Code": "seller Code",
                "DCM": "DCM",
                "KCP": "KCP",
                "TAM1_Cnt": "TAM1_Cnt",
                "TAM2_Cnt": "TAM2_Cnt",
                "TAM3_Cnt": "TAM3_Cnt",
                "TAM4_Cnt": "TAM4_Cnt",
                "TAM5_Cnt": "TAM5_Cnt",
            },
        )

        qpo_map: dict[str, dict[str, object]] = {}
        raw_data_rows = max(sheet.max_row - data_start_row + 1, 0)
        ticket_type_rows = 0
        cleaned_rows = 0
        row_number = data_start_row

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            row_type = normalize_text(row[indexes["Type"]])
            if row_type == TICKET_ONLY_TYPE:
                ticket_type_rows += 1

            stat_date = parse_cell_date(row[indexes["Stat_Date"]])
            order_id = normalize_text(row[indexes["Order_ID"]])
            ticket_no = normalize_text(row[indexes["Ticket_No"]])

            if row_type != TICKET_ONLY_TYPE or stat_date is None or not order_id or not ticket_no:
                row_number += 1
                continue

            cleaned_rows += 1
            entry = qpo_map.setdefault(
                order_id,
                {
                    "Order_ID": order_id,
                    "Latest_Stat_Date": stat_date,
                    "QPO_Order_Status": normalize_text(row[indexes["QPO_Order_Status"]]),
                    "QPO_Deal_Cnt": numeric_or_zero(row[indexes["QPO_BD_Run_Cnt"]]),
                    "Order_Level": normalize_text(row[indexes["Order_Level"]]),
                    "Seller_Code": normalize_text(row[indexes["Seller_Code"]]),
                    "DCM": normalize_text(row[indexes["DCM"]]) or None,
                    "KCP": normalize_text(row[indexes["KCP"]]) or None,
                    "Ticket_Set": set(),
                    "TAM1_Total_QPO": 0,
                    "TAM2_Total_QPO": 0,
                    "TAM3_Total_QPO": 0,
                    "TAM4_Total_QPO": 0,
                    "TAM5_Total_QPO": 0,
                    "Latest_Row_Number": row_number,
                },
            )

            entry["Ticket_Set"].add(ticket_no)
            entry["TAM1_Total_QPO"] += numeric_or_zero(row[indexes["TAM1_Cnt"]])
            entry["TAM2_Total_QPO"] += numeric_or_zero(row[indexes["TAM2_Cnt"]])
            entry["TAM3_Total_QPO"] += numeric_or_zero(row[indexes["TAM3_Cnt"]])
            entry["TAM4_Total_QPO"] += numeric_or_zero(row[indexes["TAM4_Cnt"]])
            entry["TAM5_Total_QPO"] += numeric_or_zero(row[indexes["TAM5_Cnt"]])

            latest_key = (entry["Latest_Stat_Date"], entry["Latest_Row_Number"])
            current_key = (stat_date, row_number)
            if current_key >= latest_key:
                entry["Latest_Stat_Date"] = stat_date
                entry["QPO_Order_Status"] = normalize_text(row[indexes["QPO_Order_Status"]])
                entry["QPO_Deal_Cnt"] = numeric_or_zero(row[indexes["QPO_BD_Run_Cnt"]])
                entry["Order_Level"] = normalize_text(row[indexes["Order_Level"]])
                entry["Seller_Code"] = normalize_text(row[indexes["Seller_Code"]])
                entry["DCM"] = normalize_text(row[indexes["DCM"]]) or None
                entry["KCP"] = normalize_text(row[indexes["KCP"]]) or None
                entry["Latest_Row_Number"] = row_number

            row_number += 1

        detail_rows: list[dict[str, object]] = []
        for order_id in sorted(qpo_map):
            entry = qpo_map[order_id]
            deal_cnt = int(entry["QPO_Deal_Cnt"])
            ticket_total = len(entry["Ticket_Set"])
            detail_rows.append(
                {
                    "Order_ID": entry["Order_ID"],
                    "Latest_Stat_Date": entry["Latest_Stat_Date"],
                    "QPO_Order_Status": entry["QPO_Order_Status"],
                    "QPO_Deal_Cnt": deal_cnt,
                    "Order_Level": entry["Order_Level"],
                    "Seller_Code": entry["Seller_Code"],
                    "DCM": entry["DCM"],
                    "KCP": entry["KCP"],
                    "Ticket_Total": ticket_total,
                    "TAM1_Total_QPO": entry["TAM1_Total_QPO"],
                    "TAM2_Total_QPO": entry["TAM2_Total_QPO"],
                    "TAM3_Total_QPO": entry["TAM3_Total_QPO"],
                    "TAM4_Total_QPO": entry["TAM4_Total_QPO"],
                    "TAM5_Total_QPO": entry["TAM5_Total_QPO"],
                    "TAM1_AvgPerDeal_QPO": avg_per_deal(entry["TAM1_Total_QPO"], deal_cnt),
                    "TAM2_AvgPerDeal_QPO": avg_per_deal(entry["TAM2_Total_QPO"], deal_cnt),
                    "TAM3_AvgPerDeal_QPO": avg_per_deal(entry["TAM3_Total_QPO"], deal_cnt),
                    "TAM4_AvgPerDeal_QPO": avg_per_deal(entry["TAM4_Total_QPO"], deal_cnt),
                    "TAM5_AvgPerDeal_QPO": avg_per_deal(entry["TAM5_Total_QPO"], deal_cnt),
                }
            )

        return {
            "detail_rows": detail_rows,
            "raw_data_rows": raw_data_rows,
            "ticket_type_rows": ticket_type_rows,
            "cleaned_rows": cleaned_rows,
        }
    finally:
        workbook.close()


def build_order_summary(detail_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    category_order = ["A", "AA", "B", "C", "D"]
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in detail_rows:
        grouped[normalize_text(row["Order_Level"])].append(row)
    categories = [item for item in category_order if item in grouped] + sorted(
        item for item in grouped if item not in category_order
    )
    return [
        build_summary_row(item, grouped[item], detail_rows, include_seller=False, dim_key="Order_Level")
        for item in categories
    ]


def build_dimension_summary(
    detail_rows: list[dict[str, object]],
    *,
    value_getter,
    category_order: list[str],
    empty_label: str,
) -> list[dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in detail_rows:
        key = value_getter(row) or empty_label
        grouped[key].append(row)
    categories = [item for item in category_order if item in grouped] + sorted(
        item for item in grouped if item not in category_order
    )
    return [build_summary_row(item, grouped[item], detail_rows, include_seller=True, dim_key="DIM") for item in categories]


def build_summary_row(
    category: str,
    rows: list[dict[str, object]],
    all_rows: list[dict[str, object]],
    *,
    include_seller: bool,
    dim_key: str,
) -> dict[str, object]:
    total_qpo = len(all_rows)
    total_deal = sum(int(row["QPO_Deal_Cnt"]) for row in all_rows)
    total_ticket = sum(int(row["Ticket_Total"]) for row in all_rows)
    total_seller = len(
        {normalize_text(row["Seller_Code"]) for row in all_rows if normalize_text(row["Seller_Code"])}
    )

    qpo_cnt = len(rows)
    deal_cnt = sum(int(row["QPO_Deal_Cnt"]) for row in rows)
    ticket_cnt = sum(int(row["Ticket_Total"]) for row in rows)
    seller_cnt = len({normalize_text(row["Seller_Code"]) for row in rows if normalize_text(row["Seller_Code"])})

    result: dict[str, object] = {
        dim_key: category,
        "QPO_Cnt": qpo_cnt,
        "QPO_Share": safe_div(qpo_cnt, total_qpo),
        "Deal_Cnt": deal_cnt,
        "Deal_Share": safe_div(deal_cnt, total_deal),
        "Ticket_Cnt": ticket_cnt,
        "Ticket_Share": safe_div(ticket_cnt, total_ticket),
    }

    if include_seller:
        result["Seller_Cnt"] = seller_cnt
        result["Seller_Share"] = safe_div(seller_cnt, total_seller)

    for tam in range(1, 6):
        total = sum(int(row[f"TAM{tam}_Total_QPO"]) for row in rows)
        result[f"TAM{tam}_Total"] = total
        result[f"TAM{tam}_AvgPerTicket"] = round_half_up(safe_div(total, ticket_cnt), 1)
        result[f"TAM{tam}_AvgPerQPO"] = round_half_up(safe_div(total, qpo_cnt), 1)
        result[f"TAM{tam}_AvgPerDeal"] = round_half_up(safe_div(total, deal_cnt) if deal_cnt else total, 1)
        if include_seller:
            result[f"TAM{tam}_AvgPerSeller"] = round_half_up(safe_div(total, seller_cnt), 1)

    return result


def normalize_share_sum(value: float) -> float | int:
    return 1 if abs(value - 1.0) < 1e-12 else value


def build_notes_rows(
    metadata: dict[str, object],
    order_summary: list[dict[str, object]],
    dcm_summary: list[dict[str, object]],
    kcp_summary: list[dict[str, object]],
) -> list[tuple[object, ...]]:
    return [
        ("UK104_QPO 工单处理量分析表 - Notes / 口径说明与自检", None, None, None, None),
        (None, None, None, None, None),
        ("Field_EN", "字段中文", "业务含义", "计算方式", "补充说明"),
        ("Order_ID", "QPO编号", "QPO订单唯一编号", "来自UK100 Data（Type=工单）", None),
        ("Latest_Stat_Date", "最新统计日期", "该QPO最新的统计日期", "同一Order_ID内取Stat_Date最大值对应行", None),
        ("QPO_Order_Status", "QPO订单状态", "QPO订单状态", "取最新日期对应行的QPO_Order_Status", None),
        ("QPO_Deal_Cnt", "运行Deal数量", "运行Deal数量（实时字段）", "取最新日期对应行的QPO_BD_Run_Cnt（不累加）", "Deal=0不兜底为1"),
        ("Order_Level", "订单等级", "订单等级", "取最新日期对应行的Order_Level", None),
        ("Seller_Code", "客户编号", "客户编号", "取最新日期对应行的 seller Code", None),
        ("DCM", "客户DCM等级", "客户DCM等级", "明细层允许为空；汇总层空值归类为“未标记DCM”", None),
        ("KCP", "KCP标签", "KCP标签", "取最新日期对应行的KCP；空值在汇总层归类为“未标记KCP”", None),
        ("Ticket_Total", "工单数量(按QPO内排重)", "QPO内工单数量", "仅Type=工单，按Order_ID分组，对Ticket_No去重计数(nunique)", "同一Ticket_No可能属于多个QPO，跨QPO允许重复计数"),
        ("TAM1_Total_QPO", "TAM1总次数(QPO内求和)", "QPO内TAM次数汇总", "仅Type=工单，按Order_ID分组求和(sum)", None),
        ("TAM2_Total_QPO", "TAM2总次数(QPO内求和)", "QPO内TAM次数汇总", "仅Type=工单，按Order_ID分组求和(sum)", None),
        ("TAM3_Total_QPO", "TAM3总次数(QPO内求和)", "QPO内TAM次数汇总", "仅Type=工单，按Order_ID分组求和(sum)", None),
        ("TAM4_Total_QPO", "TAM4总次数(QPO内求和)", "QPO内TAM次数汇总", "仅Type=工单，按Order_ID分组求和(sum)", None),
        ("TAM5_Total_QPO", "TAM5总次数(QPO内求和)", "QPO内TAM次数汇总", "仅Type=工单，按Order_ID分组求和(sum)", None),
        ("TAM1_AvgPerDeal_QPO", "TAM1平均/Deal(QPO口径)", "QPO内平均到Deal的TAM次数", "若QPO_Deal_Cnt>0：TAMi_Total_QPO / QPO_Deal_Cnt；若=0：直接取TAMi_Total_QPO", None),
        ("TAM2_AvgPerDeal_QPO", "TAM2平均/Deal(QPO口径)", "QPO内平均到Deal的TAM次数", "若QPO_Deal_Cnt>0：TAMi_Total_QPO / QPO_Deal_Cnt；若=0：直接取TAMi_Total_QPO", None),
        ("TAM3_AvgPerDeal_QPO", "TAM3平均/Deal(QPO口径)", "QPO内平均到Deal的TAM次数", "若QPO_Deal_Cnt>0：TAMi_Total_QPO / QPO_Deal_Cnt；若=0：直接取TAMi_Total_QPO", None),
        ("TAM4_AvgPerDeal_QPO", "TAM4平均/Deal(QPO口径)", "QPO内平均到Deal的TAM次数", "若QPO_Deal_Cnt>0：TAMi_Total_QPO / QPO_Deal_Cnt；若=0：直接取TAMi_Total_QPO", None),
        ("TAM5_AvgPerDeal_QPO", "TAM5平均/Deal(QPO口径)", "QPO内平均到Deal的TAM次数", "若QPO_Deal_Cnt>0：TAMi_Total_QPO / QPO_Deal_Cnt；若=0：直接取TAMi_Total_QPO", None),
        ("QPO_Cnt", "QPO数量", "分组下QPO数量", "对UK104_QPO按分组字段汇总：nunique(Order_ID)", None),
        ("QPO_Share", "QPO占比", "占比", "QPO / 对应总QPO（全表汇总）", None),
        ("Deal_Cnt", "Deal总量(QPO口径)", "分组下Deal总量（QPO口径）", "对UK104_QPO汇总：sum(QPO_Deal_Cnt)", None),
        ("Deal_Share", "Deal占比", "占比", "Deal / 对应总Deal（全表汇总）", None),
        ("Ticket_Cnt", "工单总量(QPO口径)", "分组下工单总量（QPO口径）", "对UK104_QPO汇总：sum(Ticket_Total)", "QPO口径工单量≠全局唯一工单数"),
        ("Ticket_Share", "工单占比", "占比", "Ticket / 对应总Ticket（全表汇总）", None),
        ("TAM1_Total", "TAM1总次数", "分组下TAM总次数", "sum(TAM1_Total_QPO)", None),
        ("TAM1_AvgPerTicket", "TAM1平均/工单", "平均到每个工单的TAM次数", "TAM1_Total / Ticket_Cnt（Ticket_Cnt=0时取0）", None),
        ("TAM1_AvgPerQPO", "TAM1平均/QPO", "平均到每个QPO的TAM次数", "TAM1_Total / QPO_Cnt（QPO_Cnt=0时取0）", None),
        ("TAM1_AvgPerDeal", "TAM1平均/Deal", "平均到每个Deal的TAM次数", "若Deal_Cnt>0：TAM1_Total / Deal_Cnt；若=0：直接取TAM1_Total", None),
        ("TAM2_Total", "TAM2总次数", "分组下TAM总次数", "sum(TAM2_Total_QPO)", None),
        ("TAM2_AvgPerTicket", "TAM2平均/工单", "平均到每个工单的TAM次数", "TAM2_Total / Ticket_Cnt（Ticket_Cnt=0时取0）", None),
        ("TAM2_AvgPerQPO", "TAM2平均/QPO", "平均到每个QPO的TAM次数", "TAM2_Total / QPO_Cnt（QPO_Cnt=0时取0）", None),
        ("TAM2_AvgPerDeal", "TAM2平均/Deal", "平均到每个Deal的TAM次数", "若Deal_Cnt>0：TAM2_Total / Deal_Cnt；若=0：直接取TAM2_Total", None),
        ("TAM3_Total", "TAM3总次数", "分组下TAM总次数", "sum(TAM3_Total_QPO)", None),
        ("TAM3_AvgPerTicket", "TAM3平均/工单", "平均到每个工单的TAM次数", "TAM3_Total / Ticket_Cnt（Ticket_Cnt=0时取0）", None),
        ("TAM3_AvgPerQPO", "TAM3平均/QPO", "平均到每个QPO的TAM次数", "TAM3_Total / QPO_Cnt（QPO_Cnt=0时取0）", None),
        ("TAM3_AvgPerDeal", "TAM3平均/Deal", "平均到每个Deal的TAM次数", "若Deal_Cnt>0：TAM3_Total / Deal_Cnt；若=0：直接取TAM3_Total", None),
        ("TAM4_Total", "TAM4总次数", "分组下TAM总次数", "sum(TAM4_Total_QPO)", None),
        ("TAM4_AvgPerTicket", "TAM4平均/工单", "平均到每个工单的TAM次数", "TAM4_Total / Ticket_Cnt（Ticket_Cnt=0时取0）", None),
        ("TAM4_AvgPerQPO", "TAM4平均/QPO", "平均到每个QPO的TAM次数", "TAM4_Total / QPO_Cnt（QPO_Cnt=0时取0）", None),
        ("TAM4_AvgPerDeal", "TAM4平均/Deal", "平均到每个Deal的TAM次数", "若Deal_Cnt>0：TAM4_Total / Deal_Cnt；若=0：直接取TAM4_Total", None),
        ("TAM5_Total", "TAM5总次数", "分组下TAM总次数", "sum(TAM5_Total_QPO)", None),
        ("TAM5_AvgPerTicket", "TAM5平均/工单", "平均到每个工单的TAM次数", "TAM5_Total / Ticket_Cnt（Ticket_Cnt=0时取0）", None),
        ("TAM5_AvgPerQPO", "TAM5平均/QPO", "平均到每个QPO的TAM次数", "TAM5_Total / QPO_Cnt（QPO_Cnt=0时取0）", None),
        ("TAM5_AvgPerDeal", "TAM5平均/Deal", "平均到每个Deal的TAM次数", "若Deal_Cnt>0：TAM5_Total / Deal_Cnt；若=0：直接取TAM5_Total", None),
        ("Seller_Cnt", "客户数", "分组下客户数", "对UK104_QPO汇总：nunique(Seller_Code)", None),
        ("Seller_Share", "客户占比", "占比", "Seller / 对应总Seller（全表汇总）", None),
        ("TAM1_AvgPerSeller", "TAM1平均/客户", "平均到每个客户的TAM次数", "TAM1_Total / Seller_Cnt（Seller_Cnt=0时取0）", None),
        ("TAM2_AvgPerSeller", "TAM2平均/客户", "平均到每个客户的TAM次数", "TAM2_Total / Seller_Cnt（Seller_Cnt=0时取0）", None),
        ("TAM3_AvgPerSeller", "TAM3平均/客户", "平均到每个客户的TAM次数", "TAM3_Total / Seller_Cnt（Seller_Cnt=0时取0）", None),
        ("TAM4_AvgPerSeller", "TAM4平均/客户", "平均到每个客户的TAM次数", "TAM4_Total / Seller_Cnt（Seller_Cnt=0时取0）", None),
        ("TAM5_AvgPerSeller", "TAM5平均/客户", "平均到每个客户的TAM次数", "TAM5_Total / Seller_Cnt（Seller_Cnt=0时取0）", None),
        (None, None, None, None, None),
        (None, None, None, None, None),
        ("自检与对账", None, None, None, None),
        ("Item", "Value", None, None, None),
        ("原始Data行数", metadata["raw_data_rows"], None, None, None),
        ("Type=工单行数", metadata["ticket_type_rows"], None, None, None),
        ("清洗后工单行数（去除Order_ID或Ticket_No为空）", metadata["cleaned_rows"], None, None, None),
        ("UK104_QPO行数（QPO数）", len(metadata["detail_rows"]), None, None, None),
        ("Ticket_Total求和（QPO口径工单量）", sum(int(row["Ticket_Total"]) for row in metadata["detail_rows"]), None, None, None),
        ("QPO_Deal_Cnt求和（QPO口径Deal总量）", sum(int(row["QPO_Deal_Cnt"]) for row in metadata["detail_rows"]), None, None, None),
        ("按订单等级汇总：QPO_Cnt求和", sum(int(row["QPO_Cnt"]) for row in order_summary), None, None, None),
        ("按客户DCM汇总：QPO_Cnt求和", sum(int(row["QPO_Cnt"]) for row in dcm_summary), None, None, None),
        ("按KCP标签汇总：QPO_Cnt求和", sum(int(row["QPO_Cnt"]) for row in kcp_summary), None, None, None),
        ("按订单等级汇总：QPO_Share之和", normalize_share_sum(sum(float(row["QPO_Share"]) for row in order_summary)), None, None, None),
        ("按客户DCM汇总：QPO_Share之和", normalize_share_sum(sum(float(row["QPO_Share"]) for row in dcm_summary)), None, None, None),
        ("按KCP标签汇总：QPO_Share之和", normalize_share_sum(sum(float(row["QPO_Share"]) for row in kcp_summary)), None, None, None),
        (None, None, None, None, None),
        (None, None, None, None, None),
        ("关键口径风险提示", None, None, None, None),
        ("同一个Ticket_No可能同时关联多个Order_ID（QPO）。因此：", None, None, None, None),
        ("1）UK104_QPO 的 Ticket_Total 在不同QPO之间允许重复计数；", None, None, None, None),
        ("2）各汇总表中 Ticket_Cnt（sum(Ticket_Total)）为“QPO口径工单量”，不等于全局唯一工单数。", None, None, None, None),
    ]


def write_sheet(
    sheet,
    columns_en: list[str],
    columns_cn: list[str],
    rows: list[dict[str, object]],
    *,
    date_columns: set[str] | None = None,
) -> None:
    sheet.append(columns_en)
    sheet.append(columns_cn)
    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    style_two_header_sheet(
        sheet,
        columns_en,
        columns_cn,
        date_columns=date_columns or set(),
        percent_columns={column for column in columns_en if "Share" in column},
        integer_columns={
            column
            for column in columns_en
            if column.endswith("_Cnt")
            or "Total" in column
            or column in {"QPO_Cnt", "Deal_Cnt", "Ticket_Cnt", "Seller_Cnt"}
        },
        decimal_formats={column: "0.0" for column in columns_en if "AvgPer" in column},
    )


def write_workbook(output_path: Path, metadata: dict[str, object]) -> None:
    detail_rows = metadata["detail_rows"]

    order_summary = build_order_summary(detail_rows)
    dcm_summary = build_dimension_summary(
        detail_rows,
        value_getter=lambda row: normalize_text(row["DCM"]),
        category_order=["A", "B", "C", "VCP", "未标记DCM"],
        empty_label="未标记DCM",
    )
    kcp_summary = build_dimension_summary(
        detail_rows,
        value_getter=lambda row: normalize_text(row["KCP"]),
        category_order=["No", "Yes", "未标记KCP"],
        empty_label="未标记KCP",
    )

    workbook = Workbook()

    detail_sheet = workbook.active
    detail_sheet.title = "UK104_QPO"
    write_sheet(detail_sheet, DETAIL_COLUMNS_EN, DETAIL_COLUMNS_CN, detail_rows, date_columns={"Latest_Stat_Date"})

    order_sheet = workbook.create_sheet("按订单等级汇总")
    write_sheet(order_sheet, SUMMARY_ORDER_COLUMNS_EN, SUMMARY_ORDER_COLUMNS_CN, order_summary)

    dcm_sheet = workbook.create_sheet("按客户DCM汇总")
    dcm_columns_en = ["DCM"] + SUMMARY_DIM_COLUMNS_EN[1:]
    dcm_rows: list[dict[str, object]] = []
    for row in dcm_summary:
        dcm_row = dict(row)
        dcm_row["DCM"] = dcm_row.pop("DIM")
        dcm_rows.append(dcm_row)
    write_sheet(dcm_sheet, dcm_columns_en, SUMMARY_DCM_COLUMNS_CN, dcm_rows)

    kcp_sheet = workbook.create_sheet("按KCP标签汇总")
    kcp_columns_en = ["KCP"] + SUMMARY_DIM_COLUMNS_EN[1:]
    kcp_rows: list[dict[str, object]] = []
    for row in kcp_summary:
        kcp_row = dict(row)
        kcp_row["KCP"] = kcp_row.pop("DIM")
        kcp_rows.append(kcp_row)
    write_sheet(kcp_sheet, kcp_columns_en, SUMMARY_KCP_COLUMNS_CN, kcp_rows)

    notes_sheet = workbook.create_sheet("Notes")
    write_notes_rows(notes_sheet, build_notes_rows(metadata, order_summary, dcm_summary, kcp_summary))

    workbook.save(output_path)
    workbook.close()


def main() -> None:
    args = parse_args()

    input_path = Path(args.input_path) if args.input_path else find_latest_input(BASE_DIR)
    if not input_path.exists():
        raise FileNotFoundError(f"未找到输入文件: {input_path}")

    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    metadata = build_workbook_data(input_path)
    write_workbook(output_path, metadata)

    print("结果已生成：")
    print(output_path)


if __name__ == "__main__":
    main()
