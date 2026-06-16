# -*- coding: utf-8 -*-
"""
根据 UK100_工作量 KPI 基础表生成 UK102_UEA_月度工作量价值结构报表。
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

from uk_export_utils import build_default_output_path
from uk_workload_utils import (
    department_sort_key,
    find_latest_input,
    get_required_indexes,
    has_value,
    load_sheet_context,
    month_sort_key,
    normalize_text,
    parse_cell_date,
    safe_div,
    style_two_header_sheet,
    write_notes_rows,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK102_UEA_月度工作量价值结构_20260430.xlsx"
RESULT_SHEET_NAME = "Result"

COLUMNS_EN = [
    "Stat_Month",
    "Department",
    "UEA",
    "Total_Case_Cnt",
    "DCM_VCP_Case",
    "DCM_VCP_Rate",
    "DCM_A_Case",
    "DCM_A_Rate",
    "DCM_B_Case",
    "DCM_B_Rate",
    "DCM_C_Case",
    "DCM_C_Rate",
    "KCP_Case",
    "KCP_Rate",
    "Non_KCP_Case",
    "Non_KCP_Rate",
    "Order_A_Case",
    "Order_A_Rate",
    "Order_B_Case",
    "Order_B_Rate",
    "Order_C_Case",
    "Order_C_Rate",
    "Order_D_Case",
    "Order_D_Rate",
    "High_Value_Order_Case",
    "High_Value_Order_Rate",
    "Low_Value_Workload_Flag",
]

COLUMNS_CN = [
    "统计月份(YYYYMM)",
    "部门",
    "跟进人编号",
    "工单总数（不去重）",
    "DCM-VCP客户工单数",
    "DCM-VCP客户占比",
    "DCM-A客户工单数",
    "DCM-A客户占比",
    "DCM-B客户工单数",
    "DCM-B客户占比",
    "DCM-C客户工单数",
    "DCM-C客户占比",
    "KCP客户工单数",
    "KCP客户占比",
    "非KCP客户工单数",
    "非KCP客户占比",
    "A等级订单工单数",
    "A等级订单占比",
    "B等级订单工单数",
    "B等级订单占比",
    "C等级订单工单数",
    "C等级订单占比",
    "D等级订单工单数",
    "D等级订单占比",
    "高价值订单工单数(A/B/C)",
    "高价值订单占比",
    "是否低价值消耗（按订单）",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="根据 UK100 基础表生成 UK102 月度工作量价值结构。")
    parser.add_argument("--input", dest="input_path", help="输入基础表路径；不传则自动识别最新 UK100 文件。")
    parser.add_argument("--output", dest="output_path", help="输出 Excel 路径；不传则自动生成带时间戳文件名。")
    return parser.parse_args()


def is_ticket_like_type(value: object) -> bool:
    text = normalize_text(value).strip().lower()
    if not text:
        return False

    compact = text.replace(" ", "").replace("_", "").replace("-", "")
    email_keywords = ("email", "mail", "邮件")
    if any(keyword in compact for keyword in email_keywords):
        return False

    return True


def normalize_dcm(value: object) -> str:
    text = normalize_text(value).strip().upper()
    if not text:
        return "C"
    if "VCP" in text:
        return "VCP"
    if text.startswith("A"):
        return "A"
    if text.startswith("B"):
        return "B"
    if text.startswith("C"):
        return "C"
    return "C"


def normalize_order_group(value: object) -> str:
    text = normalize_text(value).strip().upper()
    if text.startswith("A"):
        return "A"
    if text.startswith("B"):
        return "B"
    if text.startswith("C"):
        return "C"
    return "D"


def is_kcp(value: object) -> bool:
    text = normalize_text(value).strip().upper()
    return text in {"YES", "Y", "TRUE", "1", "是", "KCP"}


def build_output_row(month: str, department: str, uea: str, values: dict[str, int]) -> dict[str, object]:
    total = values["Total_Case_Cnt"]
    high_value = values["Order_A_Case"] + values["Order_B_Case"] + values["Order_C_Case"]
    high_rate = safe_div(high_value, total)

    return {
        "Stat_Month": month,
        "Department": department,
        "UEA": uea,
        "Total_Case_Cnt": total,
        "DCM_VCP_Case": values["DCM_VCP_Case"],
        "DCM_VCP_Rate": safe_div(values["DCM_VCP_Case"], total),
        "DCM_A_Case": values["DCM_A_Case"],
        "DCM_A_Rate": safe_div(values["DCM_A_Case"], total),
        "DCM_B_Case": values["DCM_B_Case"],
        "DCM_B_Rate": safe_div(values["DCM_B_Case"], total),
        "DCM_C_Case": values["DCM_C_Case"],
        "DCM_C_Rate": safe_div(values["DCM_C_Case"], total),
        "KCP_Case": values["KCP_Case"],
        "KCP_Rate": safe_div(values["KCP_Case"], total),
        "Non_KCP_Case": values["Non_KCP_Case"],
        "Non_KCP_Rate": safe_div(values["Non_KCP_Case"], total),
        "Order_A_Case": values["Order_A_Case"],
        "Order_A_Rate": safe_div(values["Order_A_Case"], total),
        "Order_B_Case": values["Order_B_Case"],
        "Order_B_Rate": safe_div(values["Order_B_Case"], total),
        "Order_C_Case": values["Order_C_Case"],
        "Order_C_Rate": safe_div(values["Order_C_Case"], total),
        "Order_D_Case": values["Order_D_Case"],
        "Order_D_Rate": safe_div(values["Order_D_Case"], total),
        "High_Value_Order_Case": high_value,
        "High_Value_Order_Rate": high_rate,
        "Low_Value_Workload_Flag": "是" if high_rate < 0.4 else "否",
    }


def build_summary(input_path: Path) -> list[dict[str, object]]:
    workbook, sheet, header_row_1, _, data_start_row = load_sheet_context(input_path)
    try:
        indexes = get_required_indexes(
            header_row_1,
            {
                "Stat_Date": "Stat_Date",
                "Department": "Department",
                "UEA": "UEA",
                "Type": "Type",
                "Order_ID": "Order_ID",
                "Seller_Code": "seller Code",
                "DCM": "DCM",
                "KCP": "KCP",
                "Order_Level": "Order_Level",
            },
        )

        grouped: dict[tuple[str, str, str], dict[str, int]] = defaultdict(
            lambda: {
                "Total_Case_Cnt": 0,
                "DCM_VCP_Case": 0,
                "DCM_A_Case": 0,
                "DCM_B_Case": 0,
                "DCM_C_Case": 0,
                "KCP_Case": 0,
                "Non_KCP_Case": 0,
                "Order_A_Case": 0,
                "Order_B_Case": 0,
                "Order_C_Case": 0,
                "Order_D_Case": 0,
            }
        )
        months: set[str] = set()
        departments: set[str] = set()

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            stat_date = parse_cell_date(row[indexes["Stat_Date"]])
            department = normalize_text(row[indexes["Department"]]).strip()
            uea = normalize_text(row[indexes["UEA"]]).strip()
            row_type_value = row[indexes["Type"]]
            order_id_value = row[indexes["Order_ID"]]
            seller_code_value = row[indexes["Seller_Code"]]

            if stat_date is None:
                continue
            if not department or not uea:
                continue
            if not is_ticket_like_type(row_type_value):
                continue
            if not has_value(order_id_value) or not normalize_text(order_id_value).strip():
                continue
            if not has_value(seller_code_value) or not normalize_text(seller_code_value).strip():
                continue

            month = stat_date.strftime("%Y%m")
            months.add(month)
            departments.add(department)

            key = (month, department, uea)
            current = grouped[key]
            current["Total_Case_Cnt"] += 1
            current[f"DCM_{normalize_dcm(row[indexes['DCM']])}_Case"] += 1

            if is_kcp(row[indexes["KCP"]]):
                current["KCP_Case"] += 1
            else:
                current["Non_KCP_Case"] += 1

            current[f"Order_{normalize_order_group(row[indexes['Order_Level']])}_Case"] += 1

        rows: list[dict[str, object]] = []
        for month in sorted(months, key=month_sort_key):
            for department in sorted(departments, key=department_sort_key):
                person_keys = sorted(
                    [key for key in grouped if key[0] == month and key[1] == department],
                    key=lambda item: item[2],
                )
                if not person_keys:
                    continue

                dept_totals: dict[str, int] = defaultdict(int)
                for key in person_keys:
                    for metric, value in grouped[key].items():
                        dept_totals[metric] += value

                for key in person_keys:
                    rows.append(build_output_row(month, department, key[2], grouped[key]))

                rows.append(build_output_row(month, department, "部门汇总", dept_totals))

        return rows
    finally:
        workbook.close()


def build_notes_rows() -> list[tuple[object, ...]]:
    return [
        ("Field (EN)", "字段（中文）", "统计方式 / 计算口径", "业务含义 / 管理解读"),
        ("Stat_Month", "统计月份(YYYYMM)", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按工单统计日期（Stat_Date）归属到自然月，格式YYYYMM。"),
        ("Department", "部门", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "工单跟进人所属部门维度。"),
        ("UEA", "跟进人编号", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "工单跟进人编号；“部门汇总”为同月同部门汇总行。"),
        ("Total_Case_Cnt", "工单总数（不去重）", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "当月该维度下满足过滤条件的工单总数（不去重），作为占比分母。"),
        ("DCM_VCP_Case", "DCM-VCP客户工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按客户DCM分层（VCP/A/B/C）对应的工单数量；DCM空值归入C。"),
        ("DCM_VCP_Rate", "DCM-VCP客户占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应DCM分层工单数在工单总数中的占比，用于观察客户结构。"),
        ("DCM_A_Case", "DCM-A客户工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按客户DCM分层（VCP/A/B/C）对应的工单数量；DCM空值归入C。"),
        ("DCM_A_Rate", "DCM-A客户占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应DCM分层工单数在工单总数中的占比，用于观察客户结构。"),
        ("DCM_B_Case", "DCM-B客户工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按客户DCM分层（VCP/A/B/C）对应的工单数量；DCM空值归入C。"),
        ("DCM_B_Rate", "DCM-B客户占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应DCM分层工单数在工单总数中的占比，用于观察客户结构。"),
        ("DCM_C_Case", "DCM-C客户工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按客户DCM分层（VCP/A/B/C）对应的工单数量；DCM空值归入C。"),
        ("DCM_C_Rate", "DCM-C客户占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应DCM分层工单数在工单总数中的占比，用于观察客户结构。"),
        ("KCP_Case", "KCP客户工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按KCP客户标识拆分的工单数量，用于衡量KCP客户工作量占用。"),
        ("KCP_Rate", "KCP客户占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应KCP/非KCP工单数在工单总数中的占比。"),
        ("Non_KCP_Case", "非KCP客户工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按KCP客户标识拆分的工单数量，用于衡量KCP客户工作量占用。"),
        ("Non_KCP_Rate", "非KCP客户占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应KCP/非KCP工单数在工单总数中的占比。"),
        ("Order_A_Case", "A等级订单工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按订单等级（A/B/C/D）归类后的工单数量；非A/B/C开头归入D。"),
        ("Order_A_Rate", "A等级订单占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应订单等级工单数在工单总数中的占比，用于结构分析。"),
        ("Order_B_Case", "B等级订单工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按订单等级（A/B/C/D）归类后的工单数量；非A/B/C开头归入D。"),
        ("Order_B_Rate", "B等级订单占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应订单等级工单数在工单总数中的占比，用于结构分析。"),
        ("Order_C_Case", "C等级订单工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按订单等级（A/B/C/D）归类后的工单数量；非A/B/C开头归入D。"),
        ("Order_C_Rate", "C等级订单占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应订单等级工单数在工单总数中的占比，用于结构分析。"),
        ("Order_D_Case", "D等级订单工单数", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "按订单等级（A/B/C/D）归类后的工单数量；非A/B/C开头归入D。"),
        ("Order_D_Rate", "D等级订单占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "对应订单等级工单数在工单总数中的占比，用于结构分析。"),
        ("High_Value_Order_Case", "高价值订单工单数(A/B/C)", "满足过滤条件的工单数量（Order_ID非空且客户编号非空），工单不去重；按字段条件筛选计数", "高价值订单（A/B/C）对应的工单数量。"),
        ("High_Value_Order_Rate", "高价值订单占比", "对应 *_Case ÷ Total_Case_Cnt（工单不去重）", "高价值订单工单占比，用于衡量投入产出结构（高价值占比越高越理想）。"),
        ("Low_Value_Workload_Flag", "是否低价值消耗（按订单）", "High_Value_Order_Rate < 40% → 是", "当高价值订单占比低于40%时标记为“是”，提示该维度可能存在低价值消耗风险。"),
    ]


def write_workbook(output_path: Path, rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = RESULT_SHEET_NAME
        sheet.append(COLUMNS_EN)
        sheet.append(COLUMNS_CN)

        for row in rows:
            sheet.append([row.get(column) for column in COLUMNS_EN])

        style_two_header_sheet(
            sheet,
            COLUMNS_EN,
            COLUMNS_CN,
            percent_columns={
                "DCM_VCP_Rate",
                "DCM_A_Rate",
                "DCM_B_Rate",
                "DCM_C_Rate",
                "KCP_Rate",
                "Non_KCP_Rate",
                "Order_A_Rate",
                "Order_B_Rate",
                "Order_C_Rate",
                "Order_D_Rate",
                "High_Value_Order_Rate",
            },
            integer_columns={
                "Total_Case_Cnt",
                "DCM_VCP_Case",
                "DCM_A_Case",
                "DCM_B_Case",
                "DCM_C_Case",
                "KCP_Case",
                "Non_KCP_Case",
                "Order_A_Case",
                "Order_B_Case",
                "Order_C_Case",
                "Order_D_Case",
                "High_Value_Order_Case",
            },
        )

        notes_sheet = workbook.create_sheet("Notes")
        write_notes_rows(notes_sheet, build_notes_rows())

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_path) if args.input_path else find_latest_input(BASE_DIR)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )

    rows = build_summary(input_path)
    write_workbook(output_path, rows)

    print("结果已生成：")
    print(output_path)


if __name__ == "__main__":
    main()
