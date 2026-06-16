# -*- coding: utf-8 -*-
"""
根据 UK200 基础表生成 UK209_重复提报与策略复杂度分析.xlsx。
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from collections import defaultdict, deque
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from uk_export_utils import build_default_output_path
from uk_funnel_utils import (
    detect_data_sheet,
    find_latest_input,
    get_required_indexes,
    month_sort_key,
    normalize_text,
    parse_cell_date,
    safe_div,
    style_two_header_sheet,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK209_重复提报与策略复杂度分析.xlsx"
REPORT_CODE = "UK209"
REPORT_NAME = "重复提报与策略复杂度分析"
BASE_TABLE_FAMILY = "UK200"
REPEAT_WINDOW_DAYS = 30
MULTI_WINDOW_DAYS = 7
INPUT_PREFIX = "UK200"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

SUMMARY_COLUMNS_EN = [
    "Month",
    "Total_ASIN",
    "Unique_ASIN",
    "Repeat_ASIN",
    "Repeat_ASIN_Frequency",
    "NonRepeat_ASIN",
    "Repeat_Rate",
    "Repeat_Behavior_Rate",
    "MultiFulfillment_ASIN",
    "MultiFulfillment_Rate",
    "MultiPricing_ASIN",
    "MultiPricing_Rate",
]

SUMMARY_COLUMNS_CN = [
    "月份",
    "当月提报总次数",
    "去重ASIN数",
    "发生重复行为的ASIN数",
    "重复行为次数",
    "未重复ASIN数",
    "重复ASIN占比",
    "重复行为占比",
    "多发货方式ASIN数",
    "多发货方式占比",
    "多报价方式ASIN数",
    "多报价方式占比",
]

REPEAT_DETAIL_COLUMNS_EN = [
    "Month",
    "Child_ASIN",
    "Submission_Date",
    "Previous_Submission_Date",
    "Diff_Days",
    "QPO_ID",
    "WPO_ID",
    "QPO_Status",
    "Seller_ID",
    "Seller_Name",
    "KCP_Tag",
    "CT_Owner",
    "KP_Owner",
    "OM_Owner",
    "Fulfillment_Type",
    "Pricing_Mode",
]

REPEAT_DETAIL_COLUMNS_CN = [
    "月份",
    "子ASIN",
    "提报日期",
    "上一条提报日期",
    "间隔天数",
    "QPO编号",
    "WPO编号",
    "QPO状态",
    "卖家ID",
    "卖家名称",
    "客户KCP标签",
    "CT跟进人",
    "KP跟进人",
    "OM跟进人",
    "发货方式",
    "报价方式",
]

MULTI_DETAIL_COLUMNS_EN = [
    "Month",
    "Child_ASIN",
    "Submission_Date",
    "Previous_Submission_Date",
    "Window_Days",
    "Current_Value",
    "Previous_Value",
    "QPO_ID",
    "WPO_ID",
    "Seller_ID",
    "Seller_Name",
    "KCP_Tag",
    "CT_Owner",
    "KP_Owner",
    "OM_Owner",
]

MULTI_DETAIL_COLUMNS_CN = [
    "月份",
    "子ASIN",
    "提报日期",
    "上一条提报日期",
    "窗口天数",
    "当前值",
    "上一条值",
    "QPO编号",
    "WPO编号",
    "卖家ID",
    "卖家名称",
    "客户KCP标签",
    "CT跟进人",
    "KP跟进人",
    "OM跟进人",
]


REQUIRED_ALIASES = {
    "Submission_Date": ["Submission_Date", "Submission Date", "SubmissionDate", "提交日期", "提报日期"],
    "QPO_ID": ["QPO_ID", "QPO ID", "QPO", "QPO编号"],
    "Child_ASIN": ["Child_ASIN", "Child ASIN", "ChildASIN", "子ASIN", "子asin"],
    "Pricing_Mode": ["Pricing_Mode", "Pricing Mode", "PricingMode", "报价方式"],
    "Fulfillment_Type": ["Fulfillment_Type", "Fulfillment Type", "FulfillmentType", "发货方式", "履约方式"],
}

OPTIONAL_ALIASES = {
    "WPO_ID": ["WPO_ID", "WPO ID", "WPO", "WPO编号"],
    "QPO_Status": ["QPO_Status", "QPO Status", "状态", "QPO状态"],
    "Seller_ID": ["Seller_ID", "Seller ID", "卖家ID"],
    "Seller_Name": ["Seller_Name", "Seller Name", "SellerName", "卖家名称"],
    "KCP": ["KCP", "KCP_Tag", "KCP Tag", "客户KCP标签", "KCP标签"],
    "CT_Flup": ["CT_Flup", "CT Flup", "CT_Follow_Up", "CT Owner", "CT_Owner", "CT跟进人"],
    "KP_Flup": ["KP_Flup", "KP Flup", "KP_Follow_Up", "KP Owner", "KP_Owner", "KP跟进人"],
    "OM_Flup": ["OM_Flup", "OM Flup", "OM_Follow_Up", "OM Owner", "OM_Owner", "OM跟进人"],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据 UK200 基础表生成 UK209 重复提报与策略复杂度分析。"
    )
    parser.add_argument("--input", dest="input_path", help="指定输入的 UK200 基础表路径")
    parser.add_argument("--output", dest="output_path", help="指定输出 Excel 路径")
    return parser.parse_args()


def normalize_mode(value: object) -> str:
    return normalize_text(value)


def canonicalize_header(value: object) -> str:
    text = normalize_text(value)
    return re.sub(r"[\s_\-\/\\]+", "", text).lower()


def is_blank_row(row: tuple[object, ...] | list[object] | None) -> bool:
    if row is None:
        return True
    return all(not normalize_text(value) for value in row)


def safe_get(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index < 0 or index >= len(row):
        return None
    return row[index]


def find_index_by_aliases(header_row: list[str], aliases: list[str]) -> int | None:
    header_map = {canonicalize_header(value): idx for idx, value in enumerate(header_row)}
    for alias in aliases:
        key = canonicalize_header(alias)
        if key in header_map:
            return header_map[key]
    return None


def resolve_header_indexes(header_row: list[str]) -> tuple[dict[str, int], dict[str, int | None], list[str]]:
    required_indexes: dict[str, int] = {}
    optional_indexes: dict[str, int | None] = {}
    missing_required: list[str] = []

    try:
        exact_indexes = get_required_indexes(
            header_row,
            {
                "Submission_Date": "Submission_Date",
                "QPO_ID": "QPO_ID",
                "Child_ASIN": "Child_ASIN",
                "Pricing_Mode": "Pricing_Mode",
                "Fulfillment_Type": "Fulfillment_Type",
            },
        )
    except Exception:
        exact_indexes = {}

    for field_name, aliases in REQUIRED_ALIASES.items():
        if field_name in exact_indexes:
            required_indexes[field_name] = exact_indexes[field_name]
            continue
        idx = find_index_by_aliases(header_row, aliases)
        if idx is None:
            missing_required.append(field_name)
        else:
            required_indexes[field_name] = idx

    if missing_required:
        raise ValueError(f"缺少必要字段: {', '.join(missing_required)}")

    for field_name, aliases in OPTIONAL_ALIASES.items():
        optional_indexes[field_name] = find_index_by_aliases(header_row, aliases)

    missing_optional = [field_name for field_name, idx in optional_indexes.items() if idx is None]
    return required_indexes, optional_indexes, missing_optional


def detect_data_start_row(
    row2: tuple[object, ...],
    required_indexes: dict[str, int],
) -> tuple[int, str]:
    if not row2 or is_blank_row(row2):
        return 2, "1"

    submission_idx = required_indexes["Submission_Date"]
    qpo_idx = required_indexes["QPO_ID"]
    asin_idx = required_indexes["Child_ASIN"]

    row2_date = parse_cell_date(safe_get(row2, submission_idx))
    row2_qpo = normalize_text(safe_get(row2, qpo_idx))
    row2_asin = normalize_text(safe_get(row2, asin_idx))
    row2_submission_text = normalize_text(safe_get(row2, submission_idx))
    row2_nonempty = sum(1 for value in row2 if normalize_text(value))

    header_like_keywords = {
        "submissiondate",
        "submission_date",
        "提报日期",
        "提交日期",
        "qpoid",
        "qpo_id",
        "childasin",
        "child_asin",
        "子asin",
    }

    if row2_date is not None and row2_qpo and row2_asin:
        return 2, "1"

    if canonicalize_header(row2_submission_text) in header_like_keywords:
        return 3, "1-2"

    if row2_nonempty >= 3 and row2_date is None and (not row2_qpo or not row2_asin):
        return 3, "1-2"

    return 2, "1"


def resolve_auto_input_path() -> Path:
    candidates: dict[Path, float] = {}
    search_dirs = [EXPORT_DIR, SCRIPT_DIR, BASE_DIR]

    for directory in search_dirs:
        if not directory.exists():
            continue

        try:
            auto_path = find_latest_input(directory)
            if auto_path and auto_path.exists() and auto_path.name.upper().startswith(INPUT_PREFIX):
                candidates[auto_path] = auto_path.stat().st_mtime
        except Exception:
            pass

        for pattern in (f"{INPUT_PREFIX}*.xlsx", f"{INPUT_PREFIX}*.xlsm", f"{INPUT_PREFIX}*.xls"):
            for path in directory.glob(pattern):
                if path.is_file():
                    candidates[path] = path.stat().st_mtime

    if not candidates:
        raise FileNotFoundError(
            f"未找到 {INPUT_PREFIX} 开头的基础表文件。请使用 --input 指定输入文件。"
        )

    return max(candidates.items(), key=lambda item: item[1])[0]


def load_submission_records(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = detect_data_sheet(workbook)
        row1 = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        row2 = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), None)

        if row1 is None:
            raise ValueError("输入文件未识别到表头行。")

        header_row_1 = [normalize_text(value) for value in row1]
        if not any(header_row_1):
            raise ValueError("输入文件第 1 行表头为空。")

        required_indexes, optional_indexes, missing_optional = resolve_header_indexes(header_row_1)
        data_start_row, header_row_label = detect_data_start_row(row2 or tuple(), required_indexes)

        record_map: dict[tuple[dt.date, str, str], dict[str, object]] = {}
        metadata: dict[str, object] = {
            "Data_Sheet": sheet.title,
            "Header_Row": header_row_label,
            "Loaded_Data_Rows": 0,
            "Dedup_Removed_Rows": 0,
            "Skipped_Blank_Rows": 0,
            "Skipped_Bad_Date_Rows": 0,
            "Skipped_Missing_Core_Rows": 0,
            "Missing_Optional_Fields": ", ".join(missing_optional) if missing_optional else "(none)",
        }

        for excel_row_num, row in enumerate(
            sheet.iter_rows(min_row=data_start_row, values_only=True),
            start=data_start_row,
        ):
            metadata["Loaded_Data_Rows"] = int(metadata["Loaded_Data_Rows"]) + 1

            if is_blank_row(row):
                metadata["Skipped_Blank_Rows"] = int(metadata["Skipped_Blank_Rows"]) + 1
                continue

            submission_date = parse_cell_date(safe_get(row, required_indexes["Submission_Date"]))
            if submission_date is None:
                metadata["Skipped_Bad_Date_Rows"] = int(metadata["Skipped_Bad_Date_Rows"]) + 1
                continue

            qpo_id = normalize_text(safe_get(row, required_indexes["QPO_ID"]))
            child_asin = normalize_text(safe_get(row, required_indexes["Child_ASIN"]))
            if not qpo_id or not child_asin:
                metadata["Skipped_Missing_Core_Rows"] = int(metadata["Skipped_Missing_Core_Rows"]) + 1
                continue

            candidate = {
                "submission_date": submission_date,
                "month": submission_date.strftime("%Y-%m"),
                "qpo_id": qpo_id,
                "wpo_id": normalize_text(safe_get(row, optional_indexes["WPO_ID"])),
                "child_asin": child_asin,
                "pricing_mode": normalize_mode(safe_get(row, required_indexes["Pricing_Mode"])),
                "fulfillment_type": normalize_mode(safe_get(row, required_indexes["Fulfillment_Type"])),
                "qpo_status": normalize_text(safe_get(row, optional_indexes["QPO_Status"])),
                "seller_id": normalize_text(safe_get(row, optional_indexes["Seller_ID"])),
                "seller_name": normalize_text(safe_get(row, optional_indexes["Seller_Name"])),
                "kcp_tag": normalize_text(safe_get(row, optional_indexes["KCP"])),
                "ct_owner": normalize_text(safe_get(row, optional_indexes["CT_Flup"])),
                "kp_owner": normalize_text(safe_get(row, optional_indexes["KP_Flup"])),
                "om_owner": normalize_text(safe_get(row, optional_indexes["OM_Flup"])),
                "source_row": excel_row_num,
            }

            dedup_key = (submission_date, qpo_id, child_asin)
            existing = record_map.get(dedup_key)
            if existing is None:
                record_map[dedup_key] = candidate
                continue

            metadata["Dedup_Removed_Rows"] = int(metadata["Dedup_Removed_Rows"]) + 1
            for field_name, field_value in candidate.items():
                if field_name == "source_row":
                    existing["source_row"] = min(int(existing["source_row"]), int(field_value))
                    continue
                if not existing.get(field_name) and field_value:
                    existing[field_name] = field_value

        records = sorted(
            record_map.values(),
            key=lambda item: (
                str(item["child_asin"]),
                item["submission_date"],
                int(item["source_row"]),
                str(item["qpo_id"]),
                str(item["wpo_id"]),
            ),
        )
        return records, metadata
    finally:
        workbook.close()


def annotate_records(
    records: list[dict[str, object]],
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        grouped[str(record["child_asin"])].append(record)

    repeat_detail_rows: list[dict[str, object]] = []
    multi_fulfillment_rows: list[dict[str, object]] = []
    multi_pricing_rows: list[dict[str, object]] = []

    for child_asin, child_records in grouped.items():
        child_records.sort(
            key=lambda item: (
                item["submission_date"],
                int(item["source_row"]),
                str(item["qpo_id"]),
                str(item["wpo_id"]),
            )
        )

        recent_repeat: deque[dict[str, object]] = deque()
        recent_fulfillment: deque[dict[str, object]] = deque()
        recent_pricing: deque[dict[str, object]] = deque()

        for record in child_records:
            current_date = record["submission_date"]
            assert isinstance(current_date, dt.date)

            while recent_repeat and (current_date - recent_repeat[0]["submission_date"]).days > REPEAT_WINDOW_DAYS:
                recent_repeat.popleft()

            previous_repeat = recent_repeat[-1] if recent_repeat else None
            record["is_repeat"] = previous_repeat is not None
            record["repeat_previous_date"] = previous_repeat["submission_date"] if previous_repeat else None
            record["repeat_diff_days"] = (
                (current_date - previous_repeat["submission_date"]).days if previous_repeat else None
            )
            if record["is_repeat"]:
                repeat_detail_rows.append(record.copy())
            recent_repeat.append(record)

            while recent_fulfillment and (
                current_date - recent_fulfillment[0]["submission_date"]
            ).days > MULTI_WINDOW_DAYS:
                recent_fulfillment.popleft()

            current_fulfillment = str(record["fulfillment_type"] or "")
            previous_fulfillment_record = next(
                (
                    previous_record
                    for previous_record in reversed(recent_fulfillment)
                    if current_fulfillment
                    and str(previous_record["fulfillment_type"] or "")
                    and str(previous_record["fulfillment_type"] or "") != current_fulfillment
                ),
                None,
            )
            record["is_multi_fulfillment"] = previous_fulfillment_record is not None
            record["multi_fulfillment_previous_value"] = (
                previous_fulfillment_record["fulfillment_type"] if previous_fulfillment_record else None
            )
            record["multi_fulfillment_previous_date"] = (
                previous_fulfillment_record["submission_date"] if previous_fulfillment_record else None
            )
            if record["is_multi_fulfillment"]:
                multi_fulfillment_rows.append(record.copy())
            recent_fulfillment.append(record)

            while recent_pricing and (
                current_date - recent_pricing[0]["submission_date"]
            ).days > MULTI_WINDOW_DAYS:
                recent_pricing.popleft()

            current_pricing = str(record["pricing_mode"] or "")
            previous_pricing_record = next(
                (
                    previous_record
                    for previous_record in reversed(recent_pricing)
                    if current_pricing
                    and str(previous_record["pricing_mode"] or "")
                    and str(previous_record["pricing_mode"] or "") != current_pricing
                ),
                None,
            )
            record["is_multi_pricing"] = previous_pricing_record is not None
            record["multi_pricing_previous_value"] = (
                previous_pricing_record["pricing_mode"] if previous_pricing_record else None
            )
            record["multi_pricing_previous_date"] = (
                previous_pricing_record["submission_date"] if previous_pricing_record else None
            )
            if record["is_multi_pricing"]:
                multi_pricing_rows.append(record.copy())
            recent_pricing.append(record)

    return records, repeat_detail_rows, multi_fulfillment_rows, multi_pricing_rows


def build_summary_rows(
    records: list[dict[str, object]],
    repeat_detail_rows: list[dict[str, object]],
    multi_fulfillment_rows: list[dict[str, object]],
    multi_pricing_rows: list[dict[str, object]],
) -> list[dict[str, object]]:
    total_submission_count: defaultdict[str, int] = defaultdict(int)
    unique_asins: defaultdict[str, set[str]] = defaultdict(set)
    repeat_asins: defaultdict[str, set[str]] = defaultdict(set)
    repeat_frequency: defaultdict[str, int] = defaultdict(int)
    multi_fulfillment_asins: defaultdict[str, set[str]] = defaultdict(set)
    multi_pricing_asins: defaultdict[str, set[str]] = defaultdict(set)

    for record in records:
        month = str(record["month"])
        child_asin = str(record["child_asin"])
        total_submission_count[month] += 1
        unique_asins[month].add(child_asin)

    for record in repeat_detail_rows:
        month = str(record["month"])
        child_asin = str(record["child_asin"])
        repeat_asins[month].add(child_asin)
        repeat_frequency[month] += 1

    for record in multi_fulfillment_rows:
        multi_fulfillment_asins[str(record["month"])].add(str(record["child_asin"]))

    for record in multi_pricing_rows:
        multi_pricing_asins[str(record["month"])].add(str(record["child_asin"]))

    months = sorted(unique_asins.keys(), key=month_sort_key)
    rows: list[dict[str, object]] = []
    for month in months:
        unique_count = len(unique_asins[month])
        repeat_count = len(repeat_asins[month])
        total_count = total_submission_count[month]
        multi_fulfillment_count = len(multi_fulfillment_asins[month])
        multi_pricing_count = len(multi_pricing_asins[month])

        rows.append(
            {
                "Month": month,
                "Total_ASIN": total_count,
                "Unique_ASIN": unique_count,
                "Repeat_ASIN": repeat_count,
                "Repeat_ASIN_Frequency": repeat_frequency[month],
                "NonRepeat_ASIN": unique_count - repeat_count,
                "Repeat_Rate": safe_div(repeat_count, unique_count),
                "Repeat_Behavior_Rate": safe_div(repeat_frequency[month], total_count),
                "MultiFulfillment_ASIN": multi_fulfillment_count,
                "MultiFulfillment_Rate": safe_div(multi_fulfillment_count, unique_count),
                "MultiPricing_ASIN": multi_pricing_count,
                "MultiPricing_Rate": safe_div(multi_pricing_count, unique_count),
            }
        )
    return rows


def apply_sheet_body_format(
    sheet,
    columns_en: list[str],
    *,
    percent_columns: set[str] | None = None,
    date_columns: set[str] | None = None,
    integer_columns: set[str] | None = None,
    left_align_columns: set[str] | None = None,
) -> None:
    percent_columns = percent_columns or set()
    date_columns = date_columns or set()
    integer_columns = integer_columns or set()
    left_align_columns = left_align_columns or set()
    index_map = {column: idx + 1 for idx, column in enumerate(columns_en)}

    for row_num in range(3, sheet.max_row + 1):
        for column in columns_en:
            cell = sheet.cell(row=row_num, column=index_map[column])
            cell.alignment = LEFT if column in left_align_columns else CENTER
            if column in percent_columns and cell.value is not None:
                cell.number_format = "0.00%"
            elif column in integer_columns and cell.value is not None:
                cell.number_format = "0"
            elif column in date_columns and cell.value is not None:
                cell.number_format = "yyyy-mm-dd"


def append_sheet(
    workbook: Workbook,
    title: str,
    columns_en: list[str],
    columns_cn: list[str],
    rows: list[dict[str, object]],
    *,
    percent_columns: set[str] | None = None,
    date_columns: set[str] | None = None,
    integer_columns: set[str] | None = None,
    left_align_columns: set[str] | None = None,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns_en)
    sheet.append(columns_cn)
    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    style_two_header_sheet(sheet, columns_en, columns_cn, date_columns=date_columns or set())
    apply_sheet_body_format(
        sheet,
        columns_en,
        percent_columns=percent_columns,
        date_columns=date_columns,
        integer_columns=integer_columns,
        left_align_columns=left_align_columns,
    )


def build_notes_rows(runtime_meta: dict[str, object]) -> tuple[list[tuple[object, ...]], list[tuple[object, ...]]]:
    field_rows = [
        ("Field_EN", "字段中文", "计算逻辑", "示例/说明"),
        ("Month", "月份", "Submission_Date 提取 yyyy-mm", "按触发行为所在月份统计"),
        ("Total_ASIN", "当月提报总次数", "按 Submission_Date + QPO_ID + Child_ASIN 去重后的行为条数", "不做 Child_ASIN 再去重，代表行为量"),
        ("Unique_ASIN", "去重ASIN数", "当月 distinct Child_ASIN", "代表当月被提报的独立子 ASIN 数"),
        ("Repeat_ASIN", "发生重复行为的ASIN数", "当月内至少有 1 条记录满足“往前 30 天出现过同 Child_ASIN”", "同日重复提报按后出现记录计入重复"),
        ("Repeat_ASIN_Frequency", "重复行为次数", "当月满足前 30 天重复条件的记录条数", "同一 Child_ASIN 在同月可贡献多次"),
        ("NonRepeat_ASIN", "未重复ASIN数", "Unique_ASIN - Repeat_ASIN", "当月未触发重复行为的 Child_ASIN"),
        ("Repeat_Rate", "重复ASIN占比", "Repeat_ASIN / Unique_ASIN", "零分母时按 0 处理"),
        ("Repeat_Behavior_Rate", "重复行为占比", "Repeat_ASIN_Frequency / Total_ASIN", "衡量行为层面的重复强度"),
        ("MultiFulfillment_ASIN", "多发货方式ASIN数", "同一 Child_ASIN 在 7 天窗口内出现至少 2 种 Fulfillment_Type", "按触发月份统计 distinct Child_ASIN"),
        ("MultiFulfillment_Rate", "多发货方式占比", "MultiFulfillment_ASIN / Unique_ASIN", "零分母时按 0 处理"),
        ("MultiPricing_ASIN", "多报价方式ASIN数", "同一 Child_ASIN 在 7 天窗口内出现至少 2 种 Pricing_Mode", "按触发月份统计 distinct Child_ASIN"),
        ("MultiPricing_Rate", "多报价方式占比", "MultiPricing_ASIN / Unique_ASIN", "零分母时按 0 处理"),
        ("Repeat_Detail", "重复明细", "仅保留满足前 30 天重复条件的记录", "输出上一条提报日期、间隔天数及复核字段"),
        ("MultiFulfillment_Detail", "多发货方式明细", "仅保留触发 7 天多发货方式的记录", "输出当前值与上一条不同值"),
        ("MultiPricing_Detail", "多报价方式明细", "仅保留触发 7 天多报价方式的记录", "输出当前值与上一条不同值"),
        ("示例", "样例说明", "若 A 在 2/28、3/5、3/6 提报，则 3 月 Repeat_ASIN=1、Repeat_ASIN_Frequency=2", "2/28 为首次提报，不算重复；3/5、3/6 均算重复行为"),
    ]

    runtime_rows = [
        ("Item", "Value"),
        ("Report_Code", runtime_meta.get("Report_Code")),
        ("Report_Name", runtime_meta.get("Report_Name")),
        ("Input_File", runtime_meta.get("Input_File")),
        ("Output_File", runtime_meta.get("Output_File")),
        ("Data_Sheet", runtime_meta.get("Data_Sheet")),
        ("Header_Row", runtime_meta.get("Header_Row")),
        ("Generated_At", runtime_meta.get("Generated_At")),
        ("Dedup_Key", runtime_meta.get("Dedup_Key")),
        ("Repeat_Window", runtime_meta.get("Repeat_Window")),
        ("Multi_Window", runtime_meta.get("Multi_Window")),
        ("Summary_Grain", runtime_meta.get("Summary_Grain")),
        ("Detail_Grain", runtime_meta.get("Detail_Grain")),
        ("Loaded_Data_Rows", runtime_meta.get("Loaded_Data_Rows")),
        ("Dedup_Removed_Rows", runtime_meta.get("Dedup_Removed_Rows")),
        ("Skipped_Blank_Rows", runtime_meta.get("Skipped_Blank_Rows")),
        ("Skipped_Bad_Date_Rows", runtime_meta.get("Skipped_Bad_Date_Rows")),
        ("Skipped_Missing_Core_Rows", runtime_meta.get("Skipped_Missing_Core_Rows")),
        ("Missing_Optional_Fields", runtime_meta.get("Missing_Optional_Fields")),
        ("Summary_Row_Count", runtime_meta.get("Summary_Row_Count")),
        ("Repeat_Detail_Count", runtime_meta.get("Repeat_Detail_Count")),
        ("MultiFulfillment_Detail_Count", runtime_meta.get("MultiFulfillment_Detail_Count")),
        ("MultiPricing_Detail_Count", runtime_meta.get("MultiPricing_Detail_Count")),
        ("Base_Table_Family", runtime_meta.get("Base_Table_Family")),
    ]
    return field_rows, runtime_rows


def write_notes_sheet(workbook: Workbook, runtime_meta: dict[str, object]) -> None:
    sheet = workbook.active
    sheet.title = "Notes"

    field_rows, runtime_rows = build_notes_rows(runtime_meta)

    sheet.merge_cells("A1:D1")
    sheet["A1"] = "字段口径说明"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = HEADER_FONT
    sheet["A1"].alignment = CENTER

    for row in field_rows:
        sheet.append(list(row))

    field_header_row = 2
    for cell in sheet[field_header_row]:
        cell.fill = SUBHEADER_FILL
        cell.font = BOLD_FONT
        cell.alignment = CENTER

    sheet.append([])
    runtime_title_row = sheet.max_row + 1
    sheet.merge_cells(start_row=runtime_title_row, start_column=1, end_row=runtime_title_row, end_column=4)
    title_cell = sheet.cell(row=runtime_title_row, column=1)
    title_cell.value = "运行元信息"
    title_cell.fill = HEADER_FILL
    title_cell.font = HEADER_FONT
    title_cell.alignment = CENTER

    for item, value in runtime_rows:
        sheet.append([item, value, None, None])

    runtime_header_row = runtime_title_row + 1
    for cell in sheet[runtime_header_row]:
        cell.fill = SUBHEADER_FILL
        cell.font = BOLD_FONT
        cell.alignment = CENTER

    for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
        for cell in row:
            if cell.row in (runtime_title_row, runtime_header_row):
                continue
            cell.alignment = LEFT

    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 26
    sheet.column_dimensions["C"].width = 58
    sheet.column_dimensions["D"].width = 62
    sheet.freeze_panes = "A2"


def build_repeat_detail_rows(repeat_detail_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for record in repeat_detail_rows:
        rows.append(
            {
                "Month": record["month"],
                "Child_ASIN": record["child_asin"],
                "Submission_Date": record["submission_date"],
                "Previous_Submission_Date": record["repeat_previous_date"],
                "Diff_Days": record["repeat_diff_days"],
                "QPO_ID": record["qpo_id"],
                "WPO_ID": record["wpo_id"],
                "QPO_Status": record["qpo_status"],
                "Seller_ID": record["seller_id"],
                "Seller_Name": record["seller_name"],
                "KCP_Tag": record["kcp_tag"],
                "CT_Owner": record["ct_owner"],
                "KP_Owner": record["kp_owner"],
                "OM_Owner": record["om_owner"],
                "Fulfillment_Type": record["fulfillment_type"],
                "Pricing_Mode": record["pricing_mode"],
            }
        )

    rows.sort(
        key=lambda item: (
            month_sort_key(str(item["Month"])),
            str(item["Child_ASIN"]),
            item["Submission_Date"] or dt.date.min,
            str(item["QPO_ID"]),
            str(item["WPO_ID"]),
        )
    )
    return rows


def build_multi_detail_rows(
    records: list[dict[str, object]],
    *,
    previous_date_field: str,
    previous_value_field: str,
    current_value_field: str,
) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for record in records:
        previous_date = record.get(previous_date_field)
        submission_date = record.get("submission_date")
        window_days = (
            (submission_date - previous_date).days
            if isinstance(previous_date, dt.date) and isinstance(submission_date, dt.date)
            else None
        )
        rows.append(
            {
                "Month": record["month"],
                "Child_ASIN": record["child_asin"],
                "Submission_Date": submission_date,
                "Previous_Submission_Date": previous_date,
                "Window_Days": window_days,
                "Current_Value": record.get(current_value_field),
                "Previous_Value": record.get(previous_value_field),
                "QPO_ID": record["qpo_id"],
                "WPO_ID": record["wpo_id"],
                "Seller_ID": record["seller_id"],
                "Seller_Name": record["seller_name"],
                "KCP_Tag": record["kcp_tag"],
                "CT_Owner": record["ct_owner"],
                "KP_Owner": record["kp_owner"],
                "OM_Owner": record["om_owner"],
            }
        )

    rows.sort(
        key=lambda item: (
            month_sort_key(str(item["Month"])),
            str(item["Child_ASIN"]),
            item["Submission_Date"] or dt.date.min,
            str(item["QPO_ID"]),
            str(item["WPO_ID"]),
        )
    )
    return rows


def write_workbook(
    output_path: Path,
    summary_rows: list[dict[str, object]],
    repeat_rows: list[dict[str, object]],
    multi_fulfillment_rows: list[dict[str, object]],
    multi_pricing_rows: list[dict[str, object]],
    runtime_meta: dict[str, object],
) -> None:
    workbook = Workbook()

    write_notes_sheet(workbook, runtime_meta)

    append_sheet(
        workbook,
        "Summary",
        SUMMARY_COLUMNS_EN,
        SUMMARY_COLUMNS_CN,
        summary_rows,
        percent_columns={
            "Repeat_Rate",
            "Repeat_Behavior_Rate",
            "MultiFulfillment_Rate",
            "MultiPricing_Rate",
        },
        integer_columns={
            "Total_ASIN",
            "Unique_ASIN",
            "Repeat_ASIN",
            "Repeat_ASIN_Frequency",
            "NonRepeat_ASIN",
            "MultiFulfillment_ASIN",
            "MultiPricing_ASIN",
        },
    )

    append_sheet(
        workbook,
        "Repeat_Detail",
        REPEAT_DETAIL_COLUMNS_EN,
        REPEAT_DETAIL_COLUMNS_CN,
        repeat_rows,
        date_columns={"Submission_Date", "Previous_Submission_Date"},
        integer_columns={"Diff_Days"},
        left_align_columns={
            "Month",
            "Child_ASIN",
            "QPO_ID",
            "WPO_ID",
            "QPO_Status",
            "Seller_ID",
            "Seller_Name",
            "KCP_Tag",
            "CT_Owner",
            "KP_Owner",
            "OM_Owner",
            "Fulfillment_Type",
            "Pricing_Mode",
        },
    )

    append_sheet(
        workbook,
        "MultiFulfillment_Detail",
        MULTI_DETAIL_COLUMNS_EN,
        MULTI_DETAIL_COLUMNS_CN,
        multi_fulfillment_rows,
        date_columns={"Submission_Date", "Previous_Submission_Date"},
        integer_columns={"Window_Days"},
        left_align_columns={
            "Month",
            "Child_ASIN",
            "Current_Value",
            "Previous_Value",
            "QPO_ID",
            "WPO_ID",
            "Seller_ID",
            "Seller_Name",
            "KCP_Tag",
            "CT_Owner",
            "KP_Owner",
            "OM_Owner",
        },
    )

    append_sheet(
        workbook,
        "MultiPricing_Detail",
        MULTI_DETAIL_COLUMNS_EN,
        MULTI_DETAIL_COLUMNS_CN,
        multi_pricing_rows,
        date_columns={"Submission_Date", "Previous_Submission_Date"},
        integer_columns={"Window_Days"},
        left_align_columns={
            "Month",
            "Child_ASIN",
            "Current_Value",
            "Previous_Value",
            "QPO_ID",
            "WPO_ID",
            "Seller_ID",
            "Seller_Name",
            "KCP_Tag",
            "CT_Owner",
            "KP_Owner",
            "OM_Owner",
        },
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()

    input_path = Path(args.input_path) if args.input_path else resolve_auto_input_path()
    if not input_path.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )

    records, load_meta = load_submission_records(input_path)
    records, repeat_detail_records, multi_fulfillment_detail_records, multi_pricing_detail_records = annotate_records(records)

    summary_rows = build_summary_rows(
        records,
        repeat_detail_records,
        multi_fulfillment_detail_records,
        multi_pricing_detail_records,
    )
    repeat_rows = build_repeat_detail_rows(repeat_detail_records)
    multi_fulfillment_rows = build_multi_detail_rows(
        multi_fulfillment_detail_records,
        previous_date_field="multi_fulfillment_previous_date",
        previous_value_field="multi_fulfillment_previous_value",
        current_value_field="fulfillment_type",
    )
    multi_pricing_rows = build_multi_detail_rows(
        multi_pricing_detail_records,
        previous_date_field="multi_pricing_previous_date",
        previous_value_field="multi_pricing_previous_value",
        current_value_field="pricing_mode",
    )

    runtime_meta: dict[str, object] = {
        "Report_Code": REPORT_CODE,
        "Report_Name": REPORT_NAME,
        "Input_File": str(input_path),
        "Output_File": str(output_path),
        "Data_Sheet": load_meta.get("Data_Sheet", ""),
        "Header_Row": load_meta.get("Header_Row", ""),
        "Generated_At": dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Dedup_Key": "Submission_Date + QPO_ID + Child_ASIN",
        "Repeat_Window": f"{REPEAT_WINDOW_DAYS} days",
        "Multi_Window": f"{MULTI_WINDOW_DAYS} days",
        "Summary_Grain": "Month",
        "Detail_Grain": "Child_ASIN 对应行为记录",
        "Loaded_Data_Rows": load_meta.get("Loaded_Data_Rows", 0),
        "Dedup_Removed_Rows": load_meta.get("Dedup_Removed_Rows", 0),
        "Skipped_Blank_Rows": load_meta.get("Skipped_Blank_Rows", 0),
        "Skipped_Bad_Date_Rows": load_meta.get("Skipped_Bad_Date_Rows", 0),
        "Skipped_Missing_Core_Rows": load_meta.get("Skipped_Missing_Core_Rows", 0),
        "Missing_Optional_Fields": load_meta.get("Missing_Optional_Fields", "(none)"),
        "Summary_Row_Count": len(summary_rows),
        "Repeat_Detail_Count": len(repeat_rows),
        "MultiFulfillment_Detail_Count": len(multi_fulfillment_rows),
        "MultiPricing_Detail_Count": len(multi_pricing_rows),
        "Base_Table_Family": BASE_TABLE_FAMILY,
    }

    write_workbook(
        output_path,
        summary_rows,
        repeat_rows,
        multi_fulfillment_rows,
        multi_pricing_rows,
        runtime_meta,
    )

    print("结果已生成：")
    print(output_path)


if __name__ == "__main__":
    main()
