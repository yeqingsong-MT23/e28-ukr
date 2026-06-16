# -*- coding: utf-8 -*-
"""
根据 UK200_转化漏斗KPI 基础表生成 UK208_P5价格区间转化漏斗概览.xlsx。
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from uk_export_utils import build_default_output_path
from uk_funnel_utils import (
    detect_data_sheet,
    find_latest_input,
    get_required_indexes,
    has_value,
    map_simple_type as util_map_simple_type,
    map_split_type as util_map_split_type,
    normalize_text,
    parse_cell_date,
    safe_div,
    style_two_header_sheet,
)


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
BASE_TABLE_PREFIX = "UK200"
DEFAULT_OUTPUT_NAME = "UK208_P5价格区间转化漏斗概览.xlsx"

BUCKETS = [
    "$0-$20",
    "$20-$35",
    "$35-$50",
    "$50-$80",
    "$80-$120",
    "$120-$200",
    "$200-$350",
    "$350+",
]

BUCKET_BOUNDS = [
    (0, 20, "$0-$20"),
    (20, 35, "$20-$35"),
    (35, 50, "$35-$50"),
    (50, 80, "$50-$80"),
    (80, 120, "$80-$120"),
    (120, 200, "$120-$200"),
    (200, 350, "$200-$350"),
]

STAGE_LABELS = {
    "Initial": "一审",
    "Second": "二审",
    "Third": "三审",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据 UK200 基础表生成 UK208_P5价格区间转化漏斗概览。"
    )
    parser.add_argument("--input", dest="input_path", help="指定输入的 UK200 基础表路径。")
    parser.add_argument("--output", dest="output_path", help="指定输出 Excel 路径。")
    return parser.parse_args()


def has_chinese(value: object) -> bool:
    text = str(value or "")
    return any("\u4e00" <= ch <= "\u9fff" for ch in text)


def to_number(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)

    text = str(value).strip()
    if not text:
        return None

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()

    text = (
        text.replace(",", "")
        .replace("$", "")
        .replace("£", "")
        .replace("¥", "")
        .replace("€", "")
        .replace("￥", "")
        .replace(" ", "")
    )
    text = re.sub(r"[^0-9.\-]", "", text)
    if text in {"", "-", ".", "-."}:
        return None

    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    return -number if negative else number


def bucket_p5(value: float | None) -> str | None:
    if value is None or value < 0:
        return None
    for lower, upper, name in BUCKET_BOUNDS:
        if lower <= value < upper:
            return name
    return "$350+"


def _canonical_label(value: object) -> str:
    return (
        normalize_text(value)
        .replace(" ", "")
        .replace("_", "")
        .replace("-", "")
        .lower()
    )


def _normalize_simple_type(status: str) -> str | None:
    mapped = util_map_simple_type(normalize_text(status))
    key = _canonical_label(mapped)
    if key == "success":
        return "Success"
    if key == "cancel":
        return "Cancel"
    if key == "inprogress":
        return "InProgress"
    return None


def map_split_type(status: str) -> str:
    mapped = util_map_split_type(normalize_text(status))
    key = _canonical_label(mapped)
    if key == "wootcancel":
        return "WootCancel"
    if key == "sellercancel":
        return "SellerCancel"
    if key == "success":
        return "Success"
    if key == "inprogress":
        return "InProgress"

    fallback = _normalize_simple_type(status)
    if fallback == "Cancel":
        return "Cancel"
    if fallback in {"Success", "InProgress"}:
        return fallback
    return "Other"


def _find_latest_input_by_prefix(base_dir: Path, prefix: str) -> Path | None:
    candidates: list[Path] = []
    patterns = [f"{prefix}*.xlsx", f"{prefix}*.xlsm", f"{prefix}*.xls"]
    for pattern in patterns:
        candidates.extend(path for path in base_dir.glob(pattern) if path.is_file())
    if not candidates:
        for pattern in patterns:
            candidates.extend(path for path in base_dir.rglob(pattern) if path.is_file())
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def resolve_input_path(input_arg: str | None) -> Path:
    if input_arg:
        input_path = Path(input_arg)
        if not input_path.exists():
            raise FileNotFoundError(f"输入文件不存在: {input_path}")
        return input_path

    try:
        latest_path = find_latest_input(BASE_DIR)
        if latest_path and latest_path.exists() and latest_path.name.upper().startswith(BASE_TABLE_PREFIX):
            return latest_path
    except Exception:
        latest_path = None

    scanned = _find_latest_input_by_prefix(BASE_DIR, BASE_TABLE_PREFIX)
    if scanned is not None:
        return scanned

    scanned = _find_latest_input_by_prefix(SCRIPT_DIR, BASE_TABLE_PREFIX)
    if scanned is not None:
        return scanned

    raise FileNotFoundError(
        f"未找到以 {BASE_TABLE_PREFIX} 开头的基础表文件，请使用 --input 指定输入文件。"
    )


def _mean_or_none(values: list[float | None]) -> float | None:
    valid = [value for value in values if value is not None]
    if not valid:
        return None
    return mean(valid)


def _mean_or_zero(values: list[float | None]) -> float:
    valid = [value for value in values if value is not None]
    if not valid:
        return 0.0
    return mean(valid)


def build_clean_records(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = detect_data_sheet(workbook)

        header_rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
        header_row_1 = [
            normalize_text(value) for value in (header_rows[0] if len(header_rows) >= 1 else [])
        ]
        header_row_2 = [
            normalize_text(value) for value in (header_rows[1] if len(header_rows) >= 2 else [])
        ]
        data_start_row = 3 if sum(has_chinese(value) for value in header_row_2) > 0 else 2

        indexes = get_required_indexes(
            header_row_1,
            {
                "Submission_Date": "Submission_Date",
                "QPO_ID": "QPO_ID",
                "QPO_Status": "QPO_Status",
                "Child_Status": "Child_Status",
                "Initial_Quote": "Initial_Quote",
                "Second_Quote": "Second_Quote",
                "MVM_Approved_Price": "MVM_Approved_Price",
                "P5": "P5",
                "P6": "P6",
                "M1": "M1",
            },
        )

        raw_rows = 0
        valid_rows = 0
        valid_qpo_rows = 0
        qpo_rows: dict[str, list[dict[str, object]]] = defaultdict(list)

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            raw_rows += 1
            submission_date = parse_cell_date(row[indexes["Submission_Date"]])
            if submission_date is None:
                continue
            valid_rows += 1

            qpo_id = normalize_text(row[indexes["QPO_ID"]])
            if not qpo_id:
                continue
            valid_qpo_rows += 1

            status = normalize_text(row[indexes["QPO_Status"]])
            child_status = normalize_text(row[indexes["Child_Status"]])

            qpo_rows[qpo_id].append(
                {
                    "submission_date": submission_date,
                    "status": status,
                    "child_status": child_status,
                    "is_child_cancel": child_status.lower() == "cancel",
                    "has_initial": has_value(row[indexes["Initial_Quote"]]),
                    "has_second": has_value(row[indexes["Second_Quote"]]),
                    "has_third": has_value(row[indexes["MVM_Approved_Price"]]),
                    "p5": to_number(row[indexes["P5"]]),
                    "p6": to_number(row[indexes["P6"]]),
                    "m1": to_number(row[indexes["M1"]]),
                    "split_type": map_split_type(status),
                }
            )

        deleted_type_a_rows = 0
        deleted_type_b_qpo_ids: list[str] = []
        records: list[dict[str, object]] = []

        for qpo_id, rows in qpo_rows.items():
            qpo_has_any_quote = any(
                row["has_initial"] or row["has_second"] or row["has_third"] for row in rows
            )

            kept_rows: list[dict[str, object]] = []
            for row in rows:
                should_delete_type_a = (
                    (not qpo_has_any_quote)
                    and row["split_type"] == "Success"
                    and not row["is_child_cancel"]
                    and not (row["has_initial"] or row["has_second"] or row["has_third"])
                )
                if should_delete_type_a:
                    deleted_type_a_rows += 1
                    continue
                kept_rows.append(row)

            if not kept_rows:
                continue

            if any(row["split_type"] == "Success" for row in kept_rows) and all(
                row["is_child_cancel"] for row in kept_rows
            ):
                deleted_type_b_qpo_ids.append(qpo_id)
                continue

            earliest_date = min(row["submission_date"] for row in kept_rows)
            has_initial = any(bool(row["has_initial"]) for row in kept_rows)
            has_second = any(bool(row["has_second"]) for row in kept_rows)
            has_third = any(bool(row["has_third"]) for row in kept_rows)

            stage = None
            if has_third:
                stage = "Third"
            elif has_second:
                stage = "Second"
            elif has_initial:
                stage = "Initial"

            split_types = {str(row["split_type"]) for row in kept_rows}
            if "WootCancel" in split_types:
                split_type = "WootCancel"
            elif "SellerCancel" in split_types:
                split_type = "SellerCancel"
            elif "Cancel" in split_types:
                split_type = "Cancel"
            elif "Success" in split_types:
                split_type = "Success"
            elif "InProgress" in split_types:
                split_type = "InProgress"
            else:
                split_type = "Other"

            simple_type = "Cancel" if split_type in {"WootCancel", "SellerCancel", "Cancel"} else split_type

            price_rows = [row for row in kept_rows if not row["is_child_cancel"]] or kept_rows
            avg_p5 = _mean_or_none([row["p5"] for row in price_rows])
            avg_p6 = _mean_or_none([row["p6"] for row in price_rows])
            avg_m1 = _mean_or_none([row["m1"] for row in price_rows])

            records.append(
                {
                    "qpo_id": qpo_id,
                    "submission_date": earliest_date,
                    "split_type": split_type,
                    "simple_type": simple_type,
                    "stage": stage,
                    "avg_p5": avg_p5,
                    "avg_p6": avg_p6,
                    "avg_m1": avg_m1,
                    "bucket": bucket_p5(avg_p5),
                }
            )

        metadata = {
            "input_path": str(input_path),
            "sheet_name": sheet.title,
            "raw_rows": raw_rows,
            "valid_rows": valid_rows,
            "valid_qpo_rows": valid_qpo_rows,
            "collected_qpo_count": len(qpo_rows),
            "deleted_type_a_rows": deleted_type_a_rows,
            "deleted_type_b_qpo_ids": deleted_type_b_qpo_ids,
            "deleted_type_b_qpo_count": len(deleted_type_b_qpo_ids),
            "post_clean_qpo_count": len(records),
        }
        return records, metadata
    finally:
        workbook.close()


def aggregate(records: list[dict[str, object]]) -> tuple[list[dict[str, object]], dict[str, object]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for record in records:
        bucket = record.get("bucket")
        if bucket is not None:
            grouped[str(bucket)].append(record)

    rows: list[dict[str, object]] = []
    for bucket in BUCKETS:
        bucket_records = grouped.get(bucket, [])
        submitted = len(bucket_records)

        success_records = [record for record in bucket_records if record["simple_type"] == "Success"]
        cancel_records = [record for record in bucket_records if record["simple_type"] == "Cancel"]
        inprogress_records = [
            record for record in bucket_records if record["simple_type"] == "InProgress"
        ]
        woot_records = [record for record in bucket_records if record["split_type"] == "WootCancel"]
        seller_records = [record for record in bucket_records if record["split_type"] == "SellerCancel"]

        row = {
            "P5_Bucket": bucket,
            "Submitted_QPOs": submitted,
            "Success_QPOs": len(success_records),
            "Success_Rate": safe_div(len(success_records), submitted),
            "Cancel_QPOs": len(cancel_records),
            "Cancel_Rate": safe_div(len(cancel_records), submitted),
            "InProgress_QPOs": len(inprogress_records),
            "InProgress_Rate": safe_div(len(inprogress_records), submitted),
            "WootCancel_QPOs": len(woot_records),
            "WootCancel_Share_of_Cancel": safe_div(len(woot_records), len(cancel_records)),
            "SellerCancel_QPOs": len(seller_records),
            "SellerCancel_Share_of_Cancel": safe_div(len(seller_records), len(cancel_records)),
            "Avg_P5_Success": _mean_or_zero([record["avg_p5"] for record in success_records]),
            "Avg_P6_Success": _mean_or_zero([record["avg_p6"] for record in success_records]),
            "Avg_M1_Success": _mean_or_zero([record["avg_m1"] for record in success_records]),
        }
        row["M1_over_P6"] = safe_div(row["Avg_M1_Success"], row["Avg_P6_Success"])

        for stage_name in ("Initial", "Second", "Third"):
            stage_records = [record for record in bucket_records if record["stage"] == stage_name]
            stage_success = [record for record in stage_records if record["simple_type"] == "Success"]
            stage_cancel = [record for record in stage_records if record["simple_type"] == "Cancel"]
            stage_inprogress = [
                record for record in stage_records if record["simple_type"] == "InProgress"
            ]
            stage_woot = [record for record in stage_records if record["split_type"] == "WootCancel"]
            stage_seller = [
                record for record in stage_records if record["split_type"] == "SellerCancel"
            ]

            row[f"{stage_name}_Quoted_QPOs"] = len(stage_records)
            row[f"{stage_name}_Success_QPOs"] = len(stage_success)
            row[f"{stage_name}_Success_Rate"] = safe_div(len(stage_success), len(stage_records))
            row[f"{stage_name}_Cancel_QPOs"] = len(stage_cancel)
            row[f"{stage_name}_Cancel_Rate"] = safe_div(len(stage_cancel), len(stage_records))
            row[f"{stage_name}_InProgress_QPOs"] = len(stage_inprogress)
            row[f"{stage_name}_InProgress_Rate"] = safe_div(len(stage_inprogress), len(stage_records))
            row[f"{stage_name}_WootCancel_QPOs"] = len(stage_woot)
            row[f"{stage_name}_WootCancel_Share_of_StageCancel"] = safe_div(
                len(stage_woot), len(stage_cancel)
            )
            row[f"{stage_name}_SellerCancel_QPOs"] = len(stage_seller)
            row[f"{stage_name}_SellerCancel_Share_of_StageCancel"] = safe_div(
                len(stage_seller), len(stage_cancel)
            )

        rows.append(row)

    audit = {
        "bucketed_submitted_sum": sum(int(row["Submitted_QPOs"]) for row in rows),
        "unbucketed_qpos": sum(1 for record in records if record.get("bucket") is None),
    }
    return rows, audit


def build_notes_rows(metadata: dict[str, object], audit: dict[str, object]) -> list[tuple[object, ...]]:
    rows: list[tuple[object, ...]] = [
        ("Field_English", "Field_Chinese", "Field_Group", "Definition", "Denominator/Notes"),
        (
            "P5_Bucket",
            "P5区间",
            "Bucket",
            "P5 bucket based on QPO-level Avg_P5 after anomaly cleaning.",
            "QPO先归并，再按 Avg_P5 分桶；左闭右开：[0,20)/[20,35)/[35,50)/[50,80)/[80,120)/[120,200)/[200,350)/[350,∞)。Avg_P5 缺失或 <0 不进桶。",
        ),
        (
            "Submitted_QPOs",
            "提报QPO数",
            "Overall",
            "Distinct QPO_ID count included in the bucket.",
            "分母基准；按清洗后、已分桶的唯一 QPO_ID 计数。",
        ),
        (
            "Success_QPOs",
            "成交QPO数",
            "Overall",
            "Bucket-level distinct QPO_ID count with final unified status = Success.",
            "按清洗后唯一 QPO_ID 计数。",
        ),
        (
            "Success_Rate",
            "成交占比",
            "Overall",
            "Success_QPOs / Submitted_QPOs.",
            "分母 = Submitted_QPOs。",
        ),
        (
            "Cancel_QPOs",
            "取消QPO数",
            "Overall",
            "Bucket-level distinct QPO_ID count with final unified status = Cancel.",
            "按清洗后唯一 QPO_ID 计数。",
        ),
        (
            "Cancel_Rate",
            "取消占比",
            "Overall",
            "Cancel_QPOs / Submitted_QPOs.",
            "分母 = Submitted_QPOs。",
        ),
        (
            "InProgress_QPOs",
            "跟进中QPO数",
            "Overall",
            "Bucket-level distinct QPO_ID count with final unified status = InProgress.",
            "按清洗后唯一 QPO_ID 计数。",
        ),
        (
            "InProgress_Rate",
            "跟进中占比",
            "Overall",
            "InProgress_QPOs / Submitted_QPOs.",
            "分母 = Submitted_QPOs。",
        ),
        (
            "WootCancel_QPOs",
            "WootCancel数",
            "Overall",
            "Canceled QPOs classified as WootCancel.",
            "总体取消拆分口径；分子 = WootCancel_QPOs，分母见下一行。",
        ),
        (
            "WootCancel_Share_of_Cancel",
            "WootCancel占取消比",
            "Overall",
            "WootCancel_QPOs / Cancel_QPOs.",
            "分母 = 总体 Cancel_QPOs。",
        ),
        (
            "SellerCancel_QPOs",
            "SellerCancel数",
            "Overall",
            "Canceled QPOs classified as SellerCancel.",
            "总体取消拆分口径；分子 = SellerCancel_QPOs，分母见下一行。",
        ),
        (
            "SellerCancel_Share_of_Cancel",
            "SellerCancel占取消比",
            "Overall",
            "SellerCancel_QPOs / Cancel_QPOs.",
            "分母 = 总体 Cancel_QPOs。",
        ),
        (
            "Avg_P5_Success",
            "成交平均P5",
            "Price",
            "Mean of QPO-level Avg_P5 among Success QPOs within the bucket.",
            "QPO级均值再取桶内 Success 均值，不是逐行平均。",
        ),
        (
            "Avg_P6_Success",
            "成交平均P6",
            "Price",
            "Mean of QPO-level Avg_P6 among Success QPOs within the bucket.",
            "QPO级均值再取桶内 Success 均值。",
        ),
        (
            "Avg_M1_Success",
            "成交平均M1",
            "Price",
            "Mean of QPO-level Avg_M1 among Success QPOs within the bucket.",
            "QPO级均值再取桶内 Success 均值。",
        ),
        (
            "M1_over_P6",
            "采购价占比(M1/P6)",
            "Price",
            "Avg_M1_Success / Avg_P6_Success.",
            "M1_over_P6 = Avg_M1_Success / Avg_P6_Success。",
        ),
        (
            "Price Cleaning Rule",
            "价格清洗规则",
            "RULE",
            "Price fields are converted to numeric after removing currency symbols, commas and blanks.",
            "P5/P6/M1 仅保留可转数字值；QPO层均值优先取非 Cancel child 行，如全部为 Cancel 则回退到保留行。",
        ),
        (
            "Anomaly Cleaning Rule - Type A",
            "异常删除规则A",
            "RULE",
            "Delete success child rows when the QPO has no quote-stage flag at all and the child row itself has no quote.",
            "A类删除的是子行，不是整单。",
        ),
        (
            "Anomaly Cleaning Rule - Type B",
            "异常删除规则B",
            "RULE",
            "Exclude the whole QPO when kept rows still contain Success but all remaining child rows are Cancel.",
            "B类删除的是整单 QPO_ID。",
        ),
        (
            "Bucket Boundaries",
            "价格桶固定边界",
            "RULE",
            "Fixed P5 bucket boundaries.",
            "$0-$20:[0,20); $20-$35:[20,35); $35-$50:[35,50); $50-$80:[50,80); $80-$120:[80,120); $120-$200:[120,200); $200-$350:[200,350); $350+:[350,∞)",
        ),
    ]

    for stage_name in ("Initial", "Second", "Third"):
        stage_cn = STAGE_LABELS[stage_name]
        rows.extend(
            [
                (
                    f"{stage_name}_Quoted_QPOs",
                    f"{stage_cn}报价QPO数",
                    "Stage",
                    f"Distinct QPO_ID count in {stage_name} stage under the bucket.",
                    "阶段内基数。",
                ),
                (
                    f"{stage_name}_Success_QPOs",
                    f"{stage_cn}成交QPO数",
                    "Stage",
                    f"Distinct QPO_ID count in {stage_name} stage with final status = Success.",
                    "按清洗后唯一 QPO_ID 计数。",
                ),
                (
                    f"{stage_name}_Success_Rate",
                    f"{stage_cn}成交占比",
                    "Stage",
                    f"{stage_name}_Success_QPOs / {stage_name}_Quoted_QPOs.",
                    f"阶段内 Success/Cancel/InProgress_Rate 分母均为 {stage_name}_Quoted_QPOs。",
                ),
                (
                    f"{stage_name}_Cancel_QPOs",
                    f"{stage_cn}取消QPO数",
                    "Stage",
                    f"Distinct QPO_ID count in {stage_name} stage with final status = Cancel.",
                    "按清洗后唯一 QPO_ID 计数。",
                ),
                (
                    f"{stage_name}_Cancel_Rate",
                    f"{stage_cn}取消占比",
                    "Stage",
                    f"{stage_name}_Cancel_QPOs / {stage_name}_Quoted_QPOs.",
                    f"分母 = {stage_name}_Quoted_QPOs。",
                ),
                (
                    f"{stage_name}_InProgress_QPOs",
                    f"{stage_cn}跟进中QPO数",
                    "Stage",
                    f"Distinct QPO_ID count in {stage_name} stage with final status = InProgress.",
                    "按清洗后唯一 QPO_ID 计数。",
                ),
                (
                    f"{stage_name}_InProgress_Rate",
                    f"{stage_cn}跟进中占比",
                    "Stage",
                    f"{stage_name}_InProgress_QPOs / {stage_name}_Quoted_QPOs.",
                    f"分母 = {stage_name}_Quoted_QPOs。",
                ),
                (
                    f"{stage_name}_WootCancel_QPOs",
                    f"{stage_cn}WootCancel数",
                    "Stage",
                    f"WootCancel count within {stage_name} stage.",
                    "阶段取消拆分口径。",
                ),
                (
                    f"{stage_name}_WootCancel_Share_of_StageCancel",
                    f"{stage_cn}WootCancel占阶段取消比",
                    "Stage",
                    f"{stage_name}_WootCancel_QPOs / {stage_name}_Cancel_QPOs.",
                    f"分母 = {stage_name}_Cancel_QPOs；与总体取消拆分口径分开。",
                ),
                (
                    f"{stage_name}_SellerCancel_QPOs",
                    f"{stage_cn}SellerCancel数",
                    "Stage",
                    f"SellerCancel count within {stage_name} stage.",
                    "阶段取消拆分口径。",
                ),
                (
                    f"{stage_name}_SellerCancel_Share_of_StageCancel",
                    f"{stage_cn}SellerCancel占阶段取消比",
                    "Stage",
                    f"{stage_name}_SellerCancel_QPOs / {stage_name}_Cancel_QPOs.",
                    f"分母 = {stage_name}_Cancel_QPOs；与总体取消拆分口径分开。",
                ),
            ]
        )

    rows.extend(
        [
            (
                "Input File",
                "输入文件",
                "AUDIT",
                "Resolved source workbook path.",
                str(metadata.get("input_path", "")),
            ),
            (
                "Source Sheet",
                "数据源Sheet",
                "AUDIT",
                "Detected data sheet name.",
                str(metadata.get("sheet_name", "")),
            ),
            (
                "Raw Data Rows",
                "原始数据行数",
                "AUDIT",
                "Rows iterated from the detected data sheet starting from data area.",
                str(metadata.get("raw_rows", 0)),
            ),
            (
                "Valid Date Rows",
                "有效日期行数",
                "AUDIT",
                "Rows with parsable Submission_Date.",
                str(metadata.get("valid_rows", 0)),
            ),
            (
                "Valid QPO Rows",
                "有效QPO行数",
                "AUDIT",
                "Rows with parsable date and non-empty QPO_ID.",
                str(metadata.get("valid_qpo_rows", 0)),
            ),
            (
                "Collected Unique QPOs",
                "归集后唯一QPO数",
                "AUDIT",
                "Unique QPO_ID count before anomaly deletion.",
                str(metadata.get("collected_qpo_count", 0)),
            ),
            (
                "Anomaly Deletion - Type A",
                "异常成交A类删除child行数",
                "AUDIT",
                "Deleted child rows where success rows have no quote and the whole QPO has no quote-stage flag.",
                str(metadata.get("deleted_type_a_rows", 0)),
            ),
            (
                "Anomaly Deletion - Type B",
                "异常成交B类整单剔除QPO数",
                "AUDIT",
                "Excluded QPOs where remaining rows still contain Success and all remaining child rows are Cancel.",
                str(metadata.get("deleted_type_b_qpo_count", 0)),
            ),
            (
                "Post-clean Unique QPOs",
                "清洗后QPO总数（去重）",
                "AUDIT",
                "Unique QPO_ID count after all anomaly cleaning.",
                str(metadata.get("post_clean_qpo_count", 0)),
            ),
            (
                "Bucketed Submitted Sum",
                "各区间提报QPO数求和",
                "AUDIT",
                "Sum of Submitted_QPOs across all displayed buckets.",
                str(audit.get("bucketed_submitted_sum", 0)),
            ),
            (
                "Unbucketed QPOs (missing Avg_P5)",
                "无法分桶QPO数（Avg_P5缺失或<0）",
                "AUDIT",
                "QPOs excluded from bucket summary because Avg_P5 is missing or invalid.",
                str(audit.get("unbucketed_qpos", 0)),
            ),
        ]
    )
    return rows


def write_workbook(
    rows: list[dict[str, object]],
    output_path: Path,
    metadata: dict[str, object],
    audit: dict[str, object],
) -> None:
    columns_en = [
        "P5_Bucket",
        "Submitted_QPOs",
        "Success_QPOs",
        "Success_Rate",
        "Cancel_QPOs",
        "Cancel_Rate",
        "InProgress_QPOs",
        "InProgress_Rate",
        "WootCancel_QPOs",
        "WootCancel_Share_of_Cancel",
        "SellerCancel_QPOs",
        "SellerCancel_Share_of_Cancel",
        "Avg_P5_Success",
        "Avg_P6_Success",
        "Avg_M1_Success",
        "M1_over_P6",
        "Initial_Quoted_QPOs",
        "Initial_Success_QPOs",
        "Initial_Success_Rate",
        "Initial_Cancel_QPOs",
        "Initial_Cancel_Rate",
        "Initial_InProgress_QPOs",
        "Initial_InProgress_Rate",
        "Initial_WootCancel_QPOs",
        "Initial_WootCancel_Share_of_StageCancel",
        "Initial_SellerCancel_QPOs",
        "Initial_SellerCancel_Share_of_StageCancel",
        "Second_Quoted_QPOs",
        "Second_Success_QPOs",
        "Second_Success_Rate",
        "Second_Cancel_QPOs",
        "Second_Cancel_Rate",
        "Second_InProgress_QPOs",
        "Second_InProgress_Rate",
        "Second_WootCancel_QPOs",
        "Second_WootCancel_Share_of_StageCancel",
        "Second_SellerCancel_QPOs",
        "Second_SellerCancel_Share_of_StageCancel",
        "Third_Quoted_QPOs",
        "Third_Success_QPOs",
        "Third_Success_Rate",
        "Third_Cancel_QPOs",
        "Third_Cancel_Rate",
        "Third_InProgress_QPOs",
        "Third_InProgress_Rate",
        "Third_WootCancel_QPOs",
        "Third_WootCancel_Share_of_StageCancel",
        "Third_SellerCancel_QPOs",
        "Third_SellerCancel_Share_of_StageCancel",
    ]
    columns_cn = [
        "P5区间",
        "提报QPO数",
        "成交QPO数",
        "成交占比（= 成交QPO数 / 提报QPO数）",
        "取消QPO数",
        "取消占比（= 取消QPO数 / 提报QPO数）",
        "跟进中QPO数",
        "跟进中占比（= 跟进中QPO数 / 提报QPO数）",
        "WootCancel数",
        "WootCancel占比(占取消QPO)（= WootCancel数 / 取消QPO数）",
        "SellerCancel数",
        "SellerCancel占比(占取消QPO)（= SellerCancel数 / 取消QPO数）",
        "成交平均P5（区间内 Success QPO 的 Avg_P5 均值）",
        "成交平均P6",
        "成交平均M1",
        "采购价占比(M1/P6)（= 区间层“成交平均M1 ÷ 成交平均P6”）",
        "一审报价QPO数",
        "一审成交数",
        "一审成交占比（= 一审成交数 / 一审报价QPO数）",
        "一审取消数",
        "一审取消占比（= 一审取消数 / 一审报价QPO数）",
        "一审跟进中数",
        "一审跟进中占比（= 一审跟进中数 / 一审报价QPO数）",
        "一审WootCancel数",
        "一审WootCancel占比(占阶段取消)（= 一审WootCancel数 / 一审取消数）",
        "一审SellerCancel数",
        "一审SellerCancel占比(占阶段取消)（= 一审SellerCancel数 / 一审取消数）",
        "二审报价QPO数",
        "二审成交数",
        "二审成交占比",
        "二审取消数",
        "二审取消占比",
        "二审跟进中数",
        "二审跟进中占比",
        "二审WootCancel数",
        "二审WootCancel占比(占阶段取消)",
        "二审SellerCancel数",
        "二审SellerCancel占比(占阶段取消)",
        "三审报价QPO数",
        "三审成交数",
        "三审成交占比",
        "三审取消数",
        "三审取消占比",
        "三审跟进中数",
        "三审跟进中占比",
        "三审WootCancel数",
        "三审WootCancel占比(占阶段取消)",
        "三审SellerCancel数",
        "三审SellerCancel占比(占阶段取消)",
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "UK208"
    sheet.append(columns_en)
    sheet.append(columns_cn)
    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    style_two_header_sheet(sheet, columns_en, columns_cn)
    sheet.freeze_panes = "A3"

    notes = workbook.create_sheet("Notes")
    note_rows = build_notes_rows(metadata, audit)
    for note_row in note_rows:
        notes.append(note_row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in notes[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap_alignment

    for row in notes.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap_alignment

    notes.freeze_panes = "A2"
    notes.column_dimensions["A"].width = 34
    notes.column_dimensions["B"].width = 28
    notes.column_dimensions["C"].width = 16
    notes.column_dimensions["D"].width = 58
    notes.column_dimensions["E"].width = 76

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def main() -> None:
    args = parse_args()
    input_path = resolve_input_path(args.input_path)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )

    records, metadata = build_clean_records(input_path)
    rows, audit = aggregate(records)
    write_workbook(rows, output_path, metadata, audit)

    print("生成完成：")
    print(output_path)


if __name__ == "__main__":
    main()
