# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import ast
import datetime as dt
import json
import re
import subprocess
import sys
import tempfile
import time
import urllib.error
import urllib.request
from http.client import RemoteDisconnected
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
DEFAULT_OUTPUT_DIR = BASE_DIR / "UKR_AIReport"
LOGIC_DOCS_DIR = BASE_DIR / "业务逻辑梳理"
SYNC_TRACKER_PATH = BASE_DIR / "业务逻辑梳理同步记录.md"
DEFAULT_GENERATED_EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_GENERATED_EXPORT_DIR_TEXT = str(DEFAULT_GENERATED_EXPORT_DIR)
OUTPUT_TIMESTAMP_FORMAT = "%m%d%H%M"
API_URL = "http://162.209.167.76/AssistantsApi/v5/ResponsesApi.php/create"
PROJECT_CODE = "MKP"
DEFAULT_MODEL = "gpt-5.4"
DEFAULT_API_TIMEOUT = 900
API_RETRY_ATTEMPTS = 3
API_RETRY_DELAY_SECONDS = 5

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

IDENTIFIER_RE = re.compile(r"\b[A-Za-z][A-Za-z0-9_]{2,}\b")
PURE_IDENTIFIER_RE = re.compile(r"^[A-Za-z][A-Za-z0-9_]{2,}$")
UKAI_METADATA_RE = re.compile(r"<!--\s*UKAI_METADATA\s*(\{.*?\})\s*-->", re.S)
UKAI_SYNC_TRACKER_RE = re.compile(r"<!--\s*UKAI_SYNC_TRACKER\s*(\{.*?\})\s*-->", re.S)
UKAI_AUTO_SUPPLEMENT_START = "<!-- UKAI_AUTO_SUPPLEMENT_START -->"
UKAI_AUTO_SUPPLEMENT_END = "<!-- UKAI_AUTO_SUPPLEMENT_END -->"
PLACEHOLDER_TOKEN_RE = re.compile(r"\b(?:TODO|TBD|FIXME|待补充|待实现|占位)\b", re.IGNORECASE)

FIELD_STOPWORDS = {
    "ChatGPT",
    "Anonymous",
    "Report",
    "This",
    "Summary",
    "Sheet",
    "Notes",
    "Excel",
    "Instructions",
    "GPTs",
    "GPT",
    "Date",
    "Month",
}

DIRECTIVE_KEYWORDS = (
    "现在我有一个需求",
    "需要统计",
    "需要体现的字段有",
    "请调整格式",
    "帮我增加",
    "请以",
    "导出对应结果表",
    "请你根据",
    "需要",
    "请",
    "帮我",
    "导出",
    "调整",
    "增加",
    "统计",
)

RULE_KEYWORDS = (
    "规则",
    "口径",
    "定义",
    "粒度",
    "去重",
    "重复",
    "30天",
    "前30天",
    "一周",
    "7天",
    "自然周",
    "rolling",
    "字段判断",
    "再次提交",
    "再次重复提交",
    "Date + QPO + 子ASIN",
    "Fulfillment_Type",
    "Pricing_Mode",
)

OUTPUT_KEYWORDS = (
    "Sheet",
    "双表头",
    "英文表头",
    "中文表头",
    "表头",
    "深色",
    "白色",
    "居中",
    "百分比",
    "两位小数",
    "Notes",
    "格式",
)

ANGLE_KEYWORDS = (
    "分析",
    "趋势",
    "交叉",
    "拆解",
    "管理",
    "建议",
    "Top",
    "风险",
    "联动",
    "UEA",
    "Seller",
    "部门",
    "客户",
)

CORRECTION_KEYWORDS = (
    "说错了",
    "改为",
    "已修正",
    "需要纠正",
    "不完全准确",
    "这个问题非常关键",
)

NOISE_KEYWORDS = (
    "如果你下一步要",
    "你只需要说一句",
    "一句话总结",
    "我可以帮你",
    "强烈建议",
    "这个问题非常关键",
    "很好，这个需求本质",
    "可以直接用于汇报",
    "我帮你",
    "已按你最新口径",
    "已经帮你",
)

KNOWN_ANALYSIS_FIELDS = {
    "Date",
    "QPO",
    "Parent_ASIN",
    "Child_ASIN",
    "Pricing_Mode",
    "Fulfillment_Type",
    "Total_Submission_Count",
    "Repeat_ASIN",
    "Repeat_ASIN_New",
    "Repeat_Submission_Count",
    "Repeat_Rate",
    "Repeat_Rate_vs_Submission",
    "Unique_ASIN",
    "MultiQuote_ASIN",
    "MultiFulfillment_ASIN",
    "Diff_Days",
    "Quote_Type",
    "First_Date",
    "Next_Date",
    "Week",
    "Quote_Type_Count",
    "Submission_Date",
    "QPO_ID",
}


@dataclass(frozen=True)
class ReportPattern:
    code: str
    family: str
    script_name: str
    file_prefixes: tuple[str, ...]
    output_name: str
    purpose: str
    dimensions: tuple[str, ...]
    metrics: tuple[str, ...]
    filters: tuple[str, ...]
    source_fields: tuple[str, ...]
    tags: tuple[str, ...]
    logic_doc_path: str = ""
    logic_doc_summary: str = ""


FAMILY_RULES = {
    "UK100": {
        "source_prefixes": ("UK100",),
        "source_table_rule": "识别 UK100 开头的基础表文件，不依赖完整文件名",
        "source_table_example": "UK100_工作量 KPI-*.xlsx",
        "theme": "工作量 / 工单 / 邮件效率",
        "shared_utils": (
            "uk_workload_utils.find_latest_input 自动取最新 UK100 基础表",
            "load_sheet_context 负责识别 Data Sheet、双表头和数据起始行",
            "get_required_indexes 校验字段是否存在",
            "style_two_header_sheet / write_notes_rows 负责双表头和 Notes 页格式",
            "uk_export_utils.build_default_output_path 生成带时间戳输出文件名",
        ),
        "shared_fields": (
            "Stat_Date",
            "Department",
            "UEA",
            "Type",
            "Metric_Code",
            "Metric_Name",
            "Ticket_No",
            "Order_ID",
            "seller Code",
            "DCM",
            "KCP",
            "Order_Level",
            "Handle_Time_Minutes",
            "TAM1_Cnt~TAM5_Cnt",
        ),
        "shared_filters": (
            "Type=工单 时常用于纯工单口径",
            "Type in {工单, 批量工单, U3工单, U6工单} 时常用于工单类总量口径",
            "Type=邮件 时用于邮件效率口径",
            "大量报表按月份汇总，月份来自 Stat_Date",
        ),
        "output_pattern": (
            "大多是单个 Result Sheet + Notes Sheet",
            "前两行分别为英文列名 / 中文列名",
            "冻结窗格 A3，数字/百分比格式统一处理",
        ),
    },
    "UK200": {
        "source_prefixes": ("UK200",),
        "source_table_rule": "识别 UK200 开头的基础表文件，不依赖完整文件名",
        "source_table_example": "UK200_转化漏斗KPI-*.xlsx",
        "theme": "QPO 转化漏斗 / 报价阶段 / 转化率",
        "shared_utils": (
            "uk_funnel_utils.find_latest_input 自动取最新 UK200 基础表",
            "load_stage_qpo_records 把多行明细归并成 QPO 级记录",
            "build_stage_breakdown 输出 Initial/Second/Third 三阶段指标",
            "style_two_header_sheet 统一双表头格式",
            "uk_export_utils.build_default_output_path 生成带时间戳输出文件名",
        ),
        "shared_fields": (
            "Submission_Date",
            "QPO_ID",
            "QPO_Status",
            "WPO_Current_Node（代码列 + 中文列）",
            "Order_Level",
            "Pricing_Mode",
            "Seller_ID",
            "OM_Flup",
            "KP_Flup",
            "Initial_Quote",
            "Second_Quote",
            "MVM_Approved_Price",
            "Child_Status",
            "P5",
            "P6",
            "M1",
            "Fulfillment_Type",
        ),
        "shared_filters": (
            "常见状态归类：Success / Cancel / InProgress",
            "很多报表按 Submission_Date 提取 Month，并额外补 ALL 汇总",
            "报价阶段通常按是否存在 Initial_Quote / Second_Quote / MVM_Approved_Price 推导",
            "取消拆分常见为 WootCancel 与 SellerCancel",
        ),
        "output_pattern": (
            "多数报表只有主结果 Sheet，也有脚本补充 Notes 或校验信息",
            "常见字段结构：维度列 + Submitted/Success/Cancel/InProgress + Rate + 分阶段指标",
            "多数脚本同时生成月度行和 ALL 汇总行",
        ),
    },
}


REPORT_PATTERNS = (
    ReportPattern(
        code="UK101",
        family="UK100",
        script_name="uk101.py",
        file_prefixes=("UK101",),
        output_name="UEA_个人工作量规模_结果表_月汇总_基于UK100.xlsx",
        purpose="按月 + 部门 + UEA 汇总个人工作量规模，并输出部门汇总和排名信息。",
        dimensions=("Month", "Department", "UEA"),
        metrics=("Case count", "Ticket count", "Email count", "Department summary", "Ranking"),
        filters=("按 Stat_Date 取月", "区分工单与邮件", "依赖 UK100 工作量基础表"),
        source_fields=("Stat_Date", "Department", "UEA", "Type", "Ticket_No"),
        tags=("个人工作量", "月汇总", "UEA", "排名", "工作量规模", "部门汇总"),
    ),
    ReportPattern(
        code="UK102",
        family="UK100",
        script_name="uk102.py",
        file_prefixes=("UK102",),
        output_name="UK102_UEA_月度工作量价值结构_20260430.xlsx",
        purpose="按月 + 部门 + UEA 分析工作量的 DCM、KCP、订单等级价值结构。",
        dimensions=("Stat_Month", "Department", "UEA"),
        metrics=("Total case", "DCM share", "KCP share", "Order level share", "High value order share"),
        filters=("Type in 工单类", "要求 Order_ID 和 seller Code 非空", "月份来自 Stat_Date"),
        source_fields=("Stat_Date", "Department", "UEA", "Type", "Order_ID", "seller Code", "DCM", "KCP", "Order_Level"),
        tags=("价值结构", "DCM", "KCP", "订单等级", "高价值", "UEA"),
    ),
    ReportPattern(
        code="UK103",
        family="UK100",
        script_name="uk103.py",
        file_prefixes=("UK103",),
        output_name="UK103_工单节点处理量分析表.xlsx",
        purpose="按部门 + 月份 + 节点名称分析工单节点处理量和平均强度。",
        dimensions=("Department", "Month", "Metric_Name"),
        metrics=("Distinct ticket count", "TAM1~TAM5 sums", "Average per ticket"),
        filters=("Type=工单", "Ticket_No/Metric_Name 必填", "月份来自 Stat_Date"),
        source_fields=("Stat_Date", "Department", "Type", "Metric_Name", "Ticket_No", "TAM1_Cnt", "TAM2_Cnt", "TAM3_Cnt", "TAM4_Cnt", "TAM5_Cnt"),
        tags=("节点", "工单处理量", "Metric_Name", "TAM", "平均每单", "部门"),
    ),
    ReportPattern(
        code="UK104",
        family="UK100",
        script_name="uk104.py",
        file_prefixes=("UK104",),
        output_name="UK104_QPO_工单处理量分析表_20260430.xlsx",
        purpose="先做 QPO 明细，再按订单等级、DCM、KCP 汇总工单处理量和 TAM 指标。",
        dimensions=("Order_ID detail", "Order_Level", "DCM", "KCP"),
        metrics=("QPO count", "Deal count", "Ticket count", "TAM1~TAM5 total", "Avg per ticket/QPO/Deal"),
        filters=("Type=工单", "以 Order_ID 为 QPO 粒度", "适合多张汇总 Sheet"),
        source_fields=("Stat_Date", "Type", "Order_ID", "Order_Level", "seller Code", "DCM", "KCP", "Ticket_No", "TAM1_Cnt", "TAM2_Cnt", "TAM3_Cnt", "TAM4_Cnt", "TAM5_Cnt"),
        tags=("QPO", "订单等级", "DCM", "KCP", "工单处理量", "多sheet", "明细+汇总"),
    ),
    ReportPattern(
        code="UK105",
        family="UK100",
        script_name="uk105.py",
        file_prefixes=("UK105",),
        output_name="UK105_UEA处理人邮件处理时效分析.xlsx",
        purpose="按处理人 + 邮件节点汇总邮件量、总处理时长和平均处理时长。",
        dimensions=("UEA", "Metric_Code", "Metric_Name"),
        metrics=("Email count", "Total handle minutes", "Average handle minutes"),
        filters=("Type=邮件", "Metric_Code/Metric_Name/UEA 必填"),
        source_fields=("Type", "UEA", "Metric_Code", "Metric_Name", "Handle_Time_Minutes"),
        tags=("邮件", "处理时效", "处理人", "Handle_Time_Minutes", "平均时长"),
    ),
    ReportPattern(
        code="UK201",
        family="UK200",
        script_name="uk201.py",
        file_prefixes=("UK201",),
        output_name="UK201_转化漏斗概览.xlsx",
        purpose="按月输出总提报、成交、取消、跟进中，以及按当前节点归因的取消分布。",
        dimensions=("Month", "WPO_Current_Node"),
        metrics=("Total QPO", "Deal/Cancel/InProgress count and rate", "Node cancel count and rate"),
        filters=("状态映射到 Deal/Cancel/InProgress", "取消归因取同 QPO 的最大节点", "补 ALL 汇总"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "WPO_Current_Node"),
        tags=("转化漏斗概览", "取消节点", "Node", "月度漏斗", "QPO状态"),
    ),
    ReportPattern(
        code="UK203",
        family="UK200",
        script_name="uk203.py",
        file_prefixes=("UK203",),
        output_name="UK203_Seller转化漏斗概览.xlsx",
        purpose="按 Seller 维度输出月度和历史累计转化漏斗表现。",
        dimensions=("Seller_ID", "Submit_Month"),
        metrics=("Total/Deal/Cancel/InProgress count and rate", "First/Last submit date"),
        filters=("依赖 Seller_ID", "补月度与 seller 累计两种视角"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "Seller_ID"),
        tags=("Seller", "卖家", "累计", "月度", "首末提报"),
    ),
    ReportPattern(
        code="UK204",
        family="UK200",
        script_name="uk204.py",
        file_prefixes=("UK204",),
        output_name="UK204_OM UEA转化漏斗概览.xlsx",
        purpose="按 OM_Flup 输出月度转化漏斗，并附带节点分布。",
        dimensions=("OM_Flup", "Stat_Month"),
        metrics=("Total/Success/Cancel/InProgress count and rate", "Node distribution"),
        filters=("默认只统计 2025-01-01 之后", "依赖 OM_Flup"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "OM_Flup", "WPO_Current_Node"),
        tags=("OM", "跟进人", "漏斗", "月度", "节点"),
    ),
    ReportPattern(
        code="UK205",
        family="UK200",
        script_name="uk205.py",
        file_prefixes=("UK205",),
        output_name="UK205_KP UEA转化漏斗概览.xlsx",
        purpose="按 KP_Flup 输出月度转化漏斗，并附带节点取消分布。",
        dimensions=("KP_Flup", "Stat_Month"),
        metrics=("Total/Success/Cancel/InProgress count and rate", "Node distribution"),
        filters=("默认只统计 2025-01-01 之后", "依赖 KP_Flup"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "KP_Flup", "WPO_Current_Node"),
        tags=("KP", "跟进人", "漏斗", "月度", "节点"),
    ),
    ReportPattern(
        code="UK206",
        family="UK200",
        script_name="uk206.py",
        file_prefixes=("UK206",),
        output_name="UK206｜各订单等级 × 报价阶段 转化漏斗.xlsx",
        purpose="按订单等级拆分 QPO，并分析 Initial/Second/Third 报价阶段转化。",
        dimensions=("Month", "Order_Level"),
        metrics=("Submitted/Success/Cancel/InProgress", "Stage count", "Stage success/cancel rate"),
        filters=("Order_Level 只保留 A/B/C/D", "阶段来自报价字段存在性", "补 ALL 汇总"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "Order_Level", "Initial_Quote", "Second_Quote", "MVM_Approved_Price"),
        tags=("订单等级", "报价阶段", "三阶段", "Initial", "Second", "Third", "ALL", "Order_Level", "转化"),
    ),
    ReportPattern(
        code="UK2061",
        family="UK200",
        script_name="uk2061.py",
        file_prefixes=("UK2061",),
        output_name="UK2061_各订单等级三次报价转化漏斗概览.xlsx",
        purpose="聚焦三次报价场景，按订单等级分析更深阶段的转化表现。",
        dimensions=("Month", "Order_Level"),
        metrics=("Third stage metrics", "Success/Cancel split", "Order-level funnel"),
        filters=("基于 load_stage_qpo_records 的阶段结果", "重点关注 Third 阶段"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "Order_Level", "Initial_Quote", "Second_Quote", "MVM_Approved_Price"),
        tags=("三次报价", "订单等级", "Third", "漏斗", "阶段"),
    ),
    ReportPattern(
        code="UK207",
        family="UK200",
        script_name="uk207.py",
        file_prefixes=("UK207",),
        output_name="UK207｜不同报价方式转化漏斗.xlsx",
        purpose="按 Pricing_Mode 维度输出转化漏斗，并细分阶段内取消类型。",
        dimensions=("Month", "Pricing_Mode"),
        metrics=("Submitted/Success/Cancel/InProgress", "Stage metrics", "WootCancel/SellerCancel split"),
        filters=("Pricing_Mode 非空", "阶段来自报价字段存在性", "补 ALL 汇总"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "Pricing_Mode", "Initial_Quote", "Second_Quote", "MVM_Approved_Price"),
        tags=("报价方式", "Pricing_Mode", "报价阶段", "Initial", "Second", "Third", "ALL", "取消拆分", "WootCancel", "SellerCancel"),
    ),
    ReportPattern(
        code="UK208",
        family="UK200",
        script_name="uk208.py",
        file_prefixes=("UK208",),
        output_name="UK208_P5价格区间转化漏斗概览.xlsx",
        purpose="按 P5 价格区间做转化漏斗，兼顾异常清洗与价格均值字段。",
        dimensions=("P5_Bucket", "Stage"),
        metrics=("Submitted/Success/Cancel/InProgress", "Price bucket funnel", "Avg P5/P6/M1"),
        filters=("按 P5 划分价格桶", "存在两类清洗规则", "QPO 粒度归并后分析"),
        source_fields=("Submission_Date", "QPO_ID", "QPO_Status", "Child_Status", "Initial_Quote", "Second_Quote", "MVM_Approved_Price", "P5", "P6", "M1"),
        tags=("P5", "价格区间", "价格带", "P6", "M1", "清洗"),
    ),
)


STATIC_CODE_TO_FAMILY = {
    "UK100": "UK100",
    "UK200": "UK200",
    **{pattern.code: pattern.family for pattern in REPORT_PATTERNS},
}

COMMON_CODEGEN_COMPATIBILITY_RULES = (
    "输出脚本必须能在当前代码目录通过 `python 脚本名.py --help` 启动，不能在导入阶段依赖不存在的路径、环境变量或第三方包。",
    "除标准库和 openpyxl 外，只允许使用当前代码目录已存在的 `uk_export_utils`、`uk_workload_utils`、`uk_funnel_utils` 中的真实函数，不要臆造新工具名。",
    "必须支持 `--input` 与 `--output` 参数；未传 `--input` 时自动识别最新基础表。",
    f"默认输出目录固定为 `{DEFAULT_GENERATED_EXPORT_DIR_TEXT}`；代码中应显式声明 `EXPORT_DIR = Path(r\"{DEFAULT_GENERATED_EXPORT_DIR_TEXT}\")`。",
    f"默认输出文件名必须通过 `build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)` 生成，使最终文件名在扩展名前追加时间戳，格式为 `{OUTPUT_TIMESTAMP_FORMAT}`，示例 `05071801`；同时应基于 `时间范围.txt` 在输出目录下创建对应时间范围子目录。",
    "即使过滤后没有数据，也要正常输出包含表头和 Notes 的 Excel，而不是抛异常。",
    "所有比例/除法必须做零分母保护；所有字段缺失错误要抛出明确 ValueError。",
    "读取工作簿时优先 `read_only=True, data_only=True`，并在处理完成后关闭 workbook。",
    "不要引入 pandas、numpy、xlwings 等额外依赖；保持和现有 UK 脚本同级别依赖。",
)

UTILITY_INTERFACE_HINTS = {
    "UK100": (
        "uk_export_utils.build_default_output_path(export_dir: Path, default_output_name: str, base_dir: Path | None = None) -> Path",
        "uk_workload_utils.find_latest_input(base_dir: Path) -> Path",
        "uk_workload_utils.load_sheet_context(input_path: Path) -> (workbook, sheet, header_row_1, header_row_2, data_start_row)",
        "uk_workload_utils.get_required_indexes(header_row: list[str], required_single: dict[str, str], duplicate_requirements: dict[str, tuple[str, int]] | None = None) -> dict[str, int]",
        "uk_workload_utils.normalize_text(value: object) -> str",
        "uk_workload_utils.has_value(value) -> bool",
        "uk_workload_utils.parse_cell_date(value) -> date | None",
        "uk_workload_utils.parse_cell_time(value) -> time | None",
        "uk_workload_utils.safe_div(numerator, denominator) -> float",
        "uk_workload_utils.round_half_up(value, digits=0) -> float",
        "uk_workload_utils.month_sort_key(month: str) -> tuple[int, int]",
        "uk_workload_utils.department_sort_key(department: str) -> tuple[int, str]",
        "uk_workload_utils.style_two_header_sheet(sheet, columns_en, columns_cn, *, date_columns=None, percent_columns=None, integer_columns=None, decimal_formats=None)",
        "uk_workload_utils.write_notes_rows(sheet, rows: list[tuple[object, ...]]) -> None",
    ),
    "UK200": (
        "uk_export_utils.build_default_output_path(export_dir: Path, default_output_name: str, base_dir: Path | None = None) -> Path",
        "uk_funnel_utils.find_latest_input(base_dir: Path) -> Path",
        "uk_funnel_utils.detect_data_sheet(workbook)",
        "uk_funnel_utils.get_required_indexes(header_row: list[str], required_single: dict[str, str], duplicate_requirements: dict[str, tuple[str, int]] | None = None) -> dict[str, int]",
        "uk_funnel_utils.normalize_text(value: object) -> str",
        "uk_funnel_utils.has_value(value) -> bool",
        "uk_funnel_utils.parse_cell_date(value) -> date | None",
        "uk_funnel_utils.safe_div(numerator, denominator) -> float",
        "uk_funnel_utils.month_sort_key(month: str) -> tuple[int, int]",
        "uk_funnel_utils.style_two_header_sheet(sheet, columns_en, columns_cn, *, date_columns=None)",
        "uk_funnel_utils.load_stage_qpo_records(input_path: Path) -> tuple[list[dict[str, object]], dict[str, object]]",
        "uk_funnel_utils.build_stage_breakdown(records, *, include_stage_inprogress=False, include_cancel_split=False) -> dict[str, int | float]",
        "uk_funnel_utils.map_simple_type(status: str) -> str | None",
        "uk_funnel_utils.map_split_type(status: str) -> str",
        "uk_funnel_utils.normalize_order_level(value: str) -> str",
    ),
}


def _to_tuple(value: object) -> tuple[str, ...]:
    if isinstance(value, (list, tuple)):
        return tuple(str(item) for item in value if str(item).strip())
    if value is None:
        return tuple()
    text = str(value).strip()
    return (text,) if text else tuple()


def extract_doc_summary(text: str) -> str:
    stripped = UKAI_METADATA_RE.sub("", text).strip()
    for block in re.split(r"\n\s*\n", stripped):
        line = block.strip()
        if not line or line.startswith("#"):
            continue
        return re.sub(r"\s+", " ", line)[:240]
    return ""


def extract_section_items(text: str, headings: tuple[str, ...]) -> tuple[str, ...]:
    lines = strip_metadata_block(text).splitlines()
    active = False
    items: list[str] = []
    heading_set = {item.strip() for item in headings}
    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            continue
        if line.startswith("#"):
            title = line.lstrip("#").strip()
            normalized_title = title.split("（", 1)[0].split("(", 1)[0].strip()
            active = any(key in normalized_title for key in heading_set)
            continue
        if not active:
            continue
        if line.startswith(("-", "*")):
            value = re.sub(r"^[\-\*]\s*", "", line).strip("` ").strip()
            if value:
                items.append(value)
    return tuple(dict.fromkeys(items))


def infer_family_from_doc(code: str, content: str) -> str:
    upper_content = content.upper()
    if "UK100" in upper_content:
        return "UK100"
    if "UK200" in upper_content:
        return "UK200"
    matched = re.search(r"UK(\d+)", code.upper())
    if matched:
        number = int(matched.group(1))
        if 100 <= number < 200:
            return "UK100"
        if 200 <= number < 300:
            return "UK200"
    raise ValueError(f"无法从文档自动识别基础表族：{code}")


def infer_output_name(content: str, code: str) -> str:
    for pattern in (
        r"目标输出文件[：:]\s*`([^`]+\.xlsx)`",
        r"默认输出文件名[：:]\s*`([^`]+\.xlsx)`",
        r"`([^`]*" + re.escape(code.upper()) + r"[^`]*\.xlsx)`",
    ):
        matched = re.search(pattern, content, flags=re.IGNORECASE)
        if matched:
            return matched.group(1).strip()
    return f"{code.upper()}.xlsx"


def infer_purpose(content: str) -> str:
    stripped = strip_metadata_block(content)
    for line in stripped.splitlines():
        text = line.strip()
        if not text or text.startswith("#"):
            continue
        return text[:240]
    return ""


def build_pattern_from_doc(doc_path: Path, content: str) -> ReportPattern:
    matched = UKAI_METADATA_RE.search(content)
    metadata: dict[str, object]
    if matched:
        metadata = json.loads(matched.group(1))
    else:
        code = doc_path.stem.upper()
        family = infer_family_from_doc(code, content)
        metadata = {
            "code": code,
            "family": family,
            "script_name": f"{code.lower()}.py",
            "file_prefixes": [code],
            "output_name": infer_output_name(content, code),
            "purpose": infer_purpose(content),
            "dimensions": extract_section_items(content, ("关键维度", "维度")),
            "metrics": extract_section_items(content, ("核心指标与口径", "指标")),
            "filters": extract_section_items(content, ("实现提示", "筛选条件", "过滤条件")),
            "source_fields": extract_section_items(content, ("基础文件与数据粒度", "源字段", "基础字段")),
            "tags": extract_section_items(content, ("核心分析角度", "标签")),
        }

    return ReportPattern(
        code=str(metadata["code"]).upper(),
        family=str(metadata["family"]),
        script_name=str(metadata.get("script_name") or f"{str(metadata['code']).lower()}.py"),
        file_prefixes=_to_tuple(metadata.get("file_prefixes") or [str(metadata["code"]).upper()]),
        output_name=str(metadata.get("output_name", "")),
        purpose=str(metadata.get("purpose", "")),
        dimensions=_to_tuple(metadata.get("dimensions")),
        metrics=_to_tuple(metadata.get("metrics")),
        filters=_to_tuple(metadata.get("filters")),
        source_fields=_to_tuple(metadata.get("source_fields")),
        tags=_to_tuple(metadata.get("tags")),
        logic_doc_path=doc_path.name,
        logic_doc_summary=extract_doc_summary(content),
    )


def load_report_patterns_from_docs() -> list[ReportPattern]:
    if not LOGIC_DOCS_DIR.exists():
        return []

    patterns: list[ReportPattern] = []
    for path in sorted(LOGIC_DOCS_DIR.glob("UK*.md")):
        content = path.read_text(encoding="utf-8")
        patterns.append(build_pattern_from_doc(path, content))
    return patterns


def get_available_patterns() -> tuple[ReportPattern, ...]:
    doc_patterns = load_report_patterns_from_docs()
    return tuple(doc_patterns) if doc_patterns else REPORT_PATTERNS


def build_code_to_family(patterns: tuple[ReportPattern, ...]) -> dict[str, str]:
    return {
        "UK100": "UK100",
        "UK200": "UK200",
        **{pattern.code: pattern.family for pattern in patterns},
    }


def strip_metadata_block(text: str) -> str:
    return UKAI_METADATA_RE.sub("", text).strip()


def resolve_logic_doc(
    logic_doc_ref: str,
    available_patterns: tuple[ReportPattern, ...],
) -> tuple[ReportPattern, Path, str]:
    ref = logic_doc_ref.strip()
    candidate_path = Path(ref)
    if candidate_path.exists():
        doc_path = candidate_path
    else:
        normalized = ref.upper()
        if normalized.endswith(".MD"):
            normalized = normalized[:-3]
        doc_path = LOGIC_DOCS_DIR / f"{normalized}.md"
        if not doc_path.exists():
            raise FileNotFoundError(f"未找到业务逻辑文档：{logic_doc_ref}")

    content = doc_path.read_text(encoding="utf-8")
    pattern = build_pattern_from_doc(doc_path, content)
    return pattern, doc_path, strip_metadata_block(content)


def now_text() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def log_stage(message: str) -> None:
    print(message, flush=True)


def format_mtime(path: Path) -> tuple[int | None, str]:
    if not path.exists():
        return None, "-"
    stat = path.stat()
    return stat.st_mtime_ns, dt.datetime.fromtimestamp(stat.st_mtime).isoformat(sep=" ", timespec="seconds")


def load_sync_tracker(path: Path) -> dict[str, dict[str, object]]:
    if not path.exists():
        return {}

    content = path.read_text(encoding="utf-8")
    matched = UKAI_SYNC_TRACKER_RE.search(content)
    if not matched:
        return {}

    try:
        payload = json.loads(matched.group(1))
    except json.JSONDecodeError:
        return {}

    entries = payload.get("entries", [])
    if not isinstance(entries, list):
        return {}

    tracker: dict[str, dict[str, object]] = {}
    for item in entries:
        if not isinstance(item, dict):
            continue
        code = str(item.get("code", "")).strip().upper()
        if not code:
            continue
        tracker[code] = item
    return tracker


def escape_markdown_cell(value: object) -> str:
    text = str(value or "").strip()
    if not text:
        return "-"
    return text.replace("|", "\\|").replace("\n", "<br>")


def build_sync_tracker_markdown(entries: dict[str, dict[str, object]], *, scanned_at: str) -> str:
    ordered_entries = [entries[code] for code in sorted(entries)]
    status_counts: dict[str, int] = {}
    for entry in ordered_entries:
        status = str(entry.get("status", "unknown")).strip() or "unknown"
        status_counts[status] = status_counts.get(status, 0) + 1

    metadata_payload = {
        "scanned_at": scanned_at,
        "entries": ordered_entries,
    }
    lines = [
        "<!-- UKAI_SYNC_TRACKER",
        json.dumps(metadata_payload, ensure_ascii=False, indent=2),
        "-->",
        "# 业务逻辑梳理同步记录",
        "",
        f"- 最近扫描时间：`{scanned_at}`",
        f"- 跟踪文件：`{SYNC_TRACKER_PATH.name}`",
        f"- 文档目录：`{LOGIC_DOCS_DIR.name}`",
        f"- 代码目录：`{SCRIPT_DIR.name}`",
    ]
    if status_counts:
        lines.append(
            "- 状态汇总：" + " / ".join(f"`{status}`={count}" for status, count in sorted(status_counts.items()))
        )
    lines.extend(
        [
            "",
            "| 报表编码 | 逻辑文档 | 对应脚本 | 文档更新时间 | 脚本更新时间 | 最近同步时间 | 状态 | 说明 |",
            "|---|---|---|---|---|---|---|---|",
        ]
    )
    if not ordered_entries:
        lines.append("| - | - | - | - | - | - | - | 当前没有检测到 `UK*.md` |")
    else:
        for entry in ordered_entries:
            lines.append(
                "| "
                + " | ".join(
                    [
                        escape_markdown_cell(entry.get("code")),
                        escape_markdown_cell(entry.get("logic_doc")),
                        escape_markdown_cell(entry.get("script_name")),
                        escape_markdown_cell(entry.get("logic_doc_mtime")),
                        escape_markdown_cell(entry.get("script_mtime")),
                        escape_markdown_cell(entry.get("last_synced_at")),
                        escape_markdown_cell(entry.get("status")),
                        escape_markdown_cell(entry.get("message")),
                    ]
                )
                + " |"
            )
    lines.extend(
        [
            "",
            "## 说明",
            "",
            "- 只跟踪 `业务逻辑梳理` 目录下的 `UK*.md`。",
            "- 新增文档会先按同类型（UK1 / UK2）参考其他逻辑文档，由 AI 追加一段“查漏补缺”的补充提示，但不会覆盖原主逻辑。",
            "- 自动同步默认只处理新增文档、文档更新时间发生变化的文档，或脚本缺失的文档。",
            "- 生成成功时会直接覆盖 `代码` 目录下对应的 `ukxxx.py`。",
            "- 若本次没有提供模型能力，会记录为 `blocked`；但只会在该文档再次新增/变更时自动重试。",
        ]
    )
    return "\n".join(lines) + "\n"


def write_sync_tracker(path: Path, entries: dict[str, dict[str, object]], *, scanned_at: str) -> None:
    path.write_text(build_sync_tracker_markdown(entries, scanned_at=scanned_at), encoding="utf-8")


def sort_report_code(code: str) -> tuple[int, int, str]:
    matched = re.fullmatch(r"UK(\d+)", code.upper())
    if not matched:
        return (999999, 999999, code.upper())
    digits = matched.group(1)
    primary = int(digits[:3]) if len(digits) >= 3 else int(digits)
    secondary = int(digits[3:]) if len(digits) > 3 else 0
    return (primary, secondary, code.upper())


def write_family_start_cmd(entries: dict[str, dict[str, object]], family: str) -> Path:
    start_cmd_path = BASE_DIR / ("UK100_start.cmd" if family == "UK100" else "UK200_start.cmd")
    base_script = "uk100_exe_accdb.py" if family == "UK100" else "uk200_exe_accdb.py"
    script_entries = [
        entry
        for entry in entries.values()
        if str(entry.get("code", "")).upper().startswith("UK1" if family == "UK100" else "UK2")
        and str(entry.get("script_name", "")).strip()
    ]
    script_entries.sort(key=lambda item: sort_report_code(str(item.get("code", ""))))

    lines = [
        "@echo off",
        "chcp 65001 > nul",
        'cd /d "%~dp0"',
        "",
        f'python ".\\代码\\{base_script}"',
    ]
    for entry in script_entries:
        script_name = str(entry["script_name"]).strip()
        script_path = SCRIPT_DIR / script_name
        if script_name in {"uk_ai_brief.py", "uk_export_utils.py", "uk_workload_utils.py", "uk_funnel_utils.py"}:
            continue
        if not script_path.exists():
            continue
        lines.append(f'python ".\\代码\\{script_name}"')
    lines.extend(["", "pause", ""])
    start_cmd_path.write_text("\n".join(lines), encoding="utf-8", newline="\n")
    return start_cmd_path


def refresh_start_cmd_files(entries: dict[str, dict[str, object]]) -> tuple[Path, Path]:
    return (
        write_family_start_cmd(entries, "UK100"),
        write_family_start_cmd(entries, "UK200"),
    )


def normalize_sync_message(message: str) -> str:
    text = re.sub(r"\s+", " ", message).strip()
    return text[:300] if len(text) > 300 else text


def build_sync_entry(
    *,
    code: str,
    logic_doc: Path,
    script_path: Path,
    previous_entry: dict[str, object] | None = None,
) -> dict[str, object]:
    previous_entry = previous_entry or {}
    logic_doc_mtime_ns, logic_doc_mtime = format_mtime(logic_doc)
    script_mtime_ns, script_mtime = format_mtime(script_path)
    return {
        "code": code,
        "logic_doc": logic_doc.name,
        "script_name": script_path.name,
        "logic_doc_mtime_ns": logic_doc_mtime_ns,
        "logic_doc_mtime": logic_doc_mtime,
        "script_mtime_ns": script_mtime_ns,
        "script_mtime": script_mtime,
        "last_synced_at": previous_entry.get("last_synced_at", "-"),
        "doc_augmented_at": previous_entry.get("doc_augmented_at", "-"),
        "doc_augmented_note": previous_entry.get("doc_augmented_note", ""),
        "status": previous_entry.get("status", "pending"),
        "message": previous_entry.get("message", ""),
    }


def describe_sync_reason(entry: dict[str, object], previous_entry: dict[str, object] | None) -> str:
    if previous_entry is None:
        return "新增文档"
    if previous_entry.get("logic_doc_mtime_ns") != entry.get("logic_doc_mtime_ns"):
        return "文档已更新"
    if entry.get("script_mtime_ns") is None:
        return "脚本缺失"
    return "手动触发"


def update_tracker_entry(
    pattern: ReportPattern,
    doc_path: Path,
    script_path: Path,
    *,
    status: str,
    message: str,
    last_synced_at: str = "-",
    doc_augmented_at: str | None = None,
    doc_augmented_note: str | None = None,
) -> None:
    tracker_entries = load_sync_tracker(SYNC_TRACKER_PATH)
    previous_entry = tracker_entries.get(pattern.code)
    entry = build_sync_entry(
        code=pattern.code,
        logic_doc=doc_path,
        script_path=script_path,
        previous_entry=previous_entry,
    )
    entry["status"] = status
    entry["message"] = message
    entry["last_synced_at"] = last_synced_at
    if doc_augmented_at is not None:
        entry["doc_augmented_at"] = doc_augmented_at
    if doc_augmented_note is not None:
        entry["doc_augmented_note"] = doc_augmented_note
    tracker_entries[pattern.code] = entry
    write_sync_tracker(SYNC_TRACKER_PATH, tracker_entries, scanned_at=now_text())
    refresh_start_cmd_files(tracker_entries)


def should_sync_logic_doc(entry: dict[str, object], previous_entry: dict[str, object] | None) -> bool:
    if previous_entry is None:
        return True
    if previous_entry.get("logic_doc_mtime_ns") != entry.get("logic_doc_mtime_ns"):
        return True
    return entry.get("script_mtime_ns") is None


def sync_logic_docs(
    args: argparse.Namespace,
    available_patterns: tuple[ReportPattern, ...],
) -> dict[str, object]:
    scanned_at = now_text()
    tracker_entries = load_sync_tracker(SYNC_TRACKER_PATH)
    updated_entries: dict[str, dict[str, object]] = {}
    summary = {"scanned": 0, "synced": 0, "blocked": 0, "failed": 0, "unchanged": 0}

    if not LOGIC_DOCS_DIR.exists():
        write_sync_tracker(SYNC_TRACKER_PATH, {}, scanned_at=scanned_at)
        start_cmd_paths = refresh_start_cmd_files(updated_entries)
        return {"tracker_path": SYNC_TRACKER_PATH, "start_cmd_paths": start_cmd_paths, "summary": summary, "entries": updated_entries}

    for doc_path in sorted(LOGIC_DOCS_DIR.glob("UK*.md")):
        content = doc_path.read_text(encoding="utf-8")
        pattern = build_pattern_from_doc(doc_path, content)
        previous_entry = tracker_entries.get(pattern.code)
        target_script_path = SCRIPT_DIR / pattern.script_name
        entry = build_sync_entry(
            code=pattern.code,
            logic_doc=doc_path,
            script_path=target_script_path,
            previous_entry=previous_entry,
        )
        summary["scanned"] += 1

        if should_sync_logic_doc(entry, previous_entry):
            log_stage(f"[SYNC] {pattern.code}: {describe_sync_reason(entry, previous_entry)}，开始处理。")
            if not args.model:
                entry["status"] = "blocked"
                entry["message"] = "检测到新增或更新，但本次未提供 --model，无法自动生成脚本。"
                summary["blocked"] += 1
            else:
                try:
                    sync_notes: list[str] = []
                    if previous_entry is None:
                        log_stage(f"[SYNC] {pattern.code}: 开始补充新增文档提示词。")
                        augment_message = enrich_new_logic_doc(
                            doc_path,
                            pattern,
                            args,
                            available_patterns,
                        )
                        entry["doc_augmented_at"] = scanned_at
                        entry["doc_augmented_note"] = augment_message
                        sync_notes.append(augment_message)
                        content = doc_path.read_text(encoding="utf-8")
                        pattern = build_pattern_from_doc(doc_path, content)
                        target_script_path = SCRIPT_DIR / pattern.script_name

                    log_stage(f"[SYNC] {pattern.code}: 开始生成脚本 {target_script_path.name}。")
                    prompt = build_codegen_prompt(
                        pattern,
                        strip_metadata_block(content),
                        target_script_name=target_script_path.name,
                        available_patterns=available_patterns,
                    )
                    _, response_text, validation_issues = generate_and_validate_script(
                        prompt,
                        args,
                        target_script_name=target_script_path.name,
                    )
                    if validation_issues:
                        raise RuntimeError(format_validation_issues(validation_issues))
                    existed_before = target_script_path.exists()
                    write_text(target_script_path, response_text)
                    entry = build_sync_entry(
                        code=pattern.code,
                        logic_doc=doc_path,
                        script_path=target_script_path,
                        previous_entry=entry,
                    )
                    entry["last_synced_at"] = scanned_at
                    entry["status"] = "synced"
                    sync_notes.append("已覆盖生成脚本。" if existed_before else "已生成新脚本。")
                    entry["message"] = " ".join(sync_notes)
                    log_stage(f"[SYNC] {pattern.code}: 脚本生成成功 -> {target_script_path.name}")
                    summary["synced"] += 1
                except Exception as exc:
                    entry = build_sync_entry(
                        code=pattern.code,
                        logic_doc=doc_path,
                        script_path=target_script_path,
                        previous_entry=entry,
                    )
                    entry["status"] = "failed"
                    entry["message"] = normalize_sync_message(str(exc))
                    log_stage(f"[SYNC] {pattern.code}: 处理失败 -> {entry['message']}")
                    summary["failed"] += 1
        else:
            entry["status"] = "synced"
            entry["message"] = previous_entry.get("message", "文档未变化，无需重新生成。") if previous_entry else "文档未变化，无需重新生成。"
            entry["last_synced_at"] = previous_entry.get("last_synced_at", "-") if previous_entry else "-"
            summary["unchanged"] += 1

        updated_entries[pattern.code] = entry

    write_sync_tracker(SYNC_TRACKER_PATH, updated_entries, scanned_at=scanned_at)
    start_cmd_paths = refresh_start_cmd_files(updated_entries)
    return {
        "tracker_path": SYNC_TRACKER_PATH,
        "start_cmd_paths": start_cmd_paths,
        "summary": summary,
        "entries": updated_entries,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据业务逻辑梳理目录中的提示词文档或需求文本，生成 UK 报表脚本 prompt / 结果分析 / 脚本文件。"
    )
    requirement_group = parser.add_mutually_exclusive_group(required=False)
    requirement_group.add_argument("--requirement", help="新的报表需求描述。")
    requirement_group.add_argument("--requirement-file", help="从文本文件读取需求或 AI 对话记录。")
    parser.add_argument("--family", choices=("auto", "UK100", "UK200"), default="auto")
    parser.add_argument("--source-mode", choices=("auto", "plain", "dialogue", "logic_doc"), default="auto")
    parser.add_argument("--logic-doc", help="业务逻辑文档编码或路径，例如 UK101 或 UK101.md。")
    parser.add_argument("--top-k", type=int, default=3, help="选取最相关历史报表数量，默认 3。")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--save-artifacts", action="store_true", help="显式保存 brief / prompt 文件；默认不落盘。")
    parser.add_argument("--excel-output", help="结果 Excel 路径；不传时自动生成。")
    parser.add_argument("--target-script", help="基于逻辑文档生成脚本时的目标 .py 路径。")
    parser.add_argument("--write-script", action="store_true", help="调用模型接口后，将返回结果写入目标脚本文件。")
    parser.add_argument("--skip-excel", action="store_true", help="不生成 Excel。")
    parser.add_argument("--call-api", action="store_true", help="直接调用大模型接口。")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"调用接口时使用的模型名，默认 {DEFAULT_MODEL}。")
    parser.add_argument("--sync-only", action="store_true", help="只执行业务逻辑文档自动同步，然后退出。")
    parser.add_argument("--skip-auto-sync", action="store_true", help="跳过启动时自动扫描并同步业务逻辑文档。")
    parser.add_argument("--timeout", type=int, default=DEFAULT_API_TIMEOUT, help=f"接口调用超时秒数，默认 {DEFAULT_API_TIMEOUT}。")
    parser.add_argument(
        "--skip-generated-script-check",
        action="store_true",
        help="基于逻辑文档生成脚本时，跳过本地语法/启动校验。",
    )
    parser.add_argument(
        "--codegen-repair-attempts",
        type=int,
        default=1,
        help="生成脚本未通过校验时的自动修复重试次数，默认 1。",
    )
    parser.add_argument(
        "--script-check-timeout",
        type=int,
        default=20,
        help="生成脚本本地启动校验超时秒数，默认 20。",
    )
    parser.add_argument("--max-output-tokens", type=int, default=20480000)
    parser.add_argument("--print-prompt", action="store_true", help="同时在控制台打印 prompt。")
    parser.add_argument("--print-response", action="store_true", help="调用接口后在控制台打印回复文本。")
    args = parser.parse_args()
    return args


def normalize_for_match(text: str) -> str:
    return re.sub(r"\s+", "", text).lower()


def extract_file_prefixes(text: str) -> tuple[str, ...]:
    prefixes = {match.upper() for match in re.findall(r"UK\d+", text, flags=re.IGNORECASE)}
    return tuple(sorted(prefixes))


def contains_prefix_reference(text: str, prefix: str) -> bool:
    normalized_text = normalize_for_match(text)
    normalized_prefix = normalize_for_match(prefix)
    if normalized_prefix in normalized_text:
        return True
    return any(token.startswith(normalized_prefix) for token in extract_file_prefixes(text))


def load_requirement(args: argparse.Namespace) -> tuple[str, str]:
    if args.requirement:
        return args.requirement.strip(), "inline_requirement"
    file_path = Path(args.requirement_file)
    return file_path.read_text(encoding="utf-8"), file_path.stem


def detect_source_mode(content: str, preferred_mode: str) -> str:
    if preferred_mode != "auto":
        return preferred_mode

    lowered = content.lower()
    conversation_hits = 0
    for token in (
        "copy of a conversation",
        "report conversation",
        "已按你最新口径",
        "如果你下一步要",
        "这个问题非常关键",
        "一句话总结",
        "sheet1",
        "notes 表",
        "👉",
    ):
        if token in lowered or token in content:
            conversation_hits += 1

    return "dialogue" if conversation_hits >= 2 else "plain"


def infer_family(
    requirement: str,
    specified_family: str,
    available_patterns: tuple[ReportPattern, ...],
) -> str:
    if specified_family != "auto":
        return specified_family

    code_to_family = build_code_to_family(available_patterns)
    scores = {"UK100": 0, "UK200": 0}
    for prefix in extract_file_prefixes(requirement):
        matched_family = code_to_family.get(prefix) or STATIC_CODE_TO_FAMILY.get(prefix)
        if not matched_family:
            continue
        if prefix in {"UK100", "UK200"}:
            scores[matched_family] += 8
        else:
            scores[matched_family] += 3

    text = requirement.lower()
    uk100_score = 0
    uk200_score = 0

    for keyword in ("uk100", "工作量", "工单", "邮件", "uea", "部门", "metric"):
        if keyword.lower() in text:
            uk100_score += 1
    for keyword in ("uk200", "漏斗", "转化", "qpo", "报价", "seller", "om", "kp", "pricing", "p5", "asin", "fulfillment"):
        if keyword.lower() in text:
            uk200_score += 1

    scores["UK100"] += uk100_score
    scores["UK200"] += uk200_score
    return "UK200" if scores["UK200"] > scores["UK100"] else "UK100"


def score_pattern(pattern: ReportPattern, requirement: str, family: str) -> int:
    if pattern.family != family:
        return -1

    text = normalize_for_match(requirement)
    score = 0
    if contains_prefix_reference(requirement, pattern.code):
        score += 8
    for prefix in pattern.file_prefixes:
        if contains_prefix_reference(requirement, prefix):
            score += 6

    for tag in pattern.tags:
        if normalize_for_match(tag) in text:
            score += 3
    for dimension in pattern.dimensions:
        if normalize_for_match(dimension) in text:
            score += 2
    for field in pattern.source_fields:
        if normalize_for_match(field) in text:
            score += 3
    for metric in pattern.metrics:
        if normalize_for_match(metric) in text:
            score += 2
    for rule in pattern.filters:
        if normalize_for_match(rule) in text:
            score += 2
    return score


def select_patterns(
    requirement: str,
    family: str,
    top_k: int,
    available_patterns: tuple[ReportPattern, ...],
) -> list[ReportPattern]:
    scored = [
        (score_pattern(pattern, requirement, family), pattern)
        for pattern in available_patterns
        if pattern.family == family
    ]
    scored.sort(key=lambda item: (-item[0], item[1].code))

    matched = [pattern for score, pattern in scored if score > 0][:top_k]
    if matched:
        return matched

    return [pattern for _, pattern in scored[:top_k]]


def normalize_line(line: str) -> str:
    return line.strip().replace("\ufeff", "")


def dedupe_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for item in items:
        key = normalize_for_match(item)
        if not key or key in seen:
            continue
        seen.add(key)
        ordered.append(item)
    return ordered


def is_noise_line(line: str) -> bool:
    lowered = line.lower()
    return any(keyword.lower() in lowered or keyword in line for keyword in NOISE_KEYWORDS)


def collect_latest_lines(lines: list[str], keywords: tuple[str, ...], limit: int) -> list[str]:
    results: list[str] = []
    for line in reversed(lines):
        if len(results) >= limit:
            break
        if not line or len(line) > 160:
            continue
        if is_noise_line(line):
            continue
        lowered = line.lower()
        if any(keyword.lower() in lowered or keyword in line for keyword in keywords):
            results.append(line)
    results.reverse()
    return dedupe_preserve_order(results)


def is_useful_directive_line(line: str) -> bool:
    if is_noise_line(line) or len(line) > 160:
        return False
    if line.startswith(("👉", "✅", "🧠", "⚠️", "🚀", "🔥", "###")):
        return False
    lowered = line.lower()
    if "字段有" in line or "口径" in line or "格式" in line or "统计" in line:
        return True
    if "为什么" in line or "对吗" in line:
        return True
    if re.match(r"^\d+[.、]", line):
        return True
    return any(keyword.lower() in lowered or keyword in line for keyword in DIRECTIVE_KEYWORDS)


def collect_latest_by_predicate(lines: list[str], predicate, limit: int) -> list[str]:
    results: list[str] = []
    for line in reversed(lines):
        if len(results) >= limit:
            break
        if predicate(line):
            results.append(line)
    results.reverse()
    return dedupe_preserve_order(results)


def extract_named_metrics(lines: list[str]) -> list[str]:
    metrics: list[str] = []
    for line in reversed(lines):
        if "字段有" not in line and "字段含义" not in line and "Summary表应该变成" not in line:
            continue
        if "：" in line:
            payload = line.split("：", 1)[1]
        elif ":" in line:
            payload = line.split(":", 1)[1]
        else:
            payload = line
        for part in re.split(r"[，、；,|]", payload):
            item = part.strip()
            if 2 <= len(item) <= 40:
                metrics.append(item)
        if len(metrics) >= 8:
            break
    metrics.reverse()
    return dedupe_preserve_order(metrics)


def is_candidate_identifier(token: str) -> bool:
    if token in FIELD_STOPWORDS:
        return False
    if token.startswith("UK") and token[2:].isdigit():
        return False
    if token not in KNOWN_ANALYSIS_FIELDS and "_" not in token:
        return False
    if len(token) < 3:
        return False
    return True


def extract_field_catalog(lines: list[str]) -> list[dict[str, str]]:
    positions: dict[str, int] = {}

    for index, line in enumerate(lines):
        for candidate in IDENTIFIER_RE.findall(line):
            if not is_candidate_identifier(candidate):
                continue
            positions[candidate] = index

    ordered_fields = [field_name for field_name, _ in sorted(positions.items(), key=lambda item: item[1], reverse=True)]

    catalog: list[dict[str, str]] = []
    for field_name in ordered_fields:
        index = positions[field_name]
        detail_parts: list[str] = []
        for follow_index in range(index, min(index + 4, len(lines))):
            follow_line = lines[follow_index]
            if not follow_line or is_noise_line(follow_line):
                continue
            if follow_index > index and PURE_IDENTIFIER_RE.match(follow_line):
                break
            detail_parts.append(follow_line)
        catalog.append(
            {
                "field_name": field_name,
                "definition": " ".join(detail_parts[:3]).strip(),
                "source_line": lines[index],
            }
        )

    return catalog[:20]


def build_context_profile(content: str, source_mode: str) -> dict[str, object]:
    lines = [normalize_line(line) for line in content.splitlines()]
    lines = [line for line in lines if line]

    final_directives = collect_latest_by_predicate(lines, is_useful_directive_line, 12)
    business_rules = collect_latest_lines(lines, RULE_KEYWORDS, 18)
    output_requirements = collect_latest_lines(lines, OUTPUT_KEYWORDS, 10)
    analysis_angles = collect_latest_lines(lines, ANGLE_KEYWORDS, 12)
    terminology_corrections = collect_latest_lines(lines, CORRECTION_KEYWORDS, 8)
    chinese_metrics = extract_named_metrics(lines)
    field_catalog = extract_field_catalog(lines)
    detected_prefixes = extract_file_prefixes(content)

    capability_gaps = [
        "旧版工具只能把整段文本作为普通需求描述，无法识别对话中的多轮修正。",
        "旧版工具不会抽取字段定义、术语更正、格式要求、分析角度。",
        "旧版工具不会生成结构化 Excel 结果。",
    ]

    completeness = {
        "source_mode": source_mode,
        "is_dialogue_record": source_mode == "dialogue",
        "before_upgrade_sufficient": False,
        "after_upgrade_sufficient": True,
        "capability_gaps": capability_gaps,
    }

    return {
        "detected_prefixes": detected_prefixes,
        "final_directives": final_directives,
        "business_rules": business_rules,
        "output_requirements": output_requirements,
        "analysis_angles": analysis_angles,
        "terminology_corrections": terminology_corrections,
        "field_catalog": field_catalog,
        "named_metrics": chinese_metrics,
        "completeness": completeness,
    }


def build_dynamic_sections(
    family: str,
    source_mode: str,
    context_profile: dict[str, object],
    patterns: list[ReportPattern],
) -> list[dict[str, object]]:
    sections: list[dict[str, object]] = []

    if source_mode == "dialogue":
        sections.append(
            {
                "title": "对话恢复要求",
                "items": [
                    "输入是一份 AI 对话记录，请先恢复最终有效需求。",
                    "如果对话中有旧口径与新口径冲突，必须以后出现的更正为准。",
                    "不要把中途被纠正的字段名、统计口径直接沿用到最终方案。",
                ],
            }
        )

    sections.append(
        {
            "title": "基础表判断",
            "items": [
                f"建议基于 {family} 基础表处理。",
                FAMILY_RULES[family]["source_table_rule"],
                f"基础表前缀：{', '.join(FAMILY_RULES[family]['source_prefixes'])}",
                f"基础表示例：{FAMILY_RULES[family]['source_table_example']}",
            ],
        }
    )

    if context_profile["final_directives"]:
        sections.append(
            {
                "title": "最终用户诉求",
                "items": list(context_profile["final_directives"]),
            }
        )

    if context_profile["terminology_corrections"]:
        sections.append(
            {
                "title": "术语与口径修正",
                "items": list(context_profile["terminology_corrections"]),
            }
        )

    if context_profile["business_rules"]:
        sections.append(
            {
                "title": "业务规则与计算口径",
                "items": list(context_profile["business_rules"]),
            }
        )

    if context_profile["field_catalog"]:
        sections.append(
            {
                "title": "字段与定义",
                "items": [
                    f"{field['field_name']}：{field['definition']}"
                    for field in context_profile["field_catalog"]
                ],
            }
        )

    if context_profile["named_metrics"]:
        sections.append(
            {
                "title": "中文指标要求",
                "items": list(context_profile["named_metrics"]),
            }
        )

    if context_profile["output_requirements"]:
        sections.append(
            {
                "title": "输出格式要求",
                "items": list(context_profile["output_requirements"]),
            }
        )

    if context_profile["analysis_angles"]:
        sections.append(
            {
                "title": "可扩展分析角度",
                "items": list(context_profile["analysis_angles"]),
            }
        )

    sections.append(
        {
            "title": "可复用脚本",
            "items": [
                f"{pattern.code}/{pattern.script_name}：{pattern.purpose}"
                + (f"（逻辑文档：{pattern.logic_doc_path}）" if pattern.logic_doc_path else "")
                for pattern in patterns
            ],
        }
    )

    sections.append(
        {
            "title": "逻辑文档要点",
            "items": [
                " | ".join(
                    [
                        f"{pattern.code}",
                        f"字段={', '.join(pattern.source_fields)}",
                        f"指标={', '.join(pattern.metrics)}",
                        f"摘要={pattern.logic_doc_summary or pattern.purpose}",
                    ]
                )
                for pattern in patterns
            ],
        }
    )

    sections.append(
        {
            "title": "你需要输出的内容",
            "items": [
                "恢复最终版本需求，并标明哪些口径已被后续对话覆盖。",
                "给出推荐字段、分组维度、指标、过滤条件和 Excel 输出结构。",
                "给出建议的 Python 实现结构，尽量复用现有公共工具函数。",
                "补充扩展分析角度，不要只复述用户原句。",
                "如果仍有歧义，列出待确认项。",
            ],
        }
    )

    return sections


def build_brief_payload(
    requirement: str,
    source_name: str,
    family: str,
    source_mode: str,
    patterns: list[ReportPattern],
    context_profile: dict[str, object],
) -> dict[str, object]:
    return {
        "generated_at": dt.datetime.now().isoformat(timespec="seconds"),
        "source_name": source_name,
        "source_mode": source_mode,
        "requirement": requirement,
        "family": family,
        "family_rules": FAMILY_RULES[family],
        "context_profile": context_profile,
        "recommended_reports": [asdict(pattern) for pattern in patterns],
        "dynamic_sections": build_dynamic_sections(family, source_mode, context_profile, patterns),
        "implementation_template": {
            "steps": [
                "恢复最终版本需求",
                "find_latest_input / 指定输入文件",
                "构造清洗层或行为判定层",
                "build_summary / build_detail",
                "write_workbook + Notes",
                "print 输出路径",
            ],
            "notes": [
                "如果输入是对话记录，先抽取最终有效口径，再生成 prompt。",
                "优先复用 uk_workload_utils 或 uk_funnel_utils 的公共函数。",
                "如果是 UK200 行为分析，允许在现有脚本体系基础上新增专用清洗逻辑。",
                "输出最好同时包含 Summary / Detail / Notes。",
            ],
        },
    }


def build_prompt(brief: dict[str, object]) -> str:
    family = str(brief["family"])
    family_rules = brief["family_rules"]
    dynamic_sections = brief["dynamic_sections"]
    context_profile = brief["context_profile"]

    prompt_lines = [
        "你现在是企业内部 Excel 报表开发助手。",
        "请基于下面的历史脚本规则和对话内容恢复结果，输出尽量贴近现有代码风格的实现草案。",
        "",
        f"【输入来源】\n{brief['source_name']} ({brief['source_mode']})",
        f"【建议基础表族】\n{family}",
        f"【基础表识别规则】\n{family_rules['source_table_rule']}",
        f"【基础表前缀】\n{', '.join(family_rules['source_prefixes'])}",
        f"【基础表示例】\n{family_rules['source_table_example']}",
        f"【主题】\n{family_rules['theme']}",
        "",
        "【共享工具】",
    ]
    prompt_lines.extend(f"- {item}" for item in family_rules["shared_utils"])
    prompt_lines.append("【共享字段】")
    prompt_lines.extend(f"- {item}" for item in family_rules["shared_fields"])
    prompt_lines.append("")
    prompt_lines.append("【输入原文】")
    prompt_lines.append(str(brief["requirement"]).strip())
    prompt_lines.append("")

    for section in dynamic_sections:
        prompt_lines.append(f"【{section['title']}】")
        prompt_lines.extend(f"- {item}" for item in section["items"])
        prompt_lines.append("")

    prompt_lines.append("【补充约束】")
    prompt_lines.append("- 不要把已经被后续对话修正的旧字段或旧口径继续带入最终方案。")
    prompt_lines.append("- 如果对话里出现了字段名替换，必须明确写出最终采用哪个字段。")
    prompt_lines.append("- 输出时要兼顾：最终指标表、明细表、Notes 表、格式要求。")
    prompt_lines.append("- 可以扩展分析角度、字段与管理解释，但必须基于现有对话与历史脚本。")
    prompt_lines.append("")
    prompt_lines.append("【当前工具判断】")
    prompt_lines.append(
        f"- 旧版工具是否足够：{'是' if context_profile['completeness']['before_upgrade_sufficient'] else '否'}"
    )
    prompt_lines.extend(f"- {item}" for item in context_profile["completeness"]["capability_gaps"])
    return "\n".join(prompt_lines)


def collect_family_reference_scripts(
    pattern: ReportPattern,
    available_patterns: tuple[ReportPattern, ...],
    *,
    limit: int = 6,
) -> tuple[str, ...]:
    references: list[str] = []
    for candidate in available_patterns:
        script_name = candidate.script_name.strip()
        if not script_name or candidate.code == pattern.code:
            continue
        if candidate.family != pattern.family:
            continue
        if script_name in {"uk100_exe_accdb.py", "uk200_exe_accdb.py"}:
            continue
        references.append(script_name)
    return tuple(sorted(references)[:limit])


def get_report_numeric_value(code: str) -> int:
    matched = re.fullmatch(r"UK(\d+)", code.upper())
    return int(matched.group(1)) if matched else 999999


def collect_similar_reference_docs(
    pattern: ReportPattern,
    available_patterns: tuple[ReportPattern, ...],
    *,
    limit: int = 2,
) -> list[tuple[ReportPattern, str]]:
    references: list[tuple[int, int, ReportPattern, str]] = []
    target_value = get_report_numeric_value(pattern.code)
    for candidate in available_patterns:
        if candidate.code == pattern.code or candidate.family != pattern.family:
            continue
        doc_name = candidate.logic_doc_path or f"{candidate.code}.md"
        doc_path = LOGIC_DOCS_DIR / doc_name
        if not doc_path.exists():
            continue
        content = strip_metadata_block(doc_path.read_text(encoding="utf-8")).strip()
        if not content:
            continue
        distance = abs(get_report_numeric_value(candidate.code) - target_value)
        references.append((distance, get_report_numeric_value(candidate.code), candidate, content))
    references.sort(key=lambda item: (item[0], item[1]))
    return [(candidate, content) for _, _, candidate, content in references[:limit]]


def build_reference_doc_brief(pattern: ReportPattern, content: str) -> str:
    sections = extract_section_items(content, ("关键维度", "维度", "核心指标与口径", "指标", "实现提示", "输出结构"))
    compact_items = list(sections)[:8]
    if not compact_items:
        compact_items = [extract_doc_summary(content) or pattern.purpose]
    return (
        f"{pattern.code} | 用途={pattern.purpose or '-'} | "
        f"维度={', '.join(pattern.dimensions) or '-'} | "
        f"指标={', '.join(pattern.metrics[:5]) if pattern.metrics else '-'} | "
        f"补充要点={' ; '.join(compact_items)}"
    )


def load_existing_target_script(pattern: ReportPattern) -> str:
    script_name = pattern.script_name.strip()
    if not script_name or script_name in {"uk100_exe_accdb.py", "uk200_exe_accdb.py"}:
        return ""
    script_path = SCRIPT_DIR / script_name
    if not script_path.exists():
        return ""
    return script_path.read_text(encoding="utf-8")


def build_codegen_prompt(
    pattern: ReportPattern,
    doc_body: str,
    *,
    target_script_name: str,
    available_patterns: tuple[ReportPattern, ...],
) -> str:
    family_rules = FAMILY_RULES[pattern.family]
    reference_scripts = collect_family_reference_scripts(pattern, available_patterns)
    reference_docs = collect_similar_reference_docs(pattern, available_patterns)
    existing_target_script = load_existing_target_script(pattern)
    prompt_lines = [
        "你现在是企业内部 Python 报表脚本生成助手。",
        "业务逻辑梳理目录中的 Markdown 已经是代码生成提示词，请严格执行。",
        "请依据下面的提示词文档，生成一个像 uk101.py 这种可直接落地输出 Excel 的正式 Python 脚本。",
        "输出必须是纯 Python 代码，不要加 markdown 代码块，不要加解释。",
        "",
        f"【目标脚本名】\n{target_script_name}",
        f"【报表编码】\n{pattern.code}",
        f"【所属基础表族】\n{pattern.family}",
        f"【脚本定位】\n{pattern.purpose}",
        f"【基础表识别规则】\n{family_rules['source_table_rule']}",
        f"【基础表前缀】\n{', '.join(family_rules['source_prefixes'])}",
        f"【基础表示例】\n{family_rules['source_table_example']}",
        "",
        "【生成模式】",
        "- 直接把 Markdown 文档中的提示词和业务规则落实到代码。",
        "- 输出必须是正式可运行脚本，不要只写方案、注释骨架或伪代码。",
        "- 不要生成或修改 uk100_exe_accdb.py、uk200_exe_accdb.py 这一类脚本。",
        "- 如果代码目录里已存在同名正式脚本，把它当作黄金参考脚本；在不违背当前 Markdown 口径的前提下，尽量最小改动式复刻。",
        "",
        "【优先复用的公共工具】",
    ]
    prompt_lines.extend(f"- {item}" for item in family_rules["shared_utils"])
    if reference_scripts:
        prompt_lines.append("【同族正式脚本参考】")
        prompt_lines.extend(f"- {item}" for item in reference_scripts)
    if reference_docs:
        prompt_lines.append("【相近逻辑文档参考（轻量）】")
        prompt_lines.extend(
            f"- {build_reference_doc_brief(candidate, content)}"
            for candidate, content in reference_docs
        )
    prompt_lines.append("【当前代码目录真实存在的公共函数】")
    prompt_lines.extend(f"- {item}" for item in UTILITY_INTERFACE_HINTS.get(pattern.family, ()))
    prompt_lines.append("【共享字段】")
    prompt_lines.extend(f"- {item}" for item in family_rules["shared_fields"])
    prompt_lines.append("")
    prompt_lines.append("【源提示词全文】")
    prompt_lines.append(doc_body.strip())
    prompt_lines.append("")
    if existing_target_script:
        prompt_lines.append("【黄金参考脚本全文】")
        prompt_lines.append(existing_target_script.strip())
        prompt_lines.append("")
    prompt_lines.append("【代码要求】")
    prompt_lines.append("- 必须提供 parse_args / main / 写 Excel 的完整结构。")
    prompt_lines.append("- 必须包含可实际落地的核心处理函数，例如 build_summary / build_detail / write_workbook。")
    prompt_lines.append("- 优先复用现有代码目录下的 uk_export_utils / uk_workload_utils / uk_funnel_utils。")
    prompt_lines.append("- 输出 Excel 保持中英文双表头、深色表头、白字、A3 冻结、Notes sheet。")
    prompt_lines.append("- 如果文档中提到多个 sheet、多个维度、漏斗、明细/汇总，都要完整落到代码。")
    prompt_lines.append("- 如果文档中出现“代码复刻约束”章节，必须优先遵守其中的输出文件名、Sheet 名、列顺序、固定枚举和函数结构。")
    prompt_lines.append("- 文档里已经明确的常量、列名、顺序、固定映射，不要自行改名、删减或重排。")
    prompt_lines.append(
        f"- 默认输出目录固定使用 `Path(r\"{DEFAULT_GENERATED_EXPORT_DIR_TEXT}\")`，不要改成其他目录，也不要回退到本地目录。"
    )
    prompt_lines.append(
        f"- 默认输出文件名必须用 `build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)` 生成，时间戳格式固定为 `{OUTPUT_TIMESTAMP_FORMAT}`，即文件名尾部追加类似 `-05071801`；并在输出根目录下按 `时间范围.txt` 创建时间范围子目录。"
    )
    prompt_lines.append("- 代码风格尽量对齐现有 UK 脚本。")
    prompt_lines.append("- 不要省略关键字段映射、状态映射、窗口判断、去重规则。")
    prompt_lines.append("- 如果黄金参考脚本已经满足文档口径，优先输出与其高度一致的版本，而不是重新发明一套结构。")
    prompt_lines.append("- 生成结果必须是完整可运行代码，不能包含 TODO、待补充、pass 占位实现或 NotImplementedError。")
    prompt_lines.append("- 需要显式处理空数据、坏日期、缺字段、输出目录不存在等兼容场景。")
    prompt_lines.append("- 除业务规则外，优先复用现有脚本中的列顺序、sheet 命名、格式化函数和 Notes 写法。")
    prompt_lines.append("【兼容性与稳健性要求】")
    prompt_lines.extend(f"- {item}" for item in COMMON_CODEGEN_COMPATIBILITY_RULES)
    return "\n".join(prompt_lines)


def collect_family_reference_docs(
    pattern: ReportPattern,
    available_patterns: tuple[ReportPattern, ...],
    *,
    limit: int = 2,
) -> list[tuple[ReportPattern, str]]:
    return collect_similar_reference_docs(pattern, available_patterns, limit=limit)


DOC_TEMPLATE_SECTION_RULES = (
    ("## 0. Prompt / 给 `uk_ai_brief.py` 的指令", ("prompt / 给", "prompt/给", "prompt / 给 uk_ai_brief.py 的指令")),
    ("## 1. 业务目标", ("业务目标",)),
    ("## 2. 基础文件与数据粒度", ("基础文件与数据粒度", "基础文件", "数据粒度")),
    ("## 3. 核心分析角度", ("核心分析角度", "分析角度")),
    ("## 4. 关键维度", ("关键维度", "维度")),
    ("## 5. 核心指标与口径", ("核心指标与口径", "指标与口径", "指标口径")),
    ("## 6. 漏斗 / 阶段 / 状态逻辑", ("漏斗 / 阶段 / 状态逻辑", "阶段 / 状态逻辑", "状态逻辑", "漏斗")),
    ("## 7. Excel 输出结构", ("excel 输出结构", "输出结构")),
    ("## 8. 实现提示", ("实现提示",)),
    ("## 9. 代码复刻约束", ("代码复刻约束", "复刻约束")),
)


def detect_missing_doc_sections(doc_body: str) -> list[str]:
    normalized_body = normalize_for_match(doc_body)
    missing: list[str] = []
    for canonical_title, aliases in DOC_TEMPLATE_SECTION_RULES:
        if not any(normalize_for_match(alias) in normalized_body for alias in aliases):
            missing.append(canonical_title)
    return missing


def build_doc_supplement_prompt(
    pattern: ReportPattern,
    doc_body: str,
    available_patterns: tuple[ReportPattern, ...],
) -> str:
    family_label = "UK1 开头" if pattern.family == "UK100" else "UK2 开头"
    reference_docs = collect_family_reference_docs(pattern, available_patterns)
    missing_sections = detect_missing_doc_sections(doc_body)
    prompt_lines = [
        "你现在是企业内部 Markdown 提示词补全助手。",
        "目标：仅为一个新建的 UK 报表逻辑文档补充缺失的提示词说明，不能覆盖、删改或重写原文中的主要业务逻辑。",
        "输出必须是一个可直接追加到原 Markdown 末尾的补充章节，不要输出完整文档，不要加代码块。",
        "如果原文已经足够完整，没有明显缺项，就只输出 `NO_CHANGES`。",
        "",
        f"【目标报表编码】\n{pattern.code}",
        f"【所属类型】\n{pattern.family}（{family_label}）",
        f"【对应脚本】\n{pattern.script_name}",
        "",
        "【补充原则】",
        "- 只能做查漏补缺，不能推翻原文业务目标、字段口径、Sheet 结构、命名和固定规则。",
        "- 优先补充这类内容：Prompt 指令、实现约束、输出细节、Notes 说明、校验与兼容性要求。",
        "- 如果原文已经写明某项规则，不要重复改写，只补新的缺失说明。",
        "- 输出建议以 `## AI补充提示（自动生成）` 为标题，下面只补缺失的小节。",
        "",
        "【固定模板约束】",
        "- 优先围绕以下模板章节补缺，不要发散增加无关内容：",
    ]
    prompt_lines.extend(f"- {title}" for title, _ in DOC_TEMPLATE_SECTION_RULES)
    prompt_lines.extend(
        [
            "- 已存在章节不要重写；若需要补充已存在章节，只能补“附加约束/边界条件/Notes要求”。",
            "- 输出只允许使用以下补充小节标题：`### 缺失章节补充`、`### 输出与Notes补充`、`### 实现与校验补充`。",
            "",
            "【当前检测到的缺失章节】",
        ]
    )
    if missing_sections:
        prompt_lines.extend(f"- {title}" for title in missing_sections)
    else:
        prompt_lines.append("- 无明显缺失章节，仅检查是否缺少边界条件、Notes 要求或实现约束。")
    prompt_lines.extend(
        [
            "",
            "【建议输出模板】",
            "## AI补充提示（自动生成）",
            "",
            "### 缺失章节补充",
            "- 仅补缺失章节或缺失条目",
            "",
            "### 输出与Notes补充",
            "- 仅补输出结构、Notes、边界情况",
            "",
            "### 实现与校验补充",
            "- 仅补实现限制、兼容性、校验要求",
            "",
            "【当前新文档正文】",
            doc_body.strip(),
            "",
            "【同类型参考文档】",
        ]
    )
    if reference_docs:
        for candidate, content in reference_docs:
            prompt_lines.extend(
                [
                    f"### {candidate.code} / {candidate.logic_doc_path or candidate.code + '.md'}",
                    f"- 用途：{candidate.purpose}",
                    f"- 维度：{', '.join(candidate.dimensions)}",
                    f"- 指标：{', '.join(candidate.metrics)}",
                    f"- 关键字段：{', '.join(candidate.source_fields)}",
                    content[:2500],
                    "",
                ]
            )
    else:
        prompt_lines.append("- 当前没有可用同类型参考文档。")
        prompt_lines.append("")

    prompt_lines.extend(
        [
            "【输出要求】",
            "- 仅输出“追加补充块”的 Markdown 内容。",
            "- 不要输出完整文档，不要重复原文大段内容。",
            "- 不能出现“我帮你”“下面是”之类对话口吻。",
            "- 不要增加与原文冲突的新业务口径。",
            "- 如果没有缺口，就只输出 `NO_CHANGES`。",
        ]
    )
    return "\n".join(prompt_lines)


def normalize_generated_markdown_text(text: str) -> str:
    content = text.replace("\ufeff", "").strip()
    if content.startswith("```"):
        lines = content.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        content = "\n".join(lines).strip()
    return content


def enrich_new_logic_doc(
    doc_path: Path,
    pattern: ReportPattern,
    args: argparse.Namespace,
    available_patterns: tuple[ReportPattern, ...],
) -> str:
    original_content = doc_path.read_text(encoding="utf-8")
    if UKAI_AUTO_SUPPLEMENT_START in original_content and UKAI_AUTO_SUPPLEMENT_END in original_content:
        return "文档已存在自动补充块，跳过补充。"

    doc_body = strip_metadata_block(original_content)
    prompt = build_doc_supplement_prompt(pattern, doc_body, available_patterns)
    result = call_model_api(prompt, args)
    supplement_text = normalize_generated_markdown_text(str(result["response_text"]))
    if not supplement_text or supplement_text == "NO_CHANGES":
        return "新文档已检查，无需补充提示词说明。"

    if not supplement_text.lstrip().startswith("## "):
        supplement_text = "## AI补充提示（自动生成）\n\n" + supplement_text.lstrip()

    updated_content = original_content.rstrip() + "\n\n"
    updated_content += UKAI_AUTO_SUPPLEMENT_START + "\n"
    updated_content += supplement_text.rstrip() + "\n"
    updated_content += UKAI_AUTO_SUPPLEMENT_END + "\n"
    doc_path.write_text(updated_content, encoding="utf-8")
    return "已为新文档追加同类型提示词补充。"


def write_text(path: Path, content: str) -> None:
    path.write_text(content, encoding="utf-8")


def save_codegen_prompt(output_dir: Path, pattern_code: str, prompt: str) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    prompt_path = output_dir / f"{pattern_code}_codegen_prompt.txt"
    write_text(prompt_path, prompt)
    return prompt_path


def save_artifacts(output_dir: Path, base_name: str, brief: dict[str, object], prompt: str) -> tuple[Path, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    brief_path = output_dir / f"{base_name}_brief.json"
    prompt_path = output_dir / f"{base_name}_prompt.txt"
    brief_path.write_text(json.dumps(brief, ensure_ascii=False, indent=2), encoding="utf-8")
    write_text(prompt_path, prompt)
    return brief_path, prompt_path


def extract_response_text(payload: object) -> str:
    if isinstance(payload, str):
        return payload
    if isinstance(payload, list):
        parts = [extract_response_text(item) for item in payload]
        return "\n".join(part for part in parts if part.strip())
    if isinstance(payload, dict):
        payload_type = str(payload.get("type", "")).strip()
        if payload_type == "reasoning":
            return ""
        if payload_type == "output_text" and isinstance(payload.get("text"), str):
            return str(payload["text"])
        if payload_type == "message":
            extracted = extract_response_text(payload.get("content"))
            if extracted.strip():
                return extracted

        for key in ("output", "content", "response_text", "response", "answer", "output_text", "message", "data"):
            value = payload.get(key)
            if value:
                extracted = extract_response_text(value)
                if extracted.strip():
                    return extracted
        text_value = payload.get("text")
        if isinstance(text_value, str) and text_value.strip():
            return text_value
    return json.dumps(payload, ensure_ascii=False, indent=2)


def normalize_generated_script_text(text: str) -> str:
    content = text.replace("\ufeff", "").strip()
    if content.startswith("```"):
        lines = content.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        content = "\n".join(lines).strip()

    module_docstring_pattern = re.compile(
        r'^(?P<prefix>(?:#.*\n)*)?(?P<quote>"""|\'\'\')(?P<body>.*?)(?P=quote)',
        re.DOTALL,
    )
    match = module_docstring_pattern.match(content)
    if match:
        body = match.group("body").replace("\\", "\\\\")
        prefix = match.group("prefix") or ""
        quote = match.group("quote")
        content = f"{prefix}{quote}{body}{quote}{content[match.end():]}"
    if not content.startswith("# -*- coding: utf-8 -*-"):
        if content.startswith("#!"):
            first_line, _, remainder = content.partition("\n")
            content = f"{first_line}\n# -*- coding: utf-8 -*-\n{remainder.lstrip()}"
        else:
            content = "# -*- coding: utf-8 -*-\n" + content
    return content + "\n"


def is_placeholder_function(node: ast.FunctionDef | ast.AsyncFunctionDef) -> bool:
    if len(node.body) != 1:
        return False
    statement = node.body[0]
    if isinstance(statement, ast.Pass):
        return True
    if isinstance(statement, ast.Expr) and isinstance(statement.value, ast.Constant) and statement.value.value is Ellipsis:
        return True
    if not isinstance(statement, ast.Raise):
        return False
    exc = statement.exc
    if isinstance(exc, ast.Call):
        target = exc.func
        if isinstance(target, ast.Name) and target.id == "NotImplementedError":
            return True
        if isinstance(target, ast.Attribute) and target.attr == "NotImplementedError":
            return True
    return False


def validate_generated_script_structure(script_text: str, target_script_name: str) -> list[str]:
    issues: list[str] = []
    try:
        tree = ast.parse(script_text, filename=target_script_name)
    except SyntaxError as exc:
        location = f"L{exc.lineno}:{exc.offset}" if exc.lineno else "未知位置"
        issues.append(f"语法错误（{location}）：{exc.msg}")
        return issues

    functions = {
        node.name: node
        for node in tree.body
        if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef))
    }
    for required_name in ("parse_args", "write_workbook", "main"):
        node = functions.get(required_name)
        if node is None:
            issues.append(f"缺少必要函数：{required_name}()")
            continue
        if is_placeholder_function(node):
            issues.append(f"必要函数仍是占位实现：{required_name}()")

    if 'if __name__ == "__main__":' not in script_text:
        issues.append("缺少 `if __name__ == \"__main__\":` 启动入口。")

    if PLACEHOLDER_TOKEN_RE.search(script_text):
        issues.append("代码中仍包含 TODO/TBD/待补充 等占位词。")

    return issues


def validate_generated_script_output_conventions(script_text: str) -> list[str]:
    issues: list[str] = []
    if DEFAULT_GENERATED_EXPORT_DIR_TEXT not in script_text:
        issues.append(
            f"默认导出目录不是 `{DEFAULT_GENERATED_EXPORT_DIR_TEXT}`。"
        )

    if "build_default_output_path(" not in script_text and OUTPUT_TIMESTAMP_FORMAT not in script_text:
        issues.append(
            f"默认输出文件名未体现 `{OUTPUT_TIMESTAMP_FORMAT}` 时间戳规则；请使用 `build_default_output_path(...)` 或等价实现。"
        )

    return issues


def validate_generated_script_runtime(
    script_text: str,
    target_script_name: str,
    timeout: int,
) -> list[str]:
    temp_file = tempfile.NamedTemporaryFile(
        mode="w",
        suffix=".py",
        prefix="_ukai_validate_",
        dir=SCRIPT_DIR,
        delete=False,
        encoding="utf-8",
        newline="\n",
    )
    temp_path = Path(temp_file.name)
    try:
        temp_file.write(script_text)
        temp_file.close()
        completed = subprocess.run(
            [sys.executable, str(temp_path), "--help"],
            cwd=SCRIPT_DIR,
            capture_output=True,
            text=True,
            timeout=max(10, timeout),
        )
    except subprocess.TimeoutExpired:
        return [f"`python {target_script_name} --help` 本地启动校验超时。"]
    finally:
        if not temp_file.closed:
            temp_file.close()
        temp_path.unlink(missing_ok=True)

    if completed.returncode == 0:
        return []

    detail = (completed.stderr or completed.stdout or "").strip()
    detail_lines = [line.rstrip() for line in detail.splitlines() if line.strip()]
    detail_preview = "\n".join(detail_lines[-12:]) if detail_lines else "无详细输出"
    return [f"`python {target_script_name} --help` 启动失败：\n{detail_preview}"]


def validate_generated_script(
    script_text: str,
    target_script_name: str,
    *,
    timeout: int,
) -> list[str]:
    issues = validate_generated_script_structure(script_text, target_script_name)
    if issues:
        return issues
    issues = validate_generated_script_output_conventions(script_text)
    if issues:
        return issues
    return validate_generated_script_runtime(script_text, target_script_name, timeout)


def format_validation_issues(issues: list[str]) -> str:
    return "\n".join(f"- {item}" for item in issues)


def build_script_repair_prompt(
    original_prompt: str,
    generated_script: str,
    issues: list[str],
    *,
    target_script_name: str,
) -> str:
    prompt_lines = [
        "你上一次生成的 Python 报表脚本未通过本地校验，请直接修复后重新输出完整代码。",
        "输出必须仍然是纯 Python 代码，不要加 markdown 代码块，不要加解释。",
        "",
        f"【目标脚本名】\n{target_script_name}",
        "【本地校验失败原因】",
    ]
    prompt_lines.extend(f"- {item}" for item in issues)
    prompt_lines.append("【修复要求】")
    prompt_lines.append("- 必须保留原始业务口径、Sheet 名、列顺序、输出文件名和核心函数结构。")
    prompt_lines.append("- 必须修复所有语法、导入、占位实现、默认路径兼容性、空数据兼容性和 `--help` 启动问题。")
    prompt_lines.append("- 只能调用当前代码目录真实存在的公共函数，不要新造工具函数名。")
    prompt_lines.append("- 不能输出 TODO、TBD、待补充、pass 占位、NotImplementedError。")
    prompt_lines.append(
        f"- 默认导出目录必须固定为 `Path(r\"{DEFAULT_GENERATED_EXPORT_DIR_TEXT}\")`。"
    )
    prompt_lines.append(
        f"- 默认文件名必须在扩展名前追加 `{OUTPUT_TIMESTAMP_FORMAT}` 格式时间戳，优先使用 `build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)`，并将文件输出到按 `时间范围.txt` 自动创建的时间范围子目录。"
    )
    prompt_lines.append("")
    prompt_lines.append("【原始生成要求】")
    prompt_lines.append(original_prompt)
    prompt_lines.append("")
    prompt_lines.append("【待修复脚本全文】")
    prompt_lines.append(generated_script.strip())
    return "\n".join(prompt_lines)


def generate_and_validate_script(
    prompt: str,
    args: argparse.Namespace,
    *,
    target_script_name: str,
) -> tuple[dict[str, object], str, list[str]]:
    repair_attempts = max(args.codegen_repair_attempts, 0)
    current_prompt = prompt
    result: dict[str, object] = {}
    response_text = ""
    issues: list[str] = []

    for attempt in range(repair_attempts + 1):
        log_stage(f"[AI] {target_script_name}: 正在请求模型（第 {attempt + 1}/{repair_attempts + 1} 次），超时 {args.timeout}s。")
        result = call_model_api(current_prompt, args)
        response_text = normalize_generated_script_text(str(result["response_text"]))
        if args.skip_generated_script_check:
            log_stage(f"[AI] {target_script_name}: 已收到模型回复，跳过本地校验。")
            return result, response_text, []

        log_stage(f"[AI] {target_script_name}: 已收到模型回复，开始本地校验。")
        issues = validate_generated_script(
            response_text,
            target_script_name,
            timeout=args.script_check_timeout,
        )
        if not issues:
            log_stage(f"[AI] {target_script_name}: 本地校验通过。")
            return result, response_text, []
        if attempt >= repair_attempts:
            break
        log_stage(f"[AI] {target_script_name}: 脚本未通过本地校验，正在自动修复（{attempt + 1}/{repair_attempts}）。")
        current_prompt = build_script_repair_prompt(
            prompt,
            response_text,
            issues,
            target_script_name=target_script_name,
        )

    return result, response_text, issues


def call_model_api(prompt: str, args: argparse.Namespace) -> dict[str, object]:
    payload = {
        "projectCode": PROJECT_CODE,
        "model": args.model or DEFAULT_MODEL,
        "input": prompt,
        "max_output_tokens": args.max_output_tokens,
        "store": True,
        "reasoning": {
            "effort": "medium",
        },
    }
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        API_URL,
        data=data,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    last_error: Exception | None = None
    for attempt in range(1, API_RETRY_ATTEMPTS + 1):
        try:
            with urllib.request.urlopen(request, timeout=args.timeout) as response:
                raw = response.read().decode("utf-8")
            break
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"接口调用失败：HTTP {exc.code}\n{detail}") from exc
        except (urllib.error.URLError, TimeoutError, RemoteDisconnected) as exc:
            last_error = exc
            if attempt >= API_RETRY_ATTEMPTS:
                raise RuntimeError(
                    f"接口调用失败（已重试 {API_RETRY_ATTEMPTS} 次）：{exc}"
                ) from exc
            log_stage(
                f"[AI] 接口调用异常（第 {attempt}/{API_RETRY_ATTEMPTS} 次）：{exc}；"
                f" {API_RETRY_DELAY_SECONDS}s 后自动重试。"
            )
            time.sleep(API_RETRY_DELAY_SECONDS)
    else:
        raise RuntimeError(f"接口调用失败：{last_error}")

    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        parsed = {"raw_response": raw}

    return {
        "request_payload": payload,
        "raw_response": parsed,
        "response_text": extract_response_text(parsed),
    }


def style_sheet(sheet, columns_en: list[str], columns_cn: list[str], left_align_columns: set[str] | None = None) -> None:
    left_align_columns = left_align_columns or set()

    for row_num in (1, 2):
        for col_num in range(1, len(columns_en) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

    preview_rows = min(sheet.max_row, 60)
    for idx, column in enumerate(columns_en, start=1):
        sample_lengths = [len(str(column)), len(str(columns_cn[idx - 1]))]
        for row_num in range(3, preview_rows + 1):
            value = sheet.cell(row=row_num, column=idx).value
            if value is not None:
                sample_lengths.append(len(str(value)))
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(sample_lengths) + 2, 60)

    for row_num in range(3, sheet.max_row + 1):
        for idx, column in enumerate(columns_en, start=1):
            cell = sheet.cell(row=row_num, column=idx)
            cell.alignment = LEFT if column in left_align_columns else CENTER

    sheet.freeze_panes = "A3"


def write_two_header_sheet(
    workbook: Workbook,
    title: str,
    columns_en: list[str],
    columns_cn: list[str],
    rows: list[list[object]],
    *,
    left_align_columns: set[str] | None = None,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(columns_en)
    sheet.append(columns_cn)
    for row in rows:
        sheet.append(row)
    style_sheet(sheet, columns_en, columns_cn, left_align_columns=left_align_columns)


def build_excel_rows(brief: dict[str, object], prompt: str) -> dict[str, list[list[object]]]:
    context = brief["context_profile"]
    reports = brief["recommended_reports"]

    assessment_rows = [
        ["Capability", "Before_Upgrade_Sufficient", "旧版工具是否可直接给出完善结果", "No", "旧版工具仅支持静态需求文本"],
        ["Capability", "After_Upgrade_Sufficient", "升级后是否可处理本输入", "Yes", "已支持对话恢复、字段抽取、动态 prompt、Excel 输出"],
        ["Input", "Source_Mode", "输入类型", brief["source_mode"], brief["source_name"]],
        ["Input", "Family", "建议基础表族", brief["family"], ", ".join(context["detected_prefixes"])],
        ["Input", "Detected_Prefixes", "检测到的文件前缀", ", ".join(context["detected_prefixes"]) or "-", "来源文本"],
    ]

    spec_rows: list[list[object]] = []
    for item in context["final_directives"]:
        spec_rows.append(["Final_Directive", "最终用户诉求", item, "High", brief["source_name"]])
    for item in context["terminology_corrections"]:
        spec_rows.append(["Correction", "术语或口径修正", item, "High", brief["source_name"]])
    for item in context["business_rules"]:
        spec_rows.append(["Business_Rule", "业务规则", item, "High", brief["source_name"]])
    for item in context["output_requirements"]:
        spec_rows.append(["Output_Format", "输出格式要求", item, "Medium", brief["source_name"]])
    for item in context["analysis_angles"]:
        spec_rows.append(["Analysis_Angle", "可扩展分析角度", item, "Medium", brief["source_name"]])

    field_rows = [
        [
            field["field_name"],
            field["definition"],
            field["source_line"],
            brief["source_name"],
        ]
        for field in context["field_catalog"]
    ]

    pattern_rows = [
        [
            report["code"],
            report["script_name"],
            ", ".join(report["file_prefixes"]),
            report["purpose"],
            ", ".join(report["dimensions"]),
            ", ".join(report["source_fields"]),
        ]
        for report in reports
    ]

    prompt_rows = []
    for section in brief["dynamic_sections"]:
        for item in section["items"]:
            prompt_rows.append([section["title"], item])

    notes_rows = [
        ["Before_Upgrade_Sufficient", "旧版工具是否能直接输出完善结果", "否，旧版不会恢复多轮对话中的最终口径。"],
        ["After_Upgrade_Sufficient", "升级后是否可处理", "是，现已支持读取对话记录并提炼最终规则。"],
        ["Field_Catalog", "字段目录来源", "从对话中自动抽取英文字段名，并绑定最近一次定义。"],
        ["Final_Directive", "最终诉求", "优先保留对话后段的修正规则与最新字段要求。"],
        ["Correction", "术语修正", "例如 Pricing_Mode 被更正为 Fulfillment_Type 时，以后者为准。"],
        ["Pattern_Match", "历史脚本匹配", "结合前缀、字段、维度、标签与规则做匹配，不依赖完整文件名。"],
        ["Prompt", "Prompt生成方式", "按对话内容动态扩展章节，而不是固定写死一套问题。"],
    ]

    return {
        "Assessment": assessment_rows,
        "FinalSpec": spec_rows,
        "FieldCatalog": field_rows,
        "PatternMatch": pattern_rows,
        "Prompt": prompt_rows,
        "Notes": notes_rows,
    }


def generate_excel_summary(output_path: Path, brief: dict[str, object], prompt: str) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    rows = build_excel_rows(brief, prompt)

    write_two_header_sheet(
        workbook,
        "Assessment",
        ["Category", "Item_Code", "Item_Name_CN", "Result", "Explanation"],
        ["类别", "项目编码", "项目中文名", "结果", "说明"],
        rows["Assessment"],
        left_align_columns={"Item_Name_CN", "Explanation"},
    )
    write_two_header_sheet(
        workbook,
        "FinalSpec",
        ["Spec_Type", "Spec_Name_CN", "Content", "Priority", "Source"],
        ["规格类型", "规格中文名", "内容", "优先级", "来源"],
        rows["FinalSpec"],
        left_align_columns={"Spec_Name_CN", "Content", "Source"},
    )
    write_two_header_sheet(
        workbook,
        "FieldCatalog",
        ["Field_Name", "Definition", "Source_Line", "Source_File"],
        ["字段名", "字段定义", "来源文本", "来源文件"],
        rows["FieldCatalog"],
        left_align_columns={"Definition", "Source_Line", "Source_File"},
    )
    write_two_header_sheet(
        workbook,
        "PatternMatch",
        ["Report_Code", "Script_Name", "Prefixes", "Purpose", "Dimensions", "Source_Fields"],
        ["报表编码", "脚本名", "识别前缀", "用途", "维度", "关键字段"],
        rows["PatternMatch"],
        left_align_columns={"Purpose", "Dimensions", "Source_Fields"},
    )
    write_two_header_sheet(
        workbook,
        "Prompt",
        ["Section", "Content"],
        ["章节", "内容"],
        rows["Prompt"],
        left_align_columns={"Content"},
    )
    write_two_header_sheet(
        workbook,
        "Notes",
        ["Field_EN", "字段中文", "说明"],
        ["英文字段", "字段中文", "说明"],
        rows["Notes"],
        left_align_columns={"说明"},
    )

    workbook.save(output_path)


def build_base_name(source_name: str, family: str) -> str:
    timestamp = dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    anchor = source_name if source_name and source_name != "inline_requirement" else family
    anchor = re.sub(r"[^\w\-]+", "_", anchor)
    return f"{anchor}_{timestamp}"


def main() -> None:
    args = parse_args()
    available_patterns = get_available_patterns()

    if not args.skip_auto_sync:
        sync_result = sync_logic_docs(args, available_patterns)
        sync_summary = sync_result["summary"]
        print(
            "自动同步完成："
            f"扫描 {sync_summary['scanned']} 个文档，"
            f"新增/更新已同步 {sync_summary['synced']} 个，"
            f"阻塞 {sync_summary['blocked']} 个，"
            f"失败 {sync_summary['failed']} 个，"
            f"未变化 {sync_summary['unchanged']} 个。"
        )
        print(f"同步记录：{sync_result['tracker_path']}")
        print(
            "启动脚本："
            + ", ".join(str(path) for path in sync_result.get("start_cmd_paths", ()))
        )

    if args.sync_only:
        return

    if not args.logic_doc and not (args.requirement or args.requirement_file):
        print("未提供 --logic-doc 或 --requirement / --requirement-file，本次仅执行自动同步。")
        return

    if args.logic_doc:
        pattern, doc_path, doc_body = resolve_logic_doc(args.logic_doc, available_patterns)
        target_script_path = Path(args.target_script) if args.target_script else (SCRIPT_DIR / pattern.script_name)
        prompt = build_codegen_prompt(
            pattern,
            doc_body,
            target_script_name=target_script_path.name,
            available_patterns=available_patterns,
        )
        if args.save_artifacts:
            prompt_path = save_codegen_prompt(Path(args.output_dir), pattern.code, prompt)
            print(f"已生成代码 prompt：{prompt_path}")

        if args.print_prompt or not args.call_api:
            print("\n===== PROMPT START =====")
            print(prompt)
            print("===== PROMPT END =====")

        if not args.call_api:
            return

        log_stage(f"[LOGIC_DOC] {pattern.code}: 开始定向生成 {target_script_path.name}。")
        result, response_text, validation_issues = generate_and_validate_script(
            prompt,
            args,
            target_script_name=target_script_path.name,
        )
        if args.write_script:
            if validation_issues:
                raise RuntimeError(
                    "生成脚本未通过本地校验，已停止写入：\n"
                    + format_validation_issues(validation_issues)
                )
            target_script_path.parent.mkdir(parents=True, exist_ok=True)
            write_text(target_script_path, response_text)
            update_tracker_entry(
                pattern,
                doc_path,
                target_script_path,
                status="synced",
                message="已通过定向生成更新脚本。",
                last_synced_at=now_text(),
            )
            print(f"已生成脚本：{target_script_path}")
        else:
            response_text_path = Path(args.output_dir) / f"{pattern.code}_generated_script.txt"
            response_text_path.parent.mkdir(parents=True, exist_ok=True)
            write_text(response_text_path, response_text)
            print(f"已生成脚本文本：{response_text_path}")
            if validation_issues:
                print("脚本校验未通过：")
                print(format_validation_issues(validation_issues))

        if args.print_response:
            print("\n===== RESPONSE START =====")
            print(response_text)
            print("===== RESPONSE END =====")
        return

    requirement, source_name = load_requirement(args)
    if not requirement.strip():
        raise ValueError("需求内容为空。")

    source_mode = detect_source_mode(requirement, args.source_mode)
    family = infer_family(requirement, args.family, available_patterns)
    context_profile = build_context_profile(requirement, source_mode)
    patterns = select_patterns(requirement, family, max(args.top_k, 1), available_patterns)
    brief = build_brief_payload(
        requirement=requirement.strip(),
        source_name=source_name,
        family=family,
        source_mode=source_mode,
        patterns=patterns,
        context_profile=context_profile,
    )
    prompt = build_prompt(brief)

    base_name = build_base_name(source_name, family)
    output_dir = Path(args.output_dir)
    if args.save_artifacts:
        brief_path, prompt_path = save_artifacts(output_dir, base_name, brief, prompt)
        print(f"已生成 brief：{brief_path}")
        print(f"已生成 prompt：{prompt_path}")

    if not args.skip_excel:
        excel_path = Path(args.excel_output) if args.excel_output else output_dir / f"{base_name}_analysis.xlsx"
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        generate_excel_summary(excel_path, brief, prompt)
        print(f"已生成结果 Excel：{excel_path}")

    if args.print_prompt:
        print("\n===== PROMPT START =====")
        print(prompt)
        print("===== PROMPT END =====")

    if not args.call_api:
        return

    result = call_model_api(prompt, args)
    response_text = str(result["response_text"])
    response_json_path = output_dir / f"{base_name}_response.json"
    response_text_path = output_dir / f"{base_name}_response.txt"
    response_json_path.write_text(
        json.dumps(result["raw_response"], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_text(response_text_path, response_text)

    print(f"已生成模型回复(JSON)：{response_json_path}")
    print(f"已生成模型回复(文本)：{response_text_path}")

    if args.print_response:
        print("\n===== RESPONSE START =====")
        print(response_text)
        print("===== RESPONSE END =====")


if __name__ == "__main__":
    main()
