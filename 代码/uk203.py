# -*- coding: utf-8 -*-
"""
根据 UK200 基础表生成 UK203_Seller转化漏斗概览.xlsx。

默认行为：
1. 自动读取脚本目录或其上级目录下最新的 UK200 开头基础表
2. 输出到 P:\\0_Report\\78_AI Report\\UKR_Report 下，文件名自动追加 -MMDDHHMM

示例：
python uk203.py
python uk203.py --input "C:\\path\\UK200_转化漏斗KPI-20260430.xlsx"
python uk203.py --output "C:\\path\\UK203_Seller转化漏斗概览.xlsx"
"""

from __future__ import annotations

import argparse
import datetime as dt
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from uk_export_utils import build_default_output_path
from uk_funnel_utils import (
    detect_data_sheet,
    find_latest_input,
    get_required_indexes,
    has_value,
    month_sort_key,
    normalize_text,
    parse_cell_date,
    safe_div,
    style_two_header_sheet,
)

SCRIPT_DIR = Path(__file__).resolve().parent
BASE_DIR = SCRIPT_DIR.parent
EXPORT_DIR = Path(r"P:\0_Report\78_AI Report\UKR_Report")
DEFAULT_OUTPUT_NAME = "UK203_Seller转化漏斗概览.xlsx"
MONTH_START_DATE = dt.date(2025, 1, 1)
SELLER_START_DATE = dt.date(2020, 1, 1)

CANCEL_STATUS = {
    "SellerCancel",
    "CSCancel",
    "Long time-Cancel",
    "WootCancel",
}

DEAL_STATUS = {
    "End",
    "WaitingRPO",
    "WaitingPayment",
    "WaitingDeal",
    "OnDeal",
    "WaitingPO",
}

IN_PROGRESS_STATUS = {
    "Hold",
    "New",
    "SubmitRBS",
}

STATUS_PRIORITY = {
    "Other": 0,
    "InProgress": 1,
    "Deal": 2,
    "Cancel": 3,
}

MONTH_BASE_COLUMNS_EN = [
    "Seller_ID",
    "Submit_Month",
    "Total_QPO_Count",
    "Deal_QPO_Count",
    "Cancel_QPO_Count",
    "InProgress_QPO_Count",
    "Deal_QPO_Rate",
    "Cancel_QPO_Rate",
    "InProgress_QPO_Rate",
]

MONTH_BASE_COLUMNS_CN = [
    "卖家ID",
    "提报月份",
    "当月提报QPO总数",
    "当月成交QPO数量",
    "当月取消QPO数量",
    "当月跟进中QPO数量",
    "当月成交占比",
    "当月取消占比",
    "当月跟进中占比",
]

SELLER_BASE_COLUMNS_EN = [
    "Seller_ID",
    "Total_QPO_Count",
    "Deal_QPO_Count",
    "Cancel_QPO_Count",
    "InProgress_QPO_Count",
    "Deal_QPO_Rate",
    "Cancel_QPO_Rate",
    "InProgress_QPO_Rate",
    "First_Submit_Date",
    "Last_Submit_Date",
]

SELLER_BASE_COLUMNS_CN = [
    "卖家ID",
    "2020至今提报总数",
    "2020至今成交QPO数量",
    "2020至今取消QPO数量",
    "2020至今跟进中QPO数量",
    "2020至今成交占比",
    "2020至今取消占比",
    "2020至今跟进中占比",
    "最早提报日期",
    "最近提报日期",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="根据 UK200 基础表生成 UK203_Seller转化漏斗概览。"
    )
    parser.add_argument(
        "--input",
        dest="input_path",
        help="UK200 基础表路径；不传时自动读取最新的 UK200 开头基础表",
    )
    parser.add_argument(
        "--output",
        dest="output_path",
        help=f"输出文件路径；不传时默认写入 {EXPORT_DIR}，文件名自动追加 -MMDDHHMM",
    )
    return parser.parse_args()


def has_chinese(text: object) -> bool:
    content = str(text or "")
    return any("\u4e00" <= char <= "\u9fff" for char in content)


def resolve_input_path(input_arg: str | None) -> Path:
    if input_arg:
        return Path(input_arg)

    tried_dirs: list[Path] = []
    for candidate_dir in (SCRIPT_DIR, BASE_DIR):
        if candidate_dir in tried_dirs:
            continue
        tried_dirs.append(candidate_dir)
        try:
            return find_latest_input(candidate_dir)
        except FileNotFoundError:
            continue

    tried_text = " / ".join(str(path) for path in tried_dirs)
    raise FileNotFoundError(
        f"未找到 UK200 基础表，请确认以下目录存在 UK200 开头的 Excel 文件：{tried_text}"
    )


def resolve_column_indexes(header_row: list[str]) -> dict[str, int]:
    required_single = {
        "Submission_Date": "Submission_Date",
        "QPO_ID": "QPO_ID",
        "QPO_Status": "QPO_Status",
        "Seller_ID": "Seller_ID",
    }

    duplicate_requirements = {
        "WPO_Current_Node": ("WPO_Current_Node", 1),
        "WPO_Current_Node_CN": ("WPO_Current_Node", 2),
    }

    try:
        indexes = get_required_indexes(
            header_row,
            required_single=required_single,
            duplicate_requirements=duplicate_requirements,
        )
        if "WPO_Current_Node" in indexes and "WPO_Current_Node_CN" in indexes:
            return indexes
    except Exception:
        pass

    matches: dict[str, list[int]] = defaultdict(list)
    for index, header in enumerate(header_row):
        matches[normalize_text(header)].append(index)

    indexes: dict[str, int] = {}
    missing: list[str] = []

    for key, header_name in required_single.items():
        if matches.get(header_name):
            indexes[key] = matches[header_name][0]
        else:
            missing.append(header_name)

    node_headers = matches.get("WPO_Current_Node", [])
    if len(node_headers) >= 2:
        indexes["WPO_Current_Node"] = node_headers[0]
        indexes["WPO_Current_Node_CN"] = node_headers[1]
    else:
        missing.append("WPO_Current_Node(需存在代码列和中文列各1个)")

    if missing:
        raise ValueError(f"源表缺少必要字段：{missing}")

    return indexes


def map_status_category(status: str) -> str:
    if status in CANCEL_STATUS:
        return "Cancel"
    if status in DEAL_STATUS:
        return "Deal"
    if status in IN_PROGRESS_STATUS:
        return "InProgress"
    return "Other"


def normalize_node(node: str) -> str:
    text = normalize_text(node).upper()
    return text if re.fullmatch(r"N\d{3}", text) else "Unknown"


def node_rank(node: str) -> tuple[int, int]:
    matched = re.fullmatch(r"N(\d{3})", normalize_text(node).upper())
    if not matched:
        return (0, -1)
    return (1, int(matched.group(1)))


def node_sort_key(node: str) -> tuple[int, int, str]:
    if node == "Unknown":
        return (0, -1, node)
    matched = re.fullmatch(r"N(\d{3})", normalize_text(node).upper())
    if matched:
        return (1, int(matched.group(1)), node)
    return (2, 999999, node)


def render_distribution(counter: Counter[str]) -> str:
    if not counter:
        return "无"
    return "; ".join(f"{key}:{counter[key]}" for key in sorted(counter))


def render_value_or_none(values: list[str]) -> str:
    return "无" if not values else ", ".join(values)


def build_node_columns(
    nodes_sorted: list[str],
    node_cn_map: dict[str, str],
    count_suffix: str,
    rate_suffix: str,
) -> tuple[list[str], list[str]]:
    columns_en: list[str] = []
    columns_cn: list[str] = []
    for node in nodes_sorted:
        node_name_cn = node_cn_map.get(node, node)
        columns_en.extend([f"{node}_Cancel_QPO_Count", f"{node}_Cancel_QPO_Rate"])
        columns_cn.extend([f"{node_name_cn}{count_suffix}", f"{node_name_cn}{rate_suffix}"])
    return columns_en, columns_cn


def build_summary(input_path: Path) -> tuple[dict[str, object], dict[str, object]]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        sheet = detect_data_sheet(workbook)

        header_rows = list(sheet.iter_rows(min_row=1, max_row=2, values_only=True))
        if not header_rows:
            raise ValueError("源表缺少表头，无法识别字段。")

        header_row_1 = [normalize_text(value) for value in header_rows[0]]
        header_row_2 = (
            [normalize_text(value) for value in header_rows[1]]
            if len(header_rows) >= 2
            else []
        )

        column_indexes = resolve_column_indexes(header_row_1)
        data_start_row = 3 if any(has_chinese(value) for value in header_row_2) else 2

        groups: dict[tuple[str, str], dict[str, object]] = {}
        raw_valid_rows = 0
        raw_other_status: Counter[str] = Counter()
        missing_node_cn: set[str] = set()
        node_cn_map: dict[str, str] = {}
        row_sequence = 0

        for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
            submission_date = parse_cell_date(row[column_indexes["Submission_Date"]])
            if submission_date is None:
                continue

            raw_valid_rows += 1

            seller_id = normalize_text(row[column_indexes["Seller_ID"]])
            qpo_id = normalize_text(row[column_indexes["QPO_ID"]])
            if not seller_id or not qpo_id:
                continue

            status = normalize_text(row[column_indexes["QPO_Status"]])
            category = map_status_category(status)

            raw_node = normalize_text(row[column_indexes["WPO_Current_Node"]])
            raw_node_cn_value = row[column_indexes["WPO_Current_Node_CN"]]
            raw_node_cn = normalize_text(raw_node_cn_value)
            node = normalize_node(raw_node)

            if category == "Other":
                raw_other_status[status or "<blank>"] += 1

            if node != "Unknown":
                if has_value(raw_node_cn_value):
                    node_cn_map.setdefault(node, raw_node_cn)
                else:
                    missing_node_cn.add(node)

            row_sequence += 1
            record = groups.setdefault(
                (seller_id, qpo_id),
                {
                    "first_submit_date": submission_date,
                    "rows": [],
                },
            )

            if submission_date < record["first_submit_date"]:
                record["first_submit_date"] = submission_date

            record["rows"].append(
                {
                    "seq": row_sequence,
                    "category": category,
                    "node": node,
                    "node_rank": node_rank(raw_node),
                }
            )

        dedup_records: list[dict[str, object]] = []
        nodes_sorted_set: set[str] = set()

        for (seller_id, qpo_id), record in groups.items():
            rows = record["rows"]
            final_category = max(
                rows,
                key=lambda item: (STATUS_PRIORITY[item["category"]], item["seq"]),
            )["category"]
            final_rows = [row for row in rows if row["category"] == final_category]
            final_node_row = max(
                final_rows,
                key=lambda item: (item["node_rank"], item["seq"]),
            )
            final_node = final_node_row["node"]

            if final_category == "Cancel" and final_node != "Unknown":
                nodes_sorted_set.add(final_node)

            dedup_records.append(
                {
                    "seller_id": seller_id,
                    "qpo_id": qpo_id,
                    "first_submit_date": record["first_submit_date"],
                    "final_category": final_category,
                    "final_node": final_node,
                }
            )

        nodes_sorted = sorted(nodes_sorted_set, key=node_sort_key)

        month_groups = defaultdict(
            lambda: {
                "Total": 0,
                "Deal": 0,
                "Cancel": 0,
                "InProgress": 0,
                "nodes": Counter(),
            }
        )
        seller_groups = defaultdict(
            lambda: {
                "Total": 0,
                "Deal": 0,
                "Cancel": 0,
                "InProgress": 0,
                "nodes": Counter(),
                "First_Submit_Date": None,
                "Last_Submit_Date": None,
            }
        )

        date_min: dt.date | None = None
        date_max: dt.date | None = None
        qpo_count_2025_plus = 0
        qpo_count_2020_plus = 0

        for record in dedup_records:
            first_submit_date = record["first_submit_date"]
            seller_id = record["seller_id"]
            final_category = record["final_category"]
            final_node = record["final_node"]
            submit_month = first_submit_date.strftime("%Y-%m")

            if date_min is None or first_submit_date < date_min:
                date_min = first_submit_date
            if date_max is None or first_submit_date > date_max:
                date_max = first_submit_date

            if first_submit_date >= MONTH_START_DATE:
                qpo_count_2025_plus += 1
                month_value = month_groups[(seller_id, submit_month)]
                month_value["Total"] += 1
                month_value[final_category] += 1
                if final_category == "Cancel":
                    month_value["nodes"][final_node] += 1

            if first_submit_date >= SELLER_START_DATE:
                qpo_count_2020_plus += 1
                seller_value = seller_groups[seller_id]
                seller_value["Total"] += 1
                seller_value[final_category] += 1
                if final_category == "Cancel":
                    seller_value["nodes"][final_node] += 1

                if (
                    seller_value["First_Submit_Date"] is None
                    or first_submit_date < seller_value["First_Submit_Date"]
                ):
                    seller_value["First_Submit_Date"] = first_submit_date
                if (
                    seller_value["Last_Submit_Date"] is None
                    or first_submit_date > seller_value["Last_Submit_Date"]
                ):
                    seller_value["Last_Submit_Date"] = first_submit_date

        summary = {
            "dedup_records": dedup_records,
            "nodes_sorted": nodes_sorted,
            "node_cn_map": node_cn_map,
            "month_groups": month_groups,
            "seller_groups": seller_groups,
        }
        metadata = {
            "input_path": str(input_path),
            "raw_valid_rows": raw_valid_rows,
            "dedup_qpo_count": len(dedup_records),
            "date_min": date_min,
            "date_max": date_max,
            "qpo_count_2025_plus": qpo_count_2025_plus,
            "qpo_count_2020_plus": qpo_count_2020_plus,
            "raw_other_status": raw_other_status,
            "missing_node_cn": sorted(missing_node_cn, key=node_sort_key),
        }
        return summary, metadata
    finally:
        workbook.close()


def build_month_rows(summary: dict[str, object]) -> list[dict[str, object]]:
    nodes_sorted = summary["nodes_sorted"]
    month_groups = summary["month_groups"]

    rows: list[dict[str, object]] = []
    for (seller_id, submit_month), value in sorted(
        month_groups.items(),
        key=lambda item: (item[0][0], month_sort_key(item[0][1])),
    ):
        total = value["Total"]
        deal = value["Deal"]
        cancel = value["Cancel"]
        in_progress = value["InProgress"]

        row = {
            "Seller_ID": seller_id,
            "Submit_Month": submit_month,
            "Total_QPO_Count": total,
            "Deal_QPO_Count": deal,
            "Cancel_QPO_Count": cancel,
            "InProgress_QPO_Count": in_progress,
            "Deal_QPO_Rate": safe_div(deal, total),
            "Cancel_QPO_Rate": safe_div(cancel, total),
            "InProgress_QPO_Rate": safe_div(in_progress, total),
        }

        for node in nodes_sorted:
            node_count = value["nodes"].get(node, 0)
            row[f"{node}_Cancel_QPO_Count"] = node_count
            row[f"{node}_Cancel_QPO_Rate"] = safe_div(node_count, total)

        rows.append(row)

    return rows


def build_seller_rows(summary: dict[str, object]) -> list[dict[str, object]]:
    nodes_sorted = summary["nodes_sorted"]
    seller_groups = summary["seller_groups"]

    rows: list[dict[str, object]] = []
    for seller_id, value in sorted(seller_groups.items()):
        total = value["Total"]
        deal = value["Deal"]
        cancel = value["Cancel"]
        in_progress = value["InProgress"]

        row = {
            "Seller_ID": seller_id,
            "Total_QPO_Count": total,
            "Deal_QPO_Count": deal,
            "Cancel_QPO_Count": cancel,
            "InProgress_QPO_Count": in_progress,
            "Deal_QPO_Rate": safe_div(deal, total),
            "Cancel_QPO_Rate": safe_div(cancel, total),
            "InProgress_QPO_Rate": safe_div(in_progress, total),
            "First_Submit_Date": value["First_Submit_Date"],
            "Last_Submit_Date": value["Last_Submit_Date"],
        }

        for node in nodes_sorted:
            node_count = value["nodes"].get(node, 0)
            row[f"{node}_Cancel_QPO_Count"] = node_count
            row[f"{node}_Cancel_QPO_Rate"] = safe_div(node_count, total)

        rows.append(row)

    return rows


def apply_sheet_style(
    sheet,
    columns_en: list[str],
    columns_cn: list[str],
    date_columns: set[str] | None = None,
) -> None:
    date_columns = date_columns or set()

    try:
        style_two_header_sheet(
            sheet,
            columns_en,
            columns_cn,
            date_columns=date_columns,
        )
    except Exception:
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_num in (1, 2):
            for col_num in range(1, len(columns_en) + 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

        preview_rows = min(sheet.max_row, 30)
        for idx, column in enumerate(columns_en, start=1):
            sample_lengths = [len(str(column)), len(str(columns_cn[idx - 1]))]
            for row_num in range(3, preview_rows + 1):
                value = sheet.cell(row=row_num, column=idx).value
                if value is not None:
                    sample_lengths.append(len(str(value)))
            sheet.column_dimensions[get_column_letter(idx)].width = min(max(sample_lengths) + 2, 34)

        sheet.freeze_panes = "A3"

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    count_columns = {
        "Total_QPO_Count",
        "Deal_QPO_Count",
        "Cancel_QPO_Count",
        "InProgress_QPO_Count",
    }
    rate_columns = {
        "Deal_QPO_Rate",
        "Cancel_QPO_Rate",
        "InProgress_QPO_Rate",
    }

    for column in columns_en:
        if column.endswith("_Cancel_QPO_Count"):
            count_columns.add(column)
        elif column.endswith("_Cancel_QPO_Rate"):
            rate_columns.add(column)

    column_index = {column: idx + 1 for idx, column in enumerate(columns_en)}

    for row_num in range(3, sheet.max_row + 1):
        for column in columns_en:
            cell = sheet.cell(row=row_num, column=column_index[column])
            cell.alignment = center
            if column in count_columns:
                cell.number_format = "0"
            elif column in rate_columns:
                cell.number_format = "0.00%"
            elif column in date_columns and cell.value is not None:
                cell.number_format = "yyyy-mm-dd"

    if sheet.freeze_panes is None:
        sheet.freeze_panes = "A3"


def write_month_sheet(
    workbook: Workbook,
    rows: list[dict[str, object]],
    nodes_sorted: list[str],
    node_cn_map: dict[str, str],
) -> None:
    node_columns_en, node_columns_cn = build_node_columns(
        nodes_sorted,
        node_cn_map,
        "取消QPO数量",
        "取消占当月提报比例",
    )
    columns_en = MONTH_BASE_COLUMNS_EN + node_columns_en
    columns_cn = MONTH_BASE_COLUMNS_CN + node_columns_cn

    sheet = workbook.active
    sheet.title = "UK203_Seller_Month"
    sheet.append(columns_en)
    sheet.append(columns_cn)
    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    apply_sheet_style(sheet, columns_en, columns_cn)


def write_seller_sheet(
    workbook: Workbook,
    rows: list[dict[str, object]],
    nodes_sorted: list[str],
    node_cn_map: dict[str, str],
) -> None:
    node_columns_en, node_columns_cn = build_node_columns(
        nodes_sorted,
        node_cn_map,
        "节点取消QPO数量",
        "节点取消占全部提报比例",
    )
    columns_en = SELLER_BASE_COLUMNS_EN + node_columns_en
    columns_cn = SELLER_BASE_COLUMNS_CN + node_columns_cn

    sheet = workbook.create_sheet("UK203_Seller")
    sheet.append(columns_en)
    sheet.append(columns_cn)
    for row in rows:
        sheet.append([row.get(column) for column in columns_en])

    apply_sheet_style(sheet, columns_en, columns_cn, {"First_Submit_Date", "Last_Submit_Date"})


def build_notes_rows(
    nodes_sorted: list[str],
    node_cn_map: dict[str, str],
    metadata: dict[str, object],
) -> list[list[object]]:
    rows: list[list[object]] = []

    rows.extend(
        [
            [
                "Time_Range_Month",
                "时间范围（月度）",
                "UK203_Seller_Month",
                "Meta",
                "统计时间范围",
                f"{MONTH_START_DATE.isoformat()} 至 {metadata['date_max'].isoformat()}（按排重后QPO最早提报日期）"
                if metadata["date_max"]
                else MONTH_START_DATE.isoformat(),
                None,
            ],
            [
                "Time_Range_Seller",
                "时间范围（卖家累计）",
                "UK203_Seller",
                "Meta",
                "统计时间范围",
                f"{SELLER_START_DATE.isoformat()} 至 {metadata['date_max'].isoformat()}（按排重后QPO最早提报日期）"
                if metadata["date_max"]
                else SELLER_START_DATE.isoformat(),
                None,
            ],
            [
                "QPO_Dedup_Rule",
                "QPO排重规则",
                "Both",
                "Meta",
                "排除过程记录重复",
                "按 Seller_ID + QPO_ID 排重；Submission_Date 取最早；状态优先级：Cancel > Deal > InProgress > Other；节点取最终状态对应节点（冲突优先Cancel）",
                None,
            ],
            [
                "Status_Mapping",
                "QPO_Status分类映射",
                "Both",
                "Meta",
                "将QPO状态映射到漏斗阶段",
                "Cancel: SellerCancel/CSCancel/Long time-Cancel/WootCancel; Deal: End/WaitingRPO/WaitingPayment/WaitingDeal/OnDeal/WaitingPO; InProgress: Hold/New/SubmitRBS; 其他为Other(仅自检)",
                None,
            ],
            [
                "Seller_Filter_Rule",
                "卖家过滤规则",
                "Both",
                "Meta",
                "Seller视图过滤规则",
                "Seller_ID 为空的记录不参与 Seller 视图与月度 Seller 汇总",
                None,
            ],
        ]
    )

    rows.extend(
        [
            [
                "Seller_ID",
                "卖家ID",
                "UK203_Seller_Month",
                "Core Metrics",
                "卖家编号（内）",
                "来自 UK200 Seller_ID",
                "文本/ID",
            ],
            [
                "Submit_Month",
                "提报月份",
                "UK203_Seller_Month",
                "Core Metrics",
                "提报月份(YYYY-MM)",
                "Submission_Date 截取到 YYYY-MM（基于排重后QPO最早提报日期）",
                "文本",
            ],
            [
                "Total_QPO_Count",
                "当月提报QPO总数",
                "UK203_Seller_Month",
                "Core Metrics",
                "当月提报QPO总数(排重后)",
                "当月 Seller_ID+QPO_ID 去重后计数",
                "整数",
            ],
            [
                "Deal_QPO_Count",
                "当月成交QPO数量",
                "UK203_Seller_Month",
                "Core Metrics",
                "当月成交QPO数量",
                "当月去重QPO中 QPO_Status∈Deal 状态计数",
                "整数",
            ],
            [
                "Cancel_QPO_Count",
                "当月取消QPO数量",
                "UK203_Seller_Month",
                "Core Metrics",
                "当月取消QPO数量",
                "当月去重QPO中 QPO_Status∈Cancel 状态计数",
                "整数",
            ],
            [
                "InProgress_QPO_Count",
                "当月跟进中QPO数量",
                "UK203_Seller_Month",
                "Core Metrics",
                "当月跟进中QPO数量",
                "当月去重QPO中 QPO_Status∈InProgress 状态计数",
                "整数",
            ],
            [
                "Deal_QPO_Rate",
                "当月成交占比",
                "UK203_Seller_Month",
                "Core Metrics",
                "当月成交占当月提报比例",
                "Deal_QPO_Count / Total_QPO_Count",
                "百分比，保留2位小数",
            ],
            [
                "Cancel_QPO_Rate",
                "当月取消占比",
                "UK203_Seller_Month",
                "Core Metrics",
                "当月取消占当月提报比例",
                "Cancel_QPO_Count / Total_QPO_Count",
                "百分比，保留2位小数",
            ],
            [
                "InProgress_QPO_Rate",
                "当月跟进中占比",
                "UK203_Seller_Month",
                "Core Metrics",
                "当月跟进中占当月提报比例",
                "InProgress_QPO_Count / Total_QPO_Count",
                "百分比，保留2位小数",
            ],
        ]
    )

    for node in nodes_sorted:
        node_name_cn = node_cn_map.get(node, node)
        rows.append(
            [
                f"{node}_Cancel_QPO_Count",
                f"{node_name_cn}取消QPO数量",
                "UK203_Seller_Month",
                "Cancel Node",
                f"当月在节点{node}取消的QPO数量",
                f"当月去重QPO中 cls=Cancel 且 WPO_Current_Node={node} 的计数",
                "整数",
            ]
        )
        rows.append(
            [
                f"{node}_Cancel_QPO_Rate",
                f"{node_name_cn}取消占当月提报比例",
                "UK203_Seller_Month",
                "Cancel Node",
                f"当月在节点{node}取消占当月提报比例",
                f"{node}_Cancel_QPO_Count / Total_QPO_Count（分母为当月提报总QPO）",
                "百分比，保留2位小数",
            ]
        )

    rows.extend(
        [
            [
                "Seller_ID",
                "卖家ID",
                "UK203_Seller",
                "Core Metrics",
                "卖家编号（内）",
                "来自 UK200 Seller_ID",
                "文本/ID",
            ],
            [
                "Total_QPO_Count",
                "2020至今提报总数",
                "UK203_Seller",
                "Core Metrics",
                "2020至今提报QPO总数(排重后)",
                "2020-01-01起 Seller_ID+QPO_ID 去重后计数",
                "整数",
            ],
            [
                "Deal_QPO_Count",
                "2020至今成交QPO数量",
                "UK203_Seller",
                "Core Metrics",
                "2020至今成交QPO数量",
                "去重QPO中 QPO_Status∈Deal 状态计数",
                "整数",
            ],
            [
                "Cancel_QPO_Count",
                "2020至今取消QPO数量",
                "UK203_Seller",
                "Core Metrics",
                "2020至今取消QPO数量",
                "去重QPO中 QPO_Status∈Cancel 状态计数",
                "整数",
            ],
            [
                "InProgress_QPO_Count",
                "2020至今跟进中QPO数量",
                "UK203_Seller",
                "Core Metrics",
                "2020至今跟进中QPO数量",
                "去重QPO中 QPO_Status∈InProgress 状态计数",
                "整数",
            ],
            [
                "Deal_QPO_Rate",
                "2020至今成交占比",
                "UK203_Seller",
                "Core Metrics",
                "2020至今成交占全部提报比例",
                "Deal_QPO_Count / Total_QPO_Count",
                "百分比，保留2位小数",
            ],
            [
                "Cancel_QPO_Rate",
                "2020至今取消占比",
                "UK203_Seller",
                "Core Metrics",
                "2020至今取消占全部提报比例",
                "Cancel_QPO_Count / Total_QPO_Count",
                "百分比，保留2位小数",
            ],
            [
                "InProgress_QPO_Rate",
                "2020至今跟进中占比",
                "UK203_Seller",
                "Core Metrics",
                "2020至今跟进中占全部提报比例",
                "InProgress_QPO_Count / Total_QPO_Count",
                "百分比，保留2位小数",
            ],
            [
                "First_Submit_Date",
                "最早提报日期",
                "UK203_Seller",
                "Core Metrics",
                "该卖家最早提报日期（基于去重QPO）",
                "min(Submission_Date)（按去重QPO）",
                "日期 yyyy-mm-dd",
            ],
            [
                "Last_Submit_Date",
                "最近提报日期",
                "UK203_Seller",
                "Core Metrics",
                "该卖家最近提报日期（基于去重QPO）",
                "max(Submission_Date)（按去重QPO）",
                "日期 yyyy-mm-dd",
            ],
        ]
    )

    for node in nodes_sorted:
        node_name_cn = node_cn_map.get(node, node)
        rows.append(
            [
                f"{node}_Cancel_QPO_Count",
                f"{node_name_cn}节点取消QPO数量",
                "UK203_Seller",
                "Cancel Node",
                f"2020至今在节点{node}取消的QPO数量",
                f"去重QPO中 cls=Cancel 且 WPO_Current_Node={node} 的计数",
                "整数",
            ]
        )
        rows.append(
            [
                f"{node}_Cancel_QPO_Rate",
                f"{node_name_cn}节点取消占全部提报比例",
                "UK203_Seller",
                "Cancel Node",
                f"2020至今在节点{node}取消占全部提报比例",
                f"{node}_Cancel_QPO_Count / Total_QPO_Count（分母为卖家全部提报总QPO）",
                "百分比，保留2位小数",
            ]
        )

    rows.extend(
        [
            [
                "Self_Check",
                "自检报告",
                "Notes",
                "QA",
                "输入/输出一致性检查",
                "见下方自检明细行",
                None,
            ],
            [
                "Input_File",
                "输入文件",
                "Notes",
                "QA",
                "本次使用的输入文件",
                str(metadata.get("input_path") or ""),
                None,
            ],
            [
                "UK200_Raw_Rows",
                "UK200原始有效行数",
                "Notes",
                "QA",
                "Submission_Date 可识别的原始行数",
                str(metadata["raw_valid_rows"]),
                None,
            ],
            [
                "Dedup_QPO_Count",
                "排重后QPO数(Seller+QPO)",
                "Notes",
                "QA",
                "排重后QPO数(Seller+QPO)",
                str(metadata["dedup_qpo_count"]),
                None,
            ],
            [
                "Dedup_Date_Min",
                "排重后最早提报日期",
                "Notes",
                "QA",
                "排重后最早提报日期",
                metadata["date_min"].isoformat() if metadata["date_min"] else "无",
                None,
            ],
            [
                "Dedup_Date_Max",
                "排重后最晚提报日期",
                "Notes",
                "QA",
                "排重后最晚提报日期",
                metadata["date_max"].isoformat() if metadata["date_max"] else "无",
                None,
            ],
            [
                "QPO_Count_2025_plus",
                "2025-01-01及以后QPO数",
                "Notes",
                "QA",
                "月度范围纳入QPO数",
                str(metadata["qpo_count_2025_plus"]),
                None,
            ],
            [
                "QPO_Count_2020_plus",
                "2020-01-01及以后QPO数",
                "Notes",
                "QA",
                "卖家累计范围纳入QPO数",
                str(metadata["qpo_count_2020_plus"]),
                None,
            ],
            [
                "Other_Status_Distribution",
                "Other状态分布",
                "Notes",
                "QA",
                "未映射状态枚举与数量（仅自检）",
                render_distribution(metadata["raw_other_status"]),
                None,
            ],
            [
                "Missing_Node_CN",
                "节点中文缺失列表",
                "Notes",
                "QA",
                "WPO_Current_Node 中文列为空的节点",
                render_value_or_none(metadata["missing_node_cn"]),
                None,
            ],
        ]
    )

    return rows


def write_notes_sheet(
    workbook: Workbook,
    nodes_sorted: list[str],
    node_cn_map: dict[str, str],
    metadata: dict[str, object],
) -> None:
    sheet = workbook.create_sheet("Notes")
    headers_en = [
        "Field (EN)",
        "Field (CN)",
        "Sheet",
        "Field Group",
        "Business Meaning",
        "Calculation / Rule",
        "Notes",
    ]
    headers_cn = [
        "字段英文",
        "字段中文",
        "所属Sheet",
        "字段分组",
        "业务含义",
        "计算方式 / 判断标准",
        "补充说明",
    ]

    sheet.append(headers_en)
    sheet.append(headers_cn)
    for row in build_notes_rows(nodes_sorted, node_cn_map, metadata):
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_num in (1, 2):
        for col_num in range(1, len(headers_en) + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center

    widths = [22, 22, 20, 16, 28, 72, 16]
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = width

    sheet.freeze_panes = "A3"


def write_workbook(
    output_path: Path,
    month_rows: list[dict[str, object]],
    seller_rows: list[dict[str, object]],
    nodes_sorted: list[str],
    node_cn_map: dict[str, str],
    metadata: dict[str, object],
) -> None:
    workbook = Workbook()
    try:
        write_month_sheet(workbook, month_rows, nodes_sorted, node_cn_map)
        write_seller_sheet(workbook, seller_rows, nodes_sorted, node_cn_map)
        write_notes_sheet(workbook, nodes_sorted, node_cn_map, metadata)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
    finally:
        workbook.close()


def generate_report(input_path: Path, output_path: Path) -> None:
    summary, metadata = build_summary(input_path)
    month_rows = build_month_rows(summary)
    seller_rows = build_seller_rows(summary)

    write_workbook(
        output_path=output_path,
        month_rows=month_rows,
        seller_rows=seller_rows,
        nodes_sorted=summary["nodes_sorted"],
        node_cn_map=summary["node_cn_map"],
        metadata=metadata,
    )


def main() -> None:
    args = parse_args()
    input_path = resolve_input_path(args.input_path)
    output_path = (
        Path(args.output_path)
        if args.output_path
        else build_default_output_path(EXPORT_DIR, DEFAULT_OUTPUT_NAME, BASE_DIR)
    )

    generate_report(input_path=input_path, output_path=output_path)
    print("生成完成：")
    print(output_path)


if __name__ == "__main__":
    main()
